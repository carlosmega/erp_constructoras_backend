"""
Management command to import SYM KABAT project data from Excel file 002.

Imports:
  - Account (SYM KABAT client + 18 suppliers)
  - ConstructionProject
  - ProjectZone (TAM, MAN)
  - ProjectSupplier (18 entries)
  - CostCategory (P1-P10, C1-C8) via seed_default_categories
  - ImputationCode (parsed from Imput sheet codes)
  - ImputationPeriod (JUN 2025 Q1 → FEB 2026 Q1, fortnightly)
  - ProjectExpense + ExpenseLine (1,876 rows from Imput sheet)
  - ClientEstimate (rows from Control Certificacion sheet)

Usage:
    python manage.py import_sym_kabat
    python manage.py import_sym_kabat --file /path/to/file.xlsx
    python manage.py import_sym_kabat --dry-run
"""

import io
import re
import calendar
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

# ---------------------------------------------------------------------------
# Excel file helper: patch corrupt definedNames XML before loading
# ---------------------------------------------------------------------------

def _load_workbook_patched(path: str):
    """Load xlsx patching the invalid <definedNames> block (contains #N/A, #REF!)."""
    with open(path, 'rb') as fh:
        raw = fh.read()

    buf = io.BytesIO(raw)
    with zipfile.ZipFile(buf, 'r') as zin:
        names = zin.namelist()
        files = {name: zin.read(name) for name in names}

    wb_xml = files.get('xl/workbook.xml', b'')
    patched = re.sub(
        rb'<definedNames>.*?</definedNames>',
        b'',
        wb_xml,
        flags=re.DOTALL,
    )
    files['xl/workbook.xml'] = patched

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    out.seek(0)

    return openpyxl.load_workbook(out, data_only=True, read_only=True)


# ---------------------------------------------------------------------------
# Mapping helpers
# ---------------------------------------------------------------------------

MONTH_LABELS_ES = {
    'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4,
    'MAY': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8,
    'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12,
}

MONTH_LABELS_INT = {v: k for k, v in MONTH_LABELS_ES.items()}

DOCUMENT_TYPE_MAP = {
    'factura':         0,   # INVOICE
    'nota de crédito': 1,   # CREDIT_NOTE
    'nota de credito': 1,
    'gasto sin factura': 2, # NO_INVOICE_EXPENSE
    'gastos sin factura': 2,
    'caja chica':      2,
    'recibo':          2,
    'comprobante de pago': 2,
    'prefactura':      0,   # treat as INVOICE
    'comprobante de ingresos': 2,
    'nómina':          3,   # PAYROLL
    'nomina':          3,
    'sua':             3,
    'provisión':       4,   # PROVISION
    'provision':       4,
    'cancelación':     0,   # CANCELED expense
    'cancelacion':     0,
    'deductiva':       1,   # CREDIT_NOTE
}

PAYMENT_METHOD_MAP = {
    '01': 2,  # Efectivo → CASH
    'efectivo': 2,
    '03': 1,  # Transferencia → BANK_TRANSFER
    'transferencia': 1,
    '04': 0,  # TDC → CREDIT_CARD
    'tdc': 0,
    '06': 99, # Dinero Electrónico → OTHER
    'dinero electrónico': 99,
    'dinero electronico': 99,
    '15': 99, # Condonación → OTHER
    '28': 4,  # TDD → DEBIT_CARD
    'tdd': 4,
    '30': 1,  # Aplicación Anticipos → BANK_TRANSFER
    '99': 99, # Por Definir → OTHER
    'ppd': 1, # Pago en parcialidades → BANK_TRANSFER
    'ppd-pago en parcialidades o diferido': 1,
    '27-a satisfacción del acreedor': 99,
    '31-intermediario pagos': 1,
}

PAYMENT_STATUS_MAP = {
    'pendiente':    0,  # PENDING
    'pagado':       1,  # PAID
    'atrasado':     3,  # OVERDUE
    'sobrepagado':  2,  # PARTIALLY_PAID
}

ESTIMATE_TYPE_MAP = {
    'estimacion': 0,   # ESTIMATE
    'estimación': 0,
    'anticipo':   1,   # OTHER
    'amort. anticipo': 1,
    'amortización anticipo': 1,
    'deductiva':  1,
    'retención 5%': 1,
    'retencion 5%': 1,
}

SUPPLIERS_DATA = [
    ('MAS991101DH9',  'MASYFERR'),
    ('ANA050518RL1',  'AEROENLACES NACIONALES SA DE CV'),
    ('CFC110121742',  'COMERCIALIZADORA FARMACÉUTICA DE CHIAPAS SA DE CV'),
    ('CLG130128KA0',  'E12196 COMBUSTIBLES Y LUBRICANTES GHN SA DE CV'),
    ('DLI931201MI9',  'DISTRIBUIDORA LIVERPOOL SA DE CV'),
    ('FCU220209NQ1',  'FERRETERÍA CUERVO SA DE CV'),
    ('CAHE700209ATA', 'ELVIA CABRERA HUERTA'),
    ('GUCA670908V23', 'AUGUSTO GUERRERO CRUZ'),
    ('HDM001017AS1',  'HOME DEPOT MEXICO SA DE CV'),
    ('LOZ170722KGA',  'EL LINDERO DE OZULUAMA'),
    ('MASR990207BB1', 'RICARDO MARTINEZ SALAZAR'),
    ('MAS0810247C0',  'POPAT BAZAR SA DE CV'),
    ('PFL981207AX5',  'PERFILES DE FIERRO Y LAMINA ATIYE SA DE CV'),
    ('POAJ841027D66', 'JONATHAN PORTILLO ALEJO'),
    ('ROBL4512274I1', 'ESTACION DE SERVICIO No. 6580 JOSE LUIS RODRIGUEZ'),
    ('ROGM701116LV6', 'MIGUEL ANGEL ROCHA GARCIA'),
    ('SSA010807N4A',  'SERVICIO SAROGAR SA DE CV'),
    ('UAGG890111FP7', 'GUILLERMO UMAÑA GONZALEZ'),
]


def _safe_decimal(value, default=Decimal('0.00'), max_val=None) -> Decimal:
    if value is None:
        return default
    try:
        import math
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return default
        d = Decimal(str(value))
        if d.is_nan() or d.is_infinite():
            return default
        d = d.quantize(Decimal('0.01'))
        if max_val is not None and d > Decimal(str(max_val)):
            return Decimal(str(max_val))
        return d
    except (InvalidOperation, TypeError):
        return default


def _safe_decimal4(value, default=Decimal('1.0000')) -> Decimal:
    if value is None:
        return default
    try:
        return Decimal(str(value)).quantize(Decimal('0.0001'))
    except (InvalidOperation, TypeError):
        return default


def _safe_str(value, max_len=None) -> str:
    if value is None:
        return ''
    s = str(value).strip()
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def _safe_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value)[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _parse_period_label(raw) -> tuple[int, int, int] | None:
    """
    Parse 'MMM YYYY Qn' → (year, month, periodnumber).
    Returns None if unparseable.
    """
    if not raw:
        return None
    parts = _safe_str(raw).upper().split()
    if len(parts) != 3:
        return None
    mon_str, year_str, q_str = parts
    month = MONTH_LABELS_ES.get(mon_str)
    if not month:
        return None
    try:
        year = int(year_str)
        periodnumber = int(q_str.replace('Q', ''))
    except ValueError:
        return None
    return (year, month, periodnumber)


def _map_document_type(raw) -> tuple[int, bool]:
    """Return (documenttype_code, is_canceled)."""
    key = _safe_str(raw).lower()
    if 'cancelac' in key or 'cancelad' in key:
        return 0, True
    for k, v in DOCUMENT_TYPE_MAP.items():
        if k in key:
            return v, False
    return 0, False  # default INVOICE


def _map_payment_method(raw) -> int | None:
    if not raw:
        return None
    key = _safe_str(raw).lower()
    for k, v in PAYMENT_METHOD_MAP.items():
        if key.startswith(k) or key == k:
            return v
    return 99  # OTHER


def _map_payment_status(raw) -> int:
    if not raw:
        return 0  # PENDING
    key = _safe_str(raw).lower()
    for k, v in PAYMENT_STATUS_MAP.items():
        if k in key:
            return v
    return 0


def _is_payroll_rfc(rfc: str) -> bool:
    """Detect pseudo-RFC entries like 'Nomina-OPERACIÓN'."""
    if not rfc:
        return False
    rfc_upper = rfc.upper()
    return 'NOMINA' in rfc_upper or 'NÓMINA' in rfc_upper or len(rfc) > 13


def _parse_imputation_code(code_str: str) -> dict | None:
    """
    Parse code string into components.
    Direct:   'MAN-P1-5'  → {zone_prefix: 'MAN', category: 'P1', seq: 5, costtype: 0}
    Indirect: 'C3-5'      → {zone_prefix: None,  category: 'C3', seq: 5, costtype: 1}
    Returns None if unparseable.
    """
    s = _safe_str(code_str).upper()
    if not s:
        return None

    # Pattern: ZONE-PX-N (direct)
    m = re.match(r'^([A-Z]{2,4})-([PC]\d+)-(\d+)$', s)
    if m:
        return {
            'zone_prefix': m.group(1),
            'category': m.group(2),
            'seq': int(m.group(3)),
            'costtype': 0,  # DIRECT
        }

    # Pattern: CX-N (indirect)
    m = re.match(r'^(C\d+)-(\d+)$', s)
    if m:
        return {
            'zone_prefix': None,
            'category': m.group(1),
            'seq': int(m.group(2)),
            'costtype': 1,  # INDIRECT
        }

    # Pattern: PX-N (direct without zone prefix — unusual)
    m = re.match(r'^(P\d+)-(\d+)$', s)
    if m:
        return {
            'zone_prefix': None,
            'category': m.group(1),
            'seq': int(m.group(2)),
            'costtype': 0,  # DIRECT
        }

    return None


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = 'Import SYM KABAT project data from "002. Control y Seguimiento (SYM KABAT).xlsx"'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default=None,
            help='Absolute path to the Excel file. Defaults to erp_backend/ directory.',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Parse and validate without writing to database.',
        )

    def handle(self, *args, **options):
        dry_run: bool = options['dry_run']
        file_path: str = options['file']

        if not file_path:
            # Default: parents[5] = erp_backend/ (contains the Excel file)
            base = Path(__file__).resolve().parents[5]
            file_path = str(base / '002. Control y Seguimiento (SYM KABAT).xlsx')

        if not Path(file_path).exists():
            raise CommandError(f'File not found: {file_path}')

        self._run_import(file_path, dry_run)

    @transaction.atomic
    def _run_import(self, file_path: str, dry_run: bool):

        self.stdout.write(f'Loading workbook from: {file_path}')
        wb = _load_workbook_patched(file_path)
        sheet_names = wb.sheetnames
        self.stdout.write(f'Sheets found: {sheet_names}')

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no data will be written.\n'))

        # ----------------------------------------------------------------
        # 0. Get owner user
        # ----------------------------------------------------------------
        from apps.users.models import SystemUser, SecurityRole

        owner = SystemUser.objects.filter(
            securityroleid__name='System Administrator'
        ).first() or SystemUser.objects.first()

        if not owner:
            raise CommandError(
                'No SystemUser found. Create a superuser first:\n'
                '  python manage.py createsuperuser'
            )
        self.stdout.write(f'Owner: {owner.fullname} ({owner.emailaddress1})')

        # ----------------------------------------------------------------
        # 1. Create SYM KABAT Account (client)
        # ----------------------------------------------------------------
        from apps.accounts.models import Account, CustomerTypeCode

        client_account, created = Account.objects.get_or_create(
            name='SYM KABAT',
            defaults={
                'customertypecode': CustomerTypeCode.CUSTOMER,
                'ownerid': owner,
                'createdby': owner,
                'modifiedby': owner,
            },
        )
        if not dry_run and created:
            self.stdout.write(f'  [+] Account created: SYM KABAT ({client_account.accountid})')
        else:
            self.stdout.write(f'  [=] Account found: SYM KABAT')

        # ----------------------------------------------------------------
        # 2. Create ConstructionProject
        # ----------------------------------------------------------------
        from apps.projects.models import (
            ConstructionProject, ProjectStateCode,
            ProjectTypeCode, BiddingTypeCode, PeriodTypeCode,
            ProjectZone, ProjectSupplier,
        )

        PROJECT_NAME = 'CONSTRUCCIÓN DE 50 PILAS DE CONCRETO ARMADO PARA PMI\'S'
        project, created = ConstructionProject.objects.get_or_create(
            name=PROJECT_NAME,
            defaults={
                'projectnumber': 'PRY-2025-001',
                'statecode': ProjectStateCode.ACTIVE,
                'accountid': client_account,
                'startdate': date(2025, 6, 1),
                'contractenddate': date(2026, 2, 28),
                'durationmonths': 9,
                'projecttype': ProjectTypeCode.PUBLIC,
                'biddingtype': BiddingTypeCode.DIRECT_AWARD,
                'contractamount_notax': Decimal('7818391.45') / Decimal('1.16'),
                'contractamount_withtax': Decimal('7818391.45'),
                'advancepayment_notax': Decimal('929236.37') / Decimal('1.16'),
                'advancepayment_withtax': Decimal('929236.37'),
                'exchangerate_mxn_usd': Decimal('21.0000'),
                'periodtype': PeriodTypeCode.FORTNIGHTLY,
                'ownerid': owner,
                'createdby': owner,
                'modifiedby': owner,
            },
        )
        if not dry_run and created:
            self.stdout.write(f'  [+] Project created: {project.projectnumber} - {project.name}')
        else:
            self.stdout.write(f'  [=] Project found: {project.projectnumber}')

        # ----------------------------------------------------------------
        # 3. Create ProjectZones (TAM, MAN)
        # ----------------------------------------------------------------
        zones_data = [
            ('TAM', 'Tamaulipas', 1),
            ('MAN', 'Mante', 2),
        ]
        zone_map: dict[str, ProjectZone] = {}
        for prefix, name, sortorder in zones_data:
            zone, created = ProjectZone.objects.get_or_create(
                projectid=project,
                prefix=prefix,
                defaults={
                    'name': name,
                    'sortorder': sortorder,
                    'createdby': owner,
                    'modifiedby': owner,
                },
            )
            zone_map[prefix] = zone
            tag = '[+]' if created else '[=]'
            self.stdout.write(f'  {tag} Zone: [{prefix}] {name}')

        # ----------------------------------------------------------------
        # 4. Create supplier Accounts + ProjectSuppliers
        # ----------------------------------------------------------------
        rfc_to_supplier_account: dict[str, Account] = {}
        for idx, (rfc, business_name) in enumerate(SUPPLIERS_DATA, start=1):
            # Create or find the Account for this supplier
            sup_account, _ = Account.objects.get_or_create(
                name=business_name,
                defaults={
                    'customertypecode': CustomerTypeCode.BOTH,
                    'ownerid': owner,
                    'createdby': owner,
                    'modifiedby': owner,
                },
            )
            rfc_to_supplier_account[rfc] = sup_account

            ProjectSupplier.objects.get_or_create(
                projectid=project,
                rfc=rfc,
                defaults={
                    'accountid': sup_account,
                    'suppliernumber': idx,
                    'businessname': business_name,
                    'createdby': owner,
                    'modifiedby': owner,
                },
            )

        self.stdout.write(f'  [+] Suppliers: {len(SUPPLIERS_DATA)} processed')

        # ----------------------------------------------------------------
        # 5. Seed CostCategories
        # ----------------------------------------------------------------
        from apps.budgets.models import CostCategory, ImputationCode, ImputationPeriod
        from apps.budgets.services import CostCategoryService, PeriodService

        if not CostCategory.objects.filter(projectid=project).exists():
            categories = CostCategoryService.seed_default_categories(project.projectid, owner)
            self.stdout.write(f'  [+] CostCategories seeded: {len(categories)}')
        else:
            self.stdout.write(f'  [=] CostCategories already exist')

        category_map: dict[str, CostCategory] = {
            c.code: c for c in CostCategory.objects.filter(projectid=project)
        }

        # ----------------------------------------------------------------
        # 6. Parse unique imputation codes from Imput sheet
        # ----------------------------------------------------------------
        self.stdout.write('\nParsing Imput sheet for imputation codes...')
        ws_imput = wb['Imput']

        # Collect rows — data starts at row 12 (1-indexed), headers at row 11
        raw_rows = []
        for row_idx, row in enumerate(ws_imput.iter_rows(values_only=True), start=1):
            if row_idx < 12:
                continue
            if not any(v is not None for v in row):
                continue
            raw_rows.append(row)

        unique_codes: set[str] = set()
        for row in raw_rows:
            code_val = row[2] if len(row) > 2 else None  # C. Imp. at col[2]
            if code_val:
                unique_codes.add(_safe_str(code_val).upper())

        unique_codes.discard('')
        self.stdout.write(f'  Unique imputation codes found: {len(unique_codes)}')

        # Create ImputationCode records
        existing_codes = {
            ic.code: ic
            for ic in ImputationCode.objects.filter(projectid=project)
        }

        created_codes = 0
        for code_str in sorted(unique_codes):
            if code_str in existing_codes:
                continue
            parsed = _parse_imputation_code(code_str)
            if not parsed:
                self.stdout.write(
                    self.style.WARNING(f'  [!] Cannot parse code: {code_str} — skipping')
                )
                continue

            cat_key = parsed['category']
            category = category_map.get(cat_key)
            if not category:
                self.stdout.write(
                    self.style.WARNING(f'  [!] Category not found: {cat_key} for code {code_str} — skipping')
                )
                continue

            zone_prefix = parsed.get('zone_prefix')
            zone = zone_map.get(zone_prefix) if zone_prefix else None

            if not dry_run:
                ic = ImputationCode(
                    projectid=project,
                    categoryid=category,
                    zoneid=zone,
                    costtype=parsed['costtype'],
                    code=code_str,
                    sequencenumber=parsed['seq'],
                    name=code_str,  # code as name; can be updated later
                    totalbudget=Decimal('0'),
                    totalspent=Decimal('0'),
                    remainingbudget=Decimal('0'),
                    percentused=Decimal('0'),
                    createdby=owner,
                    modifiedby=owner,
                )
                ic.save()
                existing_codes[code_str] = ic
            created_codes += 1

        self.stdout.write(f'  [+] ImputationCodes created: {created_codes}')

        # ----------------------------------------------------------------
        # 7. Initialize ImputationPeriods
        # ----------------------------------------------------------------
        if not ImputationPeriod.objects.filter(projectid=project).exists():
            if not dry_run:
                periods = PeriodService.initialize_periods(project.projectid, owner)
                self.stdout.write(f'  [+] Periods initialized: {len(periods)}')
        else:
            self.stdout.write(f'  [=] Periods already exist')

        period_map: dict[tuple, ImputationPeriod] = {
            (p.year, p.month, p.periodnumber): p
            for p in ImputationPeriod.objects.filter(projectid=project)
        }

        # ----------------------------------------------------------------
        # 8. Import expenses from Imput sheet
        # ----------------------------------------------------------------
        from apps.expenses.models import (
            ProjectExpense, ExpenseLine,
            DocumentTypeCode, ExpenseScopeCode, ExpenseStateCode,
            ClassificationStatusCode, PaymentStatusCode,
            VerificationStatusCode, CurrencyCode,
        )

        self.stdout.write(f'\nImporting {len(raw_rows)} expense rows...')

        skipped = 0
        imported = 0
        already_exists = 0

        for row_idx, row in enumerate(raw_rows):
            # col[0]=blank, col[1]=PERIODO, col[2]=C.Imp, col[3]=RFC,
            # col[4]=RAZON SOCIAL, col[5]=Descripcion, col[6]=Cantidad,
            # col[7]=Precio Unitario, col[8]=Subtotal, col[9]=IVA,
            # col[10]=Retención, col[11]=Importe Neto,
            # col[12]=Importe Neto por Factura (extra 002 col — skip),
            # col[13]=Documento, col[14]=Tipo de Cambio,
            # col[15]=Folio UUID, col[16]=Folio, col[17]=Fecha,
            # col[18]=Tipo de Pago, col[19]=Estatus de Pago,
            # col[21]=CUENTA CONTABLE, col[22]=SUBCUENTA
            period_raw    = row[1]  if len(row) > 1  else None
            code_raw      = row[2]  if len(row) > 2  else None
            rfc_raw       = row[3]  if len(row) > 3  else None
            supplier_raw  = row[4]  if len(row) > 4  else None
            desc_raw      = row[5]  if len(row) > 5  else None
            qty_raw       = row[6]  if len(row) > 6  else None
            price_raw     = row[7]  if len(row) > 7  else None
            subtotal_raw  = row[8]  if len(row) > 8  else None
            tax_raw       = row[9]  if len(row) > 9  else None
            retention_raw = row[10] if len(row) > 10 else None
            net_raw       = row[11] if len(row) > 11 else None
            doc_raw       = row[13] if len(row) > 13 else None
            exrate_raw    = row[14] if len(row) > 14 else None
            uuid_raw      = row[15] if len(row) > 15 else None
            folio_raw     = row[16] if len(row) > 16 else None
            date_raw      = row[17] if len(row) > 17 else None
            pay_method_raw = row[18] if len(row) > 18 else None
            pay_status_raw = row[19] if len(row) > 19 else None
            acct_raw      = row[21] if len(row) > 21 else None
            sub_raw       = row[22] if len(row) > 22 else None

            # Skip if no meaningful amount
            subtotal = _safe_decimal(subtotal_raw)
            net_amount = _safe_decimal(net_raw)
            if subtotal == Decimal('0') and net_amount == Decimal('0'):
                skipped += 1
                continue

            # Period lookup
            period_key = _parse_period_label(period_raw)
            period_obj = period_map.get(period_key) if period_key else None

            # Imputation code lookup
            code_str = _safe_str(code_raw).upper()
            imputation_code_obj = existing_codes.get(code_str) if code_str else None

            # Document type + canceled flag
            doc_type, is_canceled = _map_document_type(doc_raw)

            # RFC / supplier
            rfc = _safe_str(rfc_raw, 13)
            is_payroll = _is_payroll_rfc(rfc) or doc_type == DocumentTypeCode.PAYROLL
            if is_payroll:
                rfc = None
                doc_type = DocumentTypeCode.PAYROLL

            supplier_name = _safe_str(supplier_raw, 300)
            invoice_uuid  = _safe_str(uuid_raw, 36) or None
            invoice_folio = _safe_str(folio_raw, 50) or None
            invoice_date  = _safe_date(date_raw)

            # Deduplicate by UUID per project (constraint in model)
            if invoice_uuid and not dry_run:
                if ProjectExpense.objects.filter(
                    projectid=project,
                    invoiceuuid=invoice_uuid,
                ).exists():
                    already_exists += 1
                    continue

            exchange_rate  = _safe_decimal4(exrate_raw)
            tax_amount     = _safe_decimal(tax_raw)
            retention      = _safe_decimal(retention_raw)
            pay_method     = _map_payment_method(pay_method_raw)
            pay_status     = _map_payment_status(pay_status_raw)
            state_code     = ExpenseStateCode.CANCELED if is_canceled else ExpenseStateCode.ACTIVE
            acct_account   = _safe_str(acct_raw, 100) or None
            sub_account    = _safe_str(sub_raw, 100) or None
            if dry_run:
                imported += 1
                continue

            expense = ProjectExpense(
                expensescope=ExpenseScopeCode.PROJECT,
                projectid=project,
                periodid=period_obj,
                imputationcodeid=imputation_code_obj,
                classificationstatus=(
                    ClassificationStatusCode.CLASSIFIED
                    if imputation_code_obj
                    else ClassificationStatusCode.PENDING
                ),
                documenttype=doc_type,
                supplierrfc=rfc,
                suppliername=supplier_name,
                invoiceuuid=invoice_uuid,
                invoicefolio=invoice_folio,
                invoicedate=invoice_date,
                payrolltype=None,
                paymentmethod=pay_method,
                paymentstatus=pay_status,
                currency=CurrencyCode.MXN,
                exchangerate=exchange_rate,
                subtotal=subtotal,
                taxamount=tax_amount,
                retentionamount=retention,
                discountamount=Decimal('0.00'),
                netamount=net_amount,
                verificationstatus=VerificationStatusCode.PENDING,
                statecode=state_code,
                accountingaccount=acct_account,
                subaccount=sub_account,
                notes=None,
                ownerid=owner,
                createdby=owner,
                modifiedby=owner,
            )
            expense.save()

            # Create single ExpenseLine
            qty  = _safe_decimal4(qty_raw, Decimal('1.0000'))
            price = _safe_decimal4(price_raw, Decimal('0.0000'))
            desc  = _safe_str(desc_raw, 500) or supplier_name

            ExpenseLine(
                expenseid=expense,
                linenumber=1,
                description=desc,
                quantity=qty,
                unitprice=price,
                subtotal=subtotal,
                taxamount=tax_amount,
                retentionamount=retention,
                discountamount=Decimal('0.00'),
                netamount=net_amount,
                imputationcodeid=imputation_code_obj,
            ).save()

            imported += 1

        self.stdout.write(
            f'  Imported: {imported} | Skipped (zero): {skipped} | '
            f'Duplicates: {already_exists}'
        )

        # ----------------------------------------------------------------
        # 9. Import ClientEstimates from Control Certificación
        # ----------------------------------------------------------------
        from apps.expenses.models import ClientEstimate, EstimateStateCode, EstimateTypeCode

        CERT_SHEET = 'Control Certificación'
        if CERT_SHEET not in sheet_names:
            # Try alternate name without accent
            CERT_SHEET = next(
                (s for s in sheet_names if 'certificac' in s.lower()),
                None,
            )

        estimates_imported = 0
        if CERT_SHEET:
            self.stdout.write(f'\nImporting ClientEstimates from sheet "{CERT_SHEET}"...')
            ws_cert = wb[CERT_SHEET]

            # Header at rows 9-10, data from row 13
            # Cols (0-based): 0=Factura No., 1=Fecha Factura, 2=Periodo Imp.,
            # 3=APLICA, 4=Periodo Estimación, 5=Importe Estimación,
            # 6=Anticipo, 7=Amortización, 8=Otros, 9=Material, 10=Garantía,
            # 11=Total Deducciones, 12=Sin IVA, 13=IVA, 14=IVAR,
            # 15=Líquido a Pagar, 16=Fecha Pago, 17=Estatus
            estimate_num = 0
            for row_idx, row in enumerate(ws_cert.iter_rows(values_only=True), start=1):
                # col[0]=blank, col[1]=FACTURA No., col[2]=FECHA FACTURA,
                # col[3]=PERIODO Imp., col[4]=APLICA, col[5]=PERIODO ESTIMACIÓN,
                # col[6]=IMPORTE, col[7]=ANTICIPO, col[8]=AMORT. ANTICIPO,
                # col[9]=OTROS, col[10]=MATERIAL, col[11]=FONDO GARANTÍA,
                # col[12]=TOTAL DEDUCCIONES, col[13]=IMPORTE SIN IVA,
                # col[14]=IVA, col[15]=IVAR, col[16]=LIQUIDO A PAGAR,
                # col[17]=FECHA DE PAGO, col[18]=ESTATUS
                # Data rows start at row 15
                if row_idx < 15:
                    continue
                if not any(v is not None for v in row):
                    continue
                # row[1] must have invoice number
                if not row[1] if len(row) > 1 else True:
                    continue

                invoice_no   = _safe_str(row[1], 50) if len(row) > 1 else ''
                inv_date     = _safe_date(row[2]) if len(row) > 2 else None
                period_raw   = row[3]  if len(row) > 3  else None
                aplica_raw   = _safe_str(row[4]).lower() if len(row) > 4 else ''
                est_period   = _safe_str(row[5], 500) if len(row) > 5 else ''
                est_amount   = _safe_decimal(row[6]  if len(row) > 6  else None)
                advance      = _safe_decimal(row[7]  if len(row) > 7  else None)
                amort        = _safe_decimal(row[8]  if len(row) > 8  else None)
                otros        = _safe_decimal(row[9]  if len(row) > 9  else None)
                material     = _safe_decimal(row[10] if len(row) > 10 else None)
                garantia     = _safe_decimal(row[11] if len(row) > 11 else None)
                total_ded    = _safe_decimal(row[12] if len(row) > 12 else None)
                sin_iva      = _safe_decimal(row[13] if len(row) > 13 else None)
                iva          = _safe_decimal(row[14] if len(row) > 14 else None)
                ivar         = _safe_decimal(row[15] if len(row) > 15 else None)
                liquido      = _safe_decimal(row[16] if len(row) > 16 else None)
                pay_date     = _safe_date(row[17]) if len(row) > 17 else None
                estatus_raw  = _safe_str(row[18]).lower() if len(row) > 18 else ''

                if not invoice_no:
                    continue

                # Skip if all amounts zero
                if sin_iva == Decimal('0') and liquido == Decimal('0'):
                    continue

                period_key = _parse_period_label(period_raw)
                period_obj = period_map.get(period_key) if period_key else None

                # Detect canceled rows (negative amounts or description says CANCELADO)
                is_canceled_est = (
                    'cancelad' in aplica_raw or
                    'cancelac' in aplica_raw or
                    'cancelado' in est_period.lower() or
                    sin_iva < Decimal('0')
                )

                est_type = ESTIMATE_TYPE_MAP.get(aplica_raw, 0)
                pay_status_est = _map_payment_status(estatus_raw)

                estimate_num += 1

                if not dry_run:
                    if ClientEstimate.objects.filter(
                        projectid=project,
                        invoicenumber=invoice_no,
                    ).exists():
                        continue

                    ClientEstimate(
                        projectid=project,
                        periodid=period_obj,
                        estimatenumber=estimate_num,
                        invoicenumber=invoice_no,
                        invoicedate=inv_date,
                        estimationperiod=_safe_str(est_period, 50),
                        estimatetype=est_type,
                        estimatedamount=est_amount,
                        advanceamortization=amort,
                        otherdeductions=otros,
                        materialdeductions=material,
                        guaranteefund=garantia,
                        totaldeductions=total_ded,
                        amountnotax=sin_iva,
                        taxamount=iva,
                        taxretained=ivar,
                        totalinvoiced=sin_iva + iva,
                        collectableamount=liquido,
                        paymentstatus=pay_status_est,
                        paymentdate=pay_date,
                        amountpaid=Decimal('0.00'),
                        statecode=(
                            EstimateStateCode.CANCELED
                            if is_canceled_est
                            else EstimateStateCode.ACTIVE
                        ),
                        createdby=owner,
                        modifiedby=owner,
                    ).save()

                estimates_imported += 1

            self.stdout.write(f'  ClientEstimates imported: {estimates_imported}')
        else:
            self.stdout.write(
                self.style.WARNING('  Sheet "Control Certificación" not found — skipping estimates.')
            )

        # ----------------------------------------------------------------
        # Summary
        # ----------------------------------------------------------------
        self.stdout.write('\n' + '=' * 60)
        if dry_run:
            transaction.set_rollback(True)
            self.stdout.write(self.style.WARNING(
                'DRY RUN complete — nothing written to database.\n'
                f'  Would import: {imported} expenses, {estimates_imported} estimates'
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'Import complete!\n'
                f'  Project:    {project.projectnumber} — {project.name}\n'
                f'  Suppliers:  {len(SUPPLIERS_DATA)}\n'
                f'  Imp.Codes:  {created_codes} new\n'
                f'  Expenses:   {imported}\n'
                f'  Estimates:  {estimates_imported}'
            ))
