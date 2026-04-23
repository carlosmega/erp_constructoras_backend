"""
Management command to import budget data from "Costo Directo" and
"Indirect Costs" sheets in the SYM KABAT Excel file.

For each ImputationCode data row this command:
  - Creates the code if it doesn't exist yet (new zones created as needed)
  - Updates: name, description, estimatedsupplier, unitcost, quantity,
             executionmonths, totalbudget, totalspent, remainingbudget,
             percentused, personnelname, personnelrole, monthlycost, units
  - Creates ImputationCodeBudget rows (48 planned + 48 actual per code)

Usage:
    python manage.py import_budget_sheets
    python manage.py import_budget_sheets --file /path/to/file.xlsx
    python manage.py import_budget_sheets --dry-run
"""

import io
import re
import uuid
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

# ---------------------------------------------------------------------------
# Helpers shared with import_sym_kabat
# ---------------------------------------------------------------------------

def _load_workbook_patched(path: str):
    with open(path, 'rb') as fh:
        raw = fh.read()
    buf = io.BytesIO(raw)
    with zipfile.ZipFile(buf, 'r') as zin:
        files = {n: zin.read(n) for n in zin.namelist()}
    files['xl/workbook.xml'] = re.sub(
        rb'<definedNames>.*?</definedNames>',
        b'',
        files.get('xl/workbook.xml', b''),
        flags=re.DOTALL,
    )
    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for n, d in files.items():
            zout.writestr(n, d)
    out.seek(0)
    return openpyxl.load_workbook(out, data_only=True, read_only=True)


def _safe_decimal(value, default=Decimal('0.00'), max_val=None) -> Decimal:
    if value is None:
        return default
    try:
        import math
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return default
        d = Decimal(str(value))
        if d.is_nan() or d.is_infinite():
            return default
        d = d.quantize(Decimal('0.01'))
        if max_val is not None and d > Decimal(str(max_val)):
            return Decimal(str(max_val))
        return d
    except (InvalidOperation, TypeError):
        return default


def _safe_str(value, max_len=None) -> str:
    if value is None:
        return ''
    s = str(value).strip()
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


# ---------------------------------------------------------------------------
# Period label list (48 fortnightly periods JUN 2025 Q1 → MAY 2027 Q2)
# ---------------------------------------------------------------------------
MONTH_LABELS_INT = {
    1: 'ENE', 2: 'FEB', 3: 'MAR', 4: 'ABR',
    5: 'MAY', 6: 'JUN', 7: 'JUL', 8: 'AGO',
    9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DIC',
}

def _build_period_labels(start_year: int, start_month: int, count: int) -> list[str]:
    """Generate `count` fortnightly period labels starting from given year/month."""
    labels = []
    year, month = start_year, start_month
    for _ in range(count):
        mon = MONTH_LABELS_INT[month]
        labels.append(f'{mon} {year} Q1')
        labels.append(f'{mon} {year} Q2')
        month += 1
        if month > 12:
            month = 1
            year += 1
    return labels


# 48 periods = 24 months starting JUN 2025
PERIOD_LABELS_48 = _build_period_labels(2025, 6, 24)  # 48 entries

# Regex patterns to identify data rows
RE_DIRECT_CODE = re.compile(r'^[A-Z]{2,5}-[PC]\d+-\d+$')
RE_INDIRECT_CODE = re.compile(r'^C\d+-\d+$')

# Zone prefix → (full name, sortorder)
ZONE_NAMES = {
    'MAN': ('Mante', 2),
    'TAM': ('Tampico', 3),
    'VIC': ('Ciudad Victoria', 4),
    'MAT': ('Matamoros', 5),
    'MEC': ('Mecánicas de Suelos y Topografías', 6),
    'SUT': ('Mecánicas de Suelos y Topografías', 6),
}


class Command(BaseCommand):
    help = 'Import budget data from Costo Directo and Indirect Costs sheets'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true', default=False)
        parser.add_argument(
            '--project-number',
            type=str,
            default='PRY-2025-001',
            help='Project number to import into',
        )

    def handle(self, *args, **options):
        file_path = options['file']
        if not file_path:
            base = Path(__file__).resolve().parents[5]
            file_path = str(base / '002. Control y Seguimiento (SYM KABAT).xlsx')

        if not Path(file_path).exists():
            raise CommandError(f'File not found: {file_path}')

        self._run(file_path, options['dry_run'], options['project_number'])

    @transaction.atomic
    def _run(self, file_path: str, dry_run: bool, project_number: str):
        from apps.projects.models import ConstructionProject, ProjectZone
        from apps.budgets.models import (
            CostCategory, ImputationCode, ImputationPeriod,
            ImputationCodeBudget, CostTypeCode,
        )
        from apps.users.models import SystemUser

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — nothing will be written.\n'))

        # ── owner & project ────────────────────────────────────────────────
        owner = (
            SystemUser.objects.filter(securityroleid__name='System Administrator').first()
            or SystemUser.objects.first()
        )
        if not owner:
            raise CommandError('No SystemUser found. Run import_sym_kabat first.')

        try:
            project = ConstructionProject.objects.get(projectnumber=project_number)
        except ConstructionProject.DoesNotExist:
            raise CommandError(
                f'Project {project_number} not found. Run import_sym_kabat first.'
            )
        self.stdout.write(f'Project: {project.projectnumber} — {project.name}')

        # ── caches ──────────────────────────────────────────────────────────
        category_map: dict[str, CostCategory] = {
            c.code: c for c in CostCategory.objects.filter(projectid=project)
        }
        zone_map: dict[str, ProjectZone] = {
            z.prefix: z for z in ProjectZone.objects.filter(projectid=project)
        }
        period_map: dict[str, ImputationPeriod] = {
            p.label: p for p in ImputationPeriod.objects.filter(projectid=project)
        }
        code_map: dict[str, ImputationCode] = {
            ic.code: ic
            for ic in ImputationCode.objects.filter(projectid=project)
        }

        # ── load workbook ───────────────────────────────────────────────────
        self.stdout.write(f'Loading workbook: {file_path}')
        wb = _load_workbook_patched(file_path)

        total_codes = 0
        total_budgets = 0

        # ── process Costo Directo ───────────────────────────────────────────
        self.stdout.write('\n[Costo Directo]')
        cd_codes, cd_budgets = self._process_direct(
            wb['Costo Directo'],
            project, owner,
            category_map, zone_map, period_map, code_map,
            dry_run,
        )
        self.stdout.write(
            f'  Codes: {cd_codes} updated/created | Budget lines: {cd_budgets}'
        )
        total_codes += cd_codes
        total_budgets += cd_budgets

        # ── process Indirect Costs ──────────────────────────────────────────
        self.stdout.write('\n[Indirect Costs]')
        ic_codes, ic_budgets = self._process_indirect(
            wb['Indirect Costs'],
            project, owner,
            category_map, zone_map, period_map, code_map,
            dry_run,
        )
        self.stdout.write(
            f'  Codes: {ic_codes} updated/created | Budget lines: {ic_budgets}'
        )
        total_codes += ic_codes
        total_budgets += ic_budgets

        # ── summary ─────────────────────────────────────────────────────────
        self.stdout.write('\n' + '=' * 60)
        if dry_run:
            transaction.set_rollback(True)
            self.stdout.write(self.style.WARNING(
                f'DRY RUN complete — nothing written.\n'
                f'  Would process: {total_codes} codes, {total_budgets} budget lines'
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'Import complete!\n'
                f'  Codes updated/created: {total_codes}\n'
                f'  Budget lines created:  {total_budgets}'
            ))

    # ── Costo Directo processor ─────────────────────────────────────────────

    def _process_direct(
        self, ws, project, owner,
        category_map, zone_map, period_map, code_map,
        dry_run,
    ):
        """
        Column layout (Costo Directo):
          col[3]  = ImputationCode  (e.g. MAN-P1-1)
          col[4]  = Description/Name
          col[5]  = Supplier (Proveedor) — mostly None
          col[6]  = Unit cost (Costo Empresa)
          col[7]  = Quantity (Cantidad)
          col[8]  = Months (Meses)
          col[9]  = Total budget (Importe)
          col[11-58] = Planned period fractions (× Importe = planned MXN)
          col[61] = % Imputado
          col[62] = Imputado Acumulado (total actual)
          col[64-111] = Actual per-period amounts
          Data rows start at row 17.
        """
        from apps.projects.models import ProjectZone
        from apps.budgets.models import (
            ImputationCode, ImputationCodeBudget, CostTypeCode,
        )

        PLANNED_START = 11
        ACTUAL_START = 64
        N_PERIODS = 48

        codes_processed = 0
        budgets_created = 0

        bulk_budgets: list[ImputationCodeBudget] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx < 17:
                continue
            if not row or len(row) < 10:
                continue

            code_val = _safe_str(row[3]).upper()
            if not RE_DIRECT_CODE.match(code_val):
                continue  # zone row, category row, sub-header, blank

            desc      = _safe_str(row[4], 300) or code_val
            supplier  = _safe_str(row[5], 200) or None
            unitcost  = _safe_decimal(row[6])
            quantity  = _safe_decimal(row[7])
            months    = _safe_decimal(row[8])
            total_bud = _safe_decimal(row[9])
            pct_used  = _safe_decimal(row[61], max_val=99) if len(row) > 61 else Decimal('0')
            total_spt = _safe_decimal(row[62]) if len(row) > 62 else Decimal('0')
            remaining = total_bud - total_spt

            # Parse zone prefix
            parts = code_val.split('-')   # e.g. ['MAN', 'P1', '1']
            zone_prefix = parts[0]
            cat_code    = parts[1]        # e.g. 'P1'
            seq_num     = int(parts[2])

            # Ensure zone exists
            zone = zone_map.get(zone_prefix)
            if not zone and not dry_run:
                zone_name, sortorder = ZONE_NAMES.get(
                    zone_prefix, (zone_prefix, 99)
                )
                zone, _ = ProjectZone.objects.get_or_create(
                    projectid=project,
                    prefix=zone_prefix,
                    defaults={
                        'name': zone_name,
                        'sortorder': sortorder,
                        'createdby': owner,
                        'modifiedby': owner,
                    },
                )
                zone_map[zone_prefix] = zone
                self.stdout.write(f'  [+] Zone created: [{zone_prefix}] {zone.name}')

            category = category_map.get(cat_code)
            if not category:
                self.stdout.write(
                    self.style.WARNING(f'  [!] Category {cat_code} not found — skipping {code_val}')
                )
                continue

            # Get or create ImputationCode
            ic = code_map.get(code_val)
            if not dry_run:
                if ic:
                    # Update fields from budget sheet
                    ic.name = desc
                    ic.estimatedsupplier = supplier
                    ic.unitcost = unitcost
                    ic.quantity = quantity
                    ic.executionmonths = int(months) if months else None
                    ic.totalbudget = total_bud
                    ic.totalspent = total_spt
                    ic.remainingbudget = remaining
                    ic.percentused = pct_used * 100
                    ic.zoneid = zone
                    ic.modifiedby = owner
                    ic.save()
                else:
                    ic = ImputationCode(
                        projectid=project,
                        categoryid=category,
                        zoneid=zone,
                        costtype=CostTypeCode.DIRECT,
                        code=code_val,
                        sequencenumber=seq_num,
                        name=desc,
                        estimatedsupplier=supplier,
                        unitcost=unitcost,
                        quantity=quantity,
                        executionmonths=int(months) if months else None,
                        totalbudget=total_bud,
                        totalspent=total_spt,
                        remainingbudget=remaining,
                        percentused=pct_used * 100,
                        createdby=owner,
                        modifiedby=owner,
                    )
                    ic.save()
                    code_map[code_val] = ic

            codes_processed += 1

            # Build per-period budget lines
            if ic or dry_run:
                for i, label in enumerate(PERIOD_LABELS_48):
                    planned_col = PLANNED_START + i
                    actual_col  = ACTUAL_START + i

                    planned_frac = Decimal(str(row[planned_col])) if (
                        len(row) > planned_col and row[planned_col] is not None
                    ) else Decimal('0')
                    actual_amt = _safe_decimal(
                        row[actual_col] if len(row) > actual_col else None
                    )

                    # Convert fraction → MXN planned
                    planned_amt = (planned_frac * total_bud).quantize(Decimal('0.01'))

                    if planned_amt == 0 and actual_amt == 0:
                        continue  # skip empty periods

                    if not dry_run and ic:
                        bulk_budgets.append(ImputationCodeBudget(
                            imputationcodeid=ic,
                            periodid=period_map.get(label),
                            periodlabel=label,
                            plannedamount=planned_amt,
                            actualamount=actual_amt,
                        ))
                    budgets_created += 1

        # Bulk insert budget lines (ignore conflicts = re-runs are safe)
        if not dry_run and bulk_budgets:
            ImputationCodeBudget.objects.bulk_create(
                bulk_budgets, ignore_conflicts=True, batch_size=500
            )

        return codes_processed, budgets_created

    # ── Indirect Costs processor ────────────────────────────────────────────

    def _process_indirect(
        self, ws, project, owner,
        category_map, zone_map, period_map, code_map,
        dry_run,
    ):
        """
        Column layout (Indirect Costs):
          col[2]  = ImputationCode (e.g. C1-1)
          col[3]  = Area (role)
          col[4]  = Description (full name)
          col[5]  = Monthly cost (Costo mensual)
          col[6]  = Units (Unidades)
          col[7]  = Months (Meses)
          col[8]  = Total budget (Importe Total)
          col[10-57] = Planned period absolute amounts
          col[59] = % Imputado
          col[60] = Imputado Acumulado
          col[62-109] = Actual per-period amounts
          Data rows start at row 13.
        """
        from apps.budgets.models import (
            ImputationCode, ImputationCodeBudget, CostTypeCode,
            PersonnelTypeCode,
        )

        PLANNED_START = 10
        ACTUAL_START  = 62
        N_PERIODS     = 48

        codes_processed = 0
        budgets_created = 0
        bulk_budgets: list[ImputationCodeBudget] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx < 13:
                continue
            if not row or len(row) < 9:
                continue

            code_val = _safe_str(row[2]).upper()
            if not RE_INDIRECT_CODE.match(code_val):
                continue

            area      = _safe_str(row[3], 100) or None
            desc      = _safe_str(row[4], 300) or code_val
            monthly   = _safe_decimal(row[5])
            units     = _safe_decimal(row[6])
            months    = _safe_decimal(row[7])
            total_bud = _safe_decimal(row[8])
            pct_used  = _safe_decimal(row[59], max_val=99) if len(row) > 59 else Decimal('0')
            total_spt = _safe_decimal(row[60]) if len(row) > 60 else Decimal('0')
            remaining = total_bud - total_spt

            # Parse category from code e.g. 'C1-3' → 'C1'
            parts    = code_val.split('-')
            cat_code = parts[0]
            seq_num  = int(parts[1])

            category = category_map.get(cat_code)
            if not category:
                self.stdout.write(
                    self.style.WARNING(f'  [!] Category {cat_code} not found — skipping {code_val}')
                )
                continue

            # Determine if C1 (personnel) for extra fields
            is_personnel = cat_code == 'C1'

            ic = code_map.get(code_val)
            if not dry_run:
                if ic:
                    ic.name = desc
                    ic.monthlycost = monthly
                    ic.units = units
                    ic.executionmonths = int(months) if months else None
                    ic.totalbudget = total_bud
                    ic.totalspent = total_spt
                    ic.remainingbudget = remaining
                    ic.percentused = pct_used * 100
                    if is_personnel:
                        ic.personnelname = desc
                        ic.personnelrole = area
                        ic.personneltype = PersonnelTypeCode.FIELD_STAFF
                    ic.modifiedby = owner
                    ic.save()
                else:
                    ic = ImputationCode(
                        projectid=project,
                        categoryid=category,
                        zoneid=None,
                        costtype=CostTypeCode.INDIRECT,
                        code=code_val,
                        sequencenumber=seq_num,
                        name=desc,
                        monthlycost=monthly,
                        units=units,
                        executionmonths=int(months) if months else None,
                        totalbudget=total_bud,
                        totalspent=total_spt,
                        remainingbudget=remaining,
                        percentused=pct_used * 100,
                        personnelname=desc if is_personnel else None,
                        personnelrole=area if is_personnel else None,
                        personneltype=PersonnelTypeCode.FIELD_STAFF if is_personnel else None,
                        createdby=owner,
                        modifiedby=owner,
                    )
                    ic.save()
                    code_map[code_val] = ic

            codes_processed += 1

            if ic or dry_run:
                for i, label in enumerate(PERIOD_LABELS_48):
                    planned_col = PLANNED_START + i
                    actual_col  = ACTUAL_START + i

                    # Indirect: planned cols are already absolute amounts
                    planned_amt = _safe_decimal(
                        row[planned_col] if len(row) > planned_col else None
                    )
                    actual_amt = _safe_decimal(
                        row[actual_col] if len(row) > actual_col else None
                    )

                    if planned_amt == 0 and actual_amt == 0:
                        continue

                    if not dry_run and ic:
                        bulk_budgets.append(ImputationCodeBudget(
                            imputationcodeid=ic,
                            periodid=period_map.get(label),
                            periodlabel=label,
                            plannedamount=planned_amt,
                            actualamount=actual_amt,
                        ))
                    budgets_created += 1

        if not dry_run and bulk_budgets:
            ImputationCodeBudget.objects.bulk_create(
                bulk_budgets, ignore_conflicts=True, batch_size=500
            )

        return codes_processed, budgets_created
