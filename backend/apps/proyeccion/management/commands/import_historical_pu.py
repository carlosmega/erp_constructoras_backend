"""Management command to import historical P.U. from Excel into ConceptPriceCatalog.

Supports:
- Sheet "Base de datos": 747 concepts across 8 projects
- Sheet "COPY": 627 concepts across 4 projects (418 unique to this sheet)
- Deduplication: safe to re-run without creating duplicates
- --clear: wipe all HISTORICO data before importing
- --include-copy: also import concepts from the COPY sheet
- --dry-run: preview without writing
"""

import openpyxl
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.proyeccion.models import (
    ConceptPriceCatalogItem,
    ConceptPriceReference,
    CatalogSourceCode,
)
from apps.users.models import SystemUser


# ─── Sheet "Base de datos" layout ────────────────────────────────────────────
# Each tuple: (project_name, pu_col, cant_col, importe_col) — 1-based columns
SHEET1_PROJECTS = [
    ('Jenner Texcoco', 6, 7, 8),          # F, G, H
    ('Jenner Jardines', 10, 11, 12),       # J, K, L
    ('Swiss Lab Mty', 14, 15, 16),         # N, O, P
    ('Polab Morelos', 18, 19, 20),         # R, S, T
    ('Cumbres Elite', 22, 23, 24),         # V, W, X
    ('Valle', 26, 27, 28),                 # Z, AA, AB
    ('Swiss Lab Oscar Leon', 30, 31, 32),  # AD, AE, AF
    ('La Selva Tripp', 34, 35, 36),        # AH, AI, AJ
]
SHEET1_DATA_START = 8
SHEET1_DESC_COL = 3   # Column C
SHEET1_UNIT_COL = 4   # Column D

# ─── Sheet "COPY" layout ─────────────────────────────────────────────────────
# Only 4 projects, P.U. + IMPORTE (no separate CANT column per project)
# Quantity is in column E (for the Gomez Morin budget itself)
COPY_PROJECTS = [
    ('Jenner Texcoco', 7, 8),       # G=P.U., H=IMPORTE
    ('Jenner Jardines', 10, 11),     # J=P.U., K=IMPORTE
    ('Swiss Lab Mty', 13, 14),       # M=P.U., N=IMPORTE
    ('Polab Morelos', 16, 17),       # P=P.U., Q=IMPORTE
]
COPY_DATA_START = 10
COPY_DATA_END = 636
COPY_DESC_COL = 3   # Column C
COPY_UNIT_COL = 4   # Column D
COPY_CANT_COL = 5   # Column E (quantity for Gomez Morin)


def safe_decimal(value):
    """Convert a cell value to Decimal, returning None if invalid/zero."""
    if value is None or value == '' or str(value).strip() == '':
        return None
    try:
        d = Decimal(str(value).strip())
        return d if d > 0 else None
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = 'Import historical P.U. data from Excel into ConceptPriceCatalog'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_path',
            type=str,
            help='Path to the Excel file (Base de datos.xlsx)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without writing to database',
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Delete all existing HISTORICO catalog items before importing',
        )
        parser.add_argument(
            '--include-copy',
            action='store_true',
            help='Also import concepts from the COPY sheet',
        )

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        dry_run = options['dry_run']
        clear = options['clear']
        include_copy = options['include_copy']

        # ── Load workbook ────────────────────────────────────────────────
        self.stdout.write(f'Loading workbook: {excel_path}')
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.stdout.write(f'Sheets found: {wb.sheetnames}')

        # ── Parse main sheet ─────────────────────────────────────────────
        concepts = {}  # key: (desc_lower, unit_lower) -> {description, unit, references: [{...}]}
        self._parse_main_sheet(wb, concepts)

        # ── Parse COPY sheet ─────────────────────────────────────────────
        if include_copy and 'COPY' in wb.sheetnames:
            self._parse_copy_sheet(wb, concepts)

        wb.close()

        total_refs = sum(len(c['references']) for c in concepts.values())
        self.stdout.write(f'Total unique concepts: {len(concepts)}, total references: {total_refs}')

        # ── Dry run ──────────────────────────────────────────────────────
        if dry_run:
            self._print_preview(concepts)
            return

        # ── Get audit user ───────────────────────────────────────────────
        owner = SystemUser.objects.first()
        if not owner:
            self.stderr.write(self.style.ERROR(
                'No SystemUser found. Run load_dummy_data first.'
            ))
            return

        # ── Clear if requested ───────────────────────────────────────────
        if clear:
            deleted_refs = ConceptPriceReference.objects.filter(
                catalogitemid__source=CatalogSourceCode.HISTORICO
            ).delete()[0]
            deleted_items = ConceptPriceCatalogItem.objects.filter(
                source=CatalogSourceCode.HISTORICO
            ).delete()[0]
            self.stdout.write(self.style.WARNING(
                f'Cleared: {deleted_items} catalog items, {deleted_refs} references'
            ))

        # ── Import with deduplication ────────────────────────────────────
        self._import(concepts, owner)

    # ─────────────────────────────────────────────────────────────────────────
    # Parsing
    # ─────────────────────────────────────────────────────────────────────────

    def _parse_main_sheet(self, wb, concepts):
        """Parse the 'Base de datos' sheet into the concepts dict."""
        sheet_name = 'Base de datos'
        if sheet_name not in wb.sheetnames:
            self.stderr.write(f'Sheet "{sheet_name}" not found.')
            return

        ws = wb[sheet_name]
        parsed = 0
        skipped = 0

        for row_idx in range(SHEET1_DATA_START, ws.max_row + 1):
            desc = ws.cell(row=row_idx, column=SHEET1_DESC_COL).value
            unit = ws.cell(row=row_idx, column=SHEET1_UNIT_COL).value

            if not desc or not str(desc).strip():
                skipped += 1
                continue

            desc_clean = str(desc).strip()
            unit_clean = str(unit).strip() if unit else 'pza'
            key = (desc_clean.lower(), unit_clean.lower())

            if key not in concepts:
                concepts[key] = {
                    'description': desc_clean,
                    'unit': unit_clean,
                    'references': {},  # keyed by (projectname, unitprice) to deduplicate
                }

            for proj_name, pu_col, cant_col, imp_col in SHEET1_PROJECTS:
                pu = safe_decimal(ws.cell(row=row_idx, column=pu_col).value)
                cant = safe_decimal(ws.cell(row=row_idx, column=cant_col).value)
                imp = safe_decimal(ws.cell(row=row_idx, column=imp_col).value)

                if pu:
                    ref_key = (proj_name, str(pu))
                    if ref_key not in concepts[key]['references']:
                        concepts[key]['references'][ref_key] = {
                            'projectname': proj_name,
                            'unitprice': pu,
                            'quantity': cant,
                            'totalamount': imp,
                            'notes': f"Excel 'Base de datos' row {row_idx}",
                        }
            parsed += 1

        self.stdout.write(
            f'[Base de datos] Parsed {parsed} concepts, skipped {skipped} empty rows'
        )

    def _parse_copy_sheet(self, wb, concepts):
        """Parse the 'COPY' sheet, merging new concepts and references."""
        ws = wb['COPY']
        new_concepts = 0
        new_refs = 0

        for row_idx in range(COPY_DATA_START, COPY_DATA_END + 1):
            desc = ws.cell(row=row_idx, column=COPY_DESC_COL).value
            unit = ws.cell(row=row_idx, column=COPY_UNIT_COL).value

            if not desc or not str(desc).strip():
                continue

            desc_clean = str(desc).strip()
            unit_clean = str(unit).strip() if unit else 'pza'
            key = (desc_clean.lower(), unit_clean.lower())

            is_new = key not in concepts
            if is_new:
                concepts[key] = {
                    'description': desc_clean,
                    'unit': unit_clean,
                    'references': {},
                }
                new_concepts += 1

            for proj_name, pu_col, imp_col in COPY_PROJECTS:
                pu = safe_decimal(ws.cell(row=row_idx, column=pu_col).value)
                imp = safe_decimal(ws.cell(row=row_idx, column=imp_col).value)

                if pu:
                    ref_key = (proj_name, str(pu))
                    if ref_key not in concepts[key]['references']:
                        concepts[key]['references'][ref_key] = {
                            'projectname': proj_name,
                            'unitprice': pu,
                            'quantity': None,
                            'totalamount': imp,
                            'notes': f"Excel 'COPY' row {row_idx}",
                        }
                        new_refs += 1

        self.stdout.write(
            f'[COPY] Added {new_concepts} new concepts, {new_refs} new references'
        )

    # ─────────────────────────────────────────────────────────────────────────
    # Import
    # ─────────────────────────────────────────────────────────────────────────

    def _import(self, concepts, owner):
        """Import concepts into DB with deduplication against existing records."""
        # Build lookup of existing HISTORICO items
        existing = {}
        for item in ConceptPriceCatalogItem.objects.filter(
            source=CatalogSourceCode.HISTORICO, statecode=0
        ):
            key = (item.description.strip().lower(), item.unit.strip().lower())
            existing[key] = item

        # Build lookup of existing references per catalog item
        existing_refs = {}  # catalogitemid -> set of (projectname, normalized_price)
        if existing:
            for ref in ConceptPriceReference.objects.filter(
                catalogitemid__in=[item.catalogitemid for item in existing.values()],
                statecode=0
            ):
                item_id = ref.catalogitemid_id
                if item_id not in existing_refs:
                    existing_refs[item_id] = set()
                existing_refs[item_id].add(
                    (ref.projectname, f"{ref.unitprice:.4f}")
                )

        created_items = 0
        deduped_items = 0
        created_refs = 0
        skipped_refs = 0
        code_counter = self._get_next_counter()

        with transaction.atomic():
            for key, data in concepts.items():
                refs_list = list(data['references'].values())

                if key in existing:
                    # Item already exists — only add missing references
                    catalog_item = existing[key]
                    deduped_items += 1
                    item_ref_keys = existing_refs.get(catalog_item.catalogitemid, set())
                else:
                    # Create new catalog item
                    code_counter += 1
                    catalog_item = ConceptPriceCatalogItem(
                        code=f"HIST-{code_counter:05d}",
                        description=data['description'],
                        unit=data['unit'],
                        source=CatalogSourceCode.HISTORICO,
                        category='',
                        createdby=owner,
                        modifiedby=owner,
                    )
                    catalog_item.save()
                    created_items += 1
                    item_ref_keys = set()

                # Add references that don't already exist
                refs_added = False
                for ref_data in refs_list:
                    ref_key = (
                        ref_data['projectname'],
                        f"{ref_data['unitprice']:.4f}"
                    )
                    if ref_key in item_ref_keys:
                        skipped_refs += 1
                        continue

                    ConceptPriceReference(
                        catalogitemid=catalog_item,
                        projectname=ref_data['projectname'],
                        unitprice=ref_data['unitprice'],
                        quantity=ref_data.get('quantity'),
                        totalamount=ref_data.get('totalamount'),
                        notes=ref_data.get('notes', ''),
                        createdby=owner,
                        modifiedby=owner,
                    ).save()
                    created_refs += 1
                    refs_added = True

                # Recompute stats if we added any references
                if refs_added or key not in existing:
                    catalog_item.update_price_stats()
                    catalog_item.save()

        self.stdout.write(self.style.SUCCESS(
            f'\nImport complete:'
            f'\n  Catalog items created: {created_items}'
            f'\n  Catalog items already existed (deduped): {deduped_items}'
            f'\n  Price references created: {created_refs}'
            f'\n  Price references skipped (already existed): {skipped_refs}'
        ))

    # ─────────────────────────────────────────────────────────────────────────
    # Helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _get_next_counter(self):
        """Get the current max counter for HIST- codes."""
        last = (
            ConceptPriceCatalogItem.objects
            .filter(code__startswith='HIST-')
            .order_by('-code')
            .values_list('code', flat=True)
            .first()
        )
        if last:
            try:
                return int(last.split('-')[-1])
            except (ValueError, IndexError):
                return 0
        return 0

    def _print_preview(self, concepts):
        """Print a preview of what would be imported."""
        self.stdout.write(self.style.WARNING('\nDRY RUN — no data written\n'))

        # Show first 15 concepts
        for i, (key, data) in enumerate(concepts.items()):
            if i >= 15:
                break
            refs = data['references'].values()
            refs_str = ', '.join(
                f"{r['projectname']}: ${r['unitprice']}"
                for r in list(refs)[:4]
            )
            extra = f" (+{len(refs) - 4} more)" if len(refs) > 4 else ""
            self.stdout.write(
                f"  [{data['unit']}] {data['description'][:80]}..."
                f"\n    -> {len(refs)} refs: {refs_str}{extra}"
            )

        remaining = len(concepts) - 15
        if remaining > 0:
            self.stdout.write(f'\n  ... and {remaining} more concepts')

        # Check dedup against existing DB
        existing_count = ConceptPriceCatalogItem.objects.filter(
            source=CatalogSourceCode.HISTORICO, statecode=0
        ).count()
        if existing_count:
            self.stdout.write(self.style.WARNING(
                f'\n  Note: {existing_count} HISTORICO items already in DB '
                f'(duplicates will be skipped on import)'
            ))
