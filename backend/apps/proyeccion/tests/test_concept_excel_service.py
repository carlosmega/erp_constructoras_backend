"""Tests for ConceptExcelService (8-column round-trip Excel export/import)."""
import io
import pytest
import openpyxl

from apps.proyeccion.tests.factories import (
    EstimationProjectFactory,
    ConceptFamilyFactory,
    ConceptSubfamilyFactory,
    BudgetConceptFactory,
)
from apps.users.tests.factories import SystemUserFactory


# ---------------------------------------------------------------------------
# Export tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
@pytest.mark.unit
def test_export_returns_bytes():
    """export() returns non-empty bytes that openpyxl can open."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    project = EstimationProjectFactory()

    result = ConceptExcelService.export(project.estimationprojectid, user)

    assert isinstance(result, bytes)
    assert len(result) > 0
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert 'Conceptos' in wb.sheetnames


@pytest.mark.django_db
@pytest.mark.unit
def test_export_headers_in_row3():
    """Row 3 contains the 8 expected column headers."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    project = EstimationProjectFactory()

    result = ConceptExcelService.export(project.estimationprojectid, user)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb['Conceptos']
    headers = [ws.cell(row=3, column=c).value for c in range(1, 9)]
    assert headers == [
        'FAMILIA', 'COD.FAM', 'SUBFAMILIA', 'COD.SUB',
        'CODIGO', 'DESCRIPCION COMPLETA', 'UNIDAD', 'CANTIDAD',
    ]


@pytest.mark.django_db
@pytest.mark.unit
def test_export_empty_project_has_example_rows():
    """Empty project export includes example rows starting at row 4."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    project = EstimationProjectFactory()

    result = ConceptExcelService.export(project.estimationprojectid, user)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb['Conceptos']
    # At least one example row after the header
    assert ws.cell(row=4, column=5).value is not None  # CODIGO column not empty


@pytest.mark.django_db
@pytest.mark.unit
def test_export_with_concepts_writes_data_rows():
    """Export with concepts writes their data starting at row 4."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    family = ConceptFamilyFactory(name='GABINETE', code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, name='Proyecto Ejecutivo', code='GAB-01')
    concept = BudgetConceptFactory(
        subfamilyid=sf, code='A1', description='Desc test', unit='EST', quantity=1,
    )

    result = ConceptExcelService.export(sf.projectid.estimationprojectid, user)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb['Conceptos']
    assert ws.cell(row=4, column=1).value == 'GABINETE'   # FAMILIA
    assert ws.cell(row=4, column=2).value == 'GAB'         # COD.FAM
    assert ws.cell(row=4, column=3).value == 'Proyecto Ejecutivo'  # SUBFAMILIA
    assert ws.cell(row=4, column=4).value == 'GAB-01'      # COD.SUB
    assert ws.cell(row=4, column=5).value == 'A1'          # CODIGO
    assert ws.cell(row=4, column=6).value == 'Desc test'   # DESCRIPCION
    assert ws.cell(row=4, column=7).value == 'EST'         # UNIDAD
    assert ws.cell(row=4, column=8).value == 1.0           # CANTIDAD


# ---------------------------------------------------------------------------
# Analyze tests
# ---------------------------------------------------------------------------

def _make_excel(rows: list[tuple]) -> io.BytesIO:
    """Helper: build a minimal valid 8-column Excel in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='Proyecto Ejemplo')
    ws.cell(row=2, column=1, value='some-uuid')
    headers = ['FAMILIA', 'COD.FAM', 'SUBFAMILIA', 'COD.SUB',
               'CODIGO', 'DESCRIPCION COMPLETA', 'UNIDAD', 'CANTIDAD']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    for row_offset, data in enumerate(rows, 4):
        for col, val in enumerate(data, 1):
            ws.cell(row=row_offset, column=col, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.mark.django_db
@pytest.mark.unit
def test_analyze_classifies_new_row():
    """A row whose code doesn't exist yet is classified as 'new'."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    family = ConceptFamilyFactory(name='GABINETE', code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, name='Proy. Ejecutivo', code='GAB-01')
    project = sf.projectid

    buf = _make_excel([
        ('GABINETE', 'GAB', 'Proy. Ejecutivo', 'GAB-01', 'A1', 'Descripcion', 'M2', 100),
    ])

    result = ConceptExcelService.analyze(project.estimationprojectid, buf, user)

    assert result['summary']['new'] == 1
    assert result['summary']['skip'] == 0
    assert result['summary']['error'] == 0
    assert result['rows'][0]['status'] == 'new'


@pytest.mark.django_db
@pytest.mark.unit
def test_analyze_classifies_existing_code_as_skip():
    """A row whose code exists with identical values is classified as 'skip'."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    family = ConceptFamilyFactory(name='GABINETE', code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, name='Proy. Ejecutivo', code='GAB-01')
    project = sf.projectid
    # Create with same values as the Excel row so it is truly identical
    BudgetConceptFactory(subfamilyid=sf, code='A1', description='Descripcion', unit='M2', quantity=100)

    buf = _make_excel([
        ('GABINETE', 'GAB', 'Proy. Ejecutivo', 'GAB-01', 'A1', 'Descripcion', 'M2', 100),
    ])

    result = ConceptExcelService.analyze(project.estimationprojectid, buf, user)

    assert result['summary']['skip'] == 1
    assert result['rows'][0]['status'] == 'skip'


@pytest.mark.django_db
@pytest.mark.unit
def test_analyze_classifies_changed_concept_as_update():
    """A row whose code exists but with different values is classified as 'update'."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    family = ConceptFamilyFactory(name='GABINETE', code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, name='Proy. Ejecutivo', code='GAB-01')
    project = sf.projectid
    BudgetConceptFactory(subfamilyid=sf, code='A1', description='Viejo', unit='M2', quantity=50)

    buf = _make_excel([
        ('GABINETE', 'GAB', 'Proy. Ejecutivo', 'GAB-01', 'A1', 'Nuevo', 'M2', 100),
    ])

    result = ConceptExcelService.analyze(project.estimationprojectid, buf, user)

    assert result['summary']['update'] == 1
    assert result['rows'][0]['status'] == 'update'
    assert result['rows'][0]['old_description'] == 'Viejo'
    assert result['rows'][0]['old_quantity'] == 50.0


@pytest.mark.django_db
@pytest.mark.unit
def test_analyze_classifies_unknown_codsub_as_new():
    """A row with an unknown COD.SUB is classified as 'new' (subfamily auto-created on import)."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    project = EstimationProjectFactory()

    buf = _make_excel([
        ('GABINETE', 'GAB', 'Proy. Ejecutivo', 'GAB-99', 'A1', 'Desc', 'M2', 1),
    ])

    result = ConceptExcelService.analyze(project.estimationprojectid, buf, user)

    assert result['summary']['new'] == 1
    assert result['summary']['error'] == 0
    assert result['rows'][0]['status'] == 'new'


@pytest.mark.django_db
@pytest.mark.unit
def test_analyze_classifies_blank_codsub_as_error():
    """A row with no COD.SUB and no COD.FAM is classified as 'error'."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    project = EstimationProjectFactory()

    buf = _make_excel([
        ('', '', '', '', 'A1', 'Desc sin subfamilia', 'M2', 1),
    ])

    result = ConceptExcelService.analyze(project.estimationprojectid, buf, user)

    assert result['summary']['error'] == 1
    assert result['rows'][0]['status'] == 'error'


@pytest.mark.django_db
@pytest.mark.unit
def test_analyze_inherits_subfamily_columns_when_blank():
    """When FAMILIA/COD.FAM/SUBFAMILIA/COD.SUB are blank, inherit from prior row."""
    from apps.proyeccion.services import ConceptExcelService
    user = SystemUserFactory()
    family = ConceptFamilyFactory(name='GABINETE', code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, code='GAB-01')
    project = sf.projectid

    buf = _make_excel([
        ('GABINETE', 'GAB', 'Proy. Ej.', 'GAB-01', 'A1', 'Desc 1', 'M2', 100),
        ('', '', '', '', 'A2', 'Desc 2', 'M2', 200),  # blank first 4 cols
    ])

    result = ConceptExcelService.analyze(project.estimationprojectid, buf, user)

    assert result['summary']['total'] == 2
    assert result['rows'][1]['cod_sub'] == 'GAB-01'
    assert result['rows'][1]['status'] == 'new'


# ---------------------------------------------------------------------------
# Import tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
@pytest.mark.unit
def test_import_creates_new_concepts():
    """import_() creates BudgetConcept rows for each item in payload."""
    from apps.proyeccion.services import ConceptExcelService
    from apps.proyeccion.models import BudgetConcept
    from apps.proyeccion.schemas import ImportConceptExcelRequestDto

    user = SystemUserFactory()
    family = ConceptFamilyFactory(code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, code='GAB-01')
    project = sf.projectid

    payload = ImportConceptExcelRequestDto(items=[
        {'row': 4, 'familia': 'GABINETE', 'cod_fam': 'GAB', 'subfamilia': 'Proy. Ej.', 'cod_sub': 'GAB-01', 'codigo': 'A1', 'description': 'Desc', 'unit': 'M2', 'quantity': 100.0, 'status': 'new'},
        {'row': 5, 'familia': 'GABINETE', 'cod_fam': 'GAB', 'subfamilia': 'Proy. Ej.', 'cod_sub': 'GAB-01', 'codigo': 'A2', 'description': 'Desc 2', 'unit': 'KG', 'quantity': 50.0, 'status': 'new'},
    ])

    result = ConceptExcelService.import_(project.estimationprojectid, payload, user)

    assert result['created'] == 2
    assert result['updated'] == 0
    assert result['skipped'] == 0
    assert BudgetConcept.objects.filter(projectid=project, statecode=0).count() == 2


@pytest.mark.django_db
@pytest.mark.unit
def test_import_autocreates_family_and_subfamily():
    """import_() auto-creates ConceptFamily and ConceptSubfamily when they don't exist."""
    from apps.proyeccion.services import ConceptExcelService
    from apps.proyeccion.models import BudgetConcept, ConceptFamily, ConceptSubfamily
    from apps.proyeccion.schemas import ImportConceptExcelRequestDto

    user = SystemUserFactory()
    project = EstimationProjectFactory()

    payload = ImportConceptExcelRequestDto(items=[
        {'row': 4, 'familia': 'GABINETE', 'cod_fam': 'GAB', 'subfamilia': 'Proy. Ejecutivo', 'cod_sub': 'GAB-01', 'codigo': 'A1', 'description': 'Desc', 'unit': 'M2', 'quantity': 1.0, 'status': 'new'},
    ])

    result = ConceptExcelService.import_(project.estimationprojectid, payload, user)

    assert result['created'] == 1
    assert ConceptFamily.objects.filter(projectid=project, code='GAB').exists()
    assert ConceptSubfamily.objects.filter(projectid=project, code='GAB-01').exists()
    assert BudgetConcept.objects.filter(projectid=project, code='A1').exists()


@pytest.mark.django_db
@pytest.mark.unit
def test_import_updates_existing_concept():
    """import_() updates description/unit/quantity for status='update' items."""
    from apps.proyeccion.services import ConceptExcelService
    from apps.proyeccion.models import BudgetConcept
    from apps.proyeccion.schemas import ImportConceptExcelRequestDto

    user = SystemUserFactory()
    family = ConceptFamilyFactory(code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, code='GAB-01')
    project = sf.projectid
    BudgetConceptFactory(subfamilyid=sf, code='A1', description='Viejo', unit='M2', quantity=50)

    payload = ImportConceptExcelRequestDto(items=[
        {'row': 4, 'familia': 'GABINETE', 'cod_fam': 'GAB', 'subfamilia': 'Proy. Ej.', 'cod_sub': 'GAB-01', 'codigo': 'A1', 'description': 'Nuevo', 'unit': 'KG', 'quantity': 200.0, 'status': 'update'},
    ])

    result = ConceptExcelService.import_(project.estimationprojectid, payload, user)

    assert result['updated'] == 1
    assert result['created'] == 0
    concept = BudgetConcept.objects.get(projectid=project, code='A1')
    assert concept.description == 'Nuevo'
    assert concept.unit == 'KG'
    assert float(concept.quantity) == 200.0


@pytest.mark.django_db
@pytest.mark.unit
def test_import_autogenerates_code_when_blank():
    """import_() auto-generates a code when the codigo field is empty."""
    from apps.proyeccion.services import ConceptExcelService
    from apps.proyeccion.models import BudgetConcept
    from apps.proyeccion.schemas import ImportConceptExcelRequestDto

    user = SystemUserFactory()
    family = ConceptFamilyFactory(code='GAB')
    sf = ConceptSubfamilyFactory(familyid=family, code='GAB-01')
    project = sf.projectid

    payload = ImportConceptExcelRequestDto(items=[
        {'row': 4, 'familia': 'GABINETE', 'cod_fam': 'GAB', 'subfamilia': 'Proy. Ej.', 'cod_sub': 'GAB-01', 'codigo': '', 'description': 'Sin codigo', 'unit': 'M2', 'quantity': 1.0, 'status': 'new'},
    ])

    ConceptExcelService.import_(project.estimationprojectid, payload, user)

    concept = BudgetConcept.objects.get(projectid=project)
    assert concept.code  # not empty
    assert 'GAB-01' in concept.code  # uses the subfamily code as prefix
