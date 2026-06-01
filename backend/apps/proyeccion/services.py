"""Budget estimation (proyeccion) business logic service layer."""

from typing import Optional
from uuid import UUID
from decimal import Decimal
from datetime import datetime
from collections import defaultdict
from django.db import models, transaction
from django.db.models import QuerySet, Max, Sum, Q, F
from django.utils import timezone

from apps.proyeccion.models import (
    EstimationProject,
    EstimationStateCode,
    ConceptFamily,
    ConceptSubfamily,
    BudgetConcept,
    UnitCostBreakdown,
    IndirectCostDetail,
    IndirectCostTemplate,
    OfferAlternative,
    ExternalCostItem,
    SupplyCatalogItem,
    EquipmentYield,
    WorkPlanEntry,
    WorkPlanEntryType,
    BreakdownCategoryCode,
    ChecklistStatusCode,
    ProjectSizeCode,
    ConceptPriceCatalogItem,
    ConceptPriceReference,
    CatalogSourceCode,
    FamilyTemplateSet,
    FamilyTemplateItem,
    EstimationFinancialSettings,
    EstimationBillingRule,
)
from apps.proyeccion.schemas import (
    CreateConceptFamilyDto,
    UpdateConceptFamilyDto,
    CreateConceptSubfamilyDto,
    UpdateConceptSubfamilyDto,
    CreateBudgetConceptDto,
    UpdateBudgetConceptDto,
    CreateUnitCostBreakdownDto,
    UpdateUnitCostBreakdownDto,
    CreateIndirectCostDetailDto,
    UpdateIndirectCostDetailDto,
    CreateOfferAlternativeDto,
    UpdateOfferAlternativeDto,
    UpdateExternalCostItemDto,
    CreateSupplyCatalogItemDto,
    UpdateSupplyCatalogItemDto,
    CreateEquipmentYieldDto,
    UpdateEquipmentYieldDto,
    CreateWorkPlanEntryDto,
    UpdateWorkPlanEntryDto,
    SaveProjectAsTemplateDto,
    ApplyFamilyTemplateDto,
)
from core.exceptions import ValidationError, NotFound


class EstimationProjectService:
    """Service for EstimationProject CRUD and conversion to ConstructionProject."""

    @staticmethod
    def list_projects(user, statecode=None, search=None):
        """List estimation projects with optional filters."""
        qs = EstimationProject.objects.select_related('accountid', 'ownerid', 'createdby', 'modifiedby')
        if statecode is not None:
            qs = qs.filter(statecode=statecode)
        else:
            qs = qs.exclude(statecode=EstimationStateCode.CANCELED)
        if search:
            qs = qs.filter(models.Q(name__icontains=search) | models.Q(estimationnumber__icontains=search))
        return qs

    @staticmethod
    def get_project(project_id, user):
        """Get a single estimation project by ID."""
        try:
            return EstimationProject.objects.select_related(
                'accountid', 'ownerid', 'generatedprojectid', 'createdby', 'modifiedby'
            ).get(estimationprojectid=project_id)
        except EstimationProject.DoesNotExist:
            raise NotFound(f"Estimation project with ID {project_id} not found")

    @staticmethod
    def create_project(dto, user):
        """Create a new estimation project with auto-generated number (EST-YYYY-NNN)."""
        year = datetime.now().year
        max_num = EstimationProject.objects.filter(
            estimationnumber__startswith=f'EST-{year}-'
        ).count()
        next_num = max_num + 1
        estimation_number = f'EST-{year}-{next_num:03d}'

        project = EstimationProject(
            estimationnumber=estimation_number,
            name=dto.name,
            description=dto.description,
            accountid_id=dto.accountid,
            opportunityid_id=dto.opportunityid,
            presentationdate=dto.presentationdate,
            estimatedstartdate=dto.estimatedstartdate,
            estimatedenddate=dto.estimatedenddate,
            durationmonths=dto.durationmonths or 0,
            projecttype=dto.projecttype or 0,
            biddingtype=dto.biddingtype or 0,
            periodtype=dto.periodtype or 0,
            estimatedcontractamount=dto.estimatedcontractamount or 0,
            exchangerate_mxn_usd=dto.exchangerate_mxn_usd,
            ownerid=user,
            createdby=user,
            modifiedby=user,
        )
        project.save()
        return project

    @staticmethod
    def update_project(project_id, dto, user):
        """Update an estimation project. Cannot update if already converted.

        Changing ``profitpercent`` triggers a re-prorate of indirect costs so
        every concept picks up the new utility rate.
        """
        project = EstimationProjectService.get_project(project_id, user)

        if project.statecode == EstimationStateCode.CONVERTED:
            raise AlreadyConvertedError(project.generatedprojectid_id)

        update_fields = [
            'name', 'description', 'presentationdate', 'estimatedstartdate',
            'estimatedenddate', 'durationmonths', 'projecttype', 'biddingtype',
            'periodtype', 'estimatedcontractamount', 'exchangerate_mxn_usd',
            'profitpercent', 'statecode',
        ]
        old_profitpercent = project.profitpercent
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(project, field, value)

        # Handle FK fields separately
        if dto.accountid is not None:
            project.accountid_id = dto.accountid
        if dto.opportunityid is not None:
            project.opportunityid_id = dto.opportunityid

        project.modifiedby = user
        project.save()

        # When profitpercent changes, re-prorate so utility ripples to every concept.
        if project.profitpercent != old_profitpercent:
            IndirectCostDetailService.prorate_to_concepts(project.estimationprojectid, user)

        return project

    @staticmethod
    def delete_project(project_id, user):
        """Soft-delete an estimation project (set to Canceled). Cannot delete if converted."""
        project = EstimationProjectService.get_project(project_id, user)
        if project.statecode == EstimationStateCode.CONVERTED:
            raise AlreadyConvertedError(project.generatedprojectid_id)
        project.statecode = EstimationStateCode.CANCELED
        project.modifiedby = user
        project.save()
        return project

class AlreadyConvertedError(Exception):
    """Raised when the caller tries to convert an estimation that is already CONVERTED.

    Carries the existing ``projectid`` so the API layer can return 409 with the link.
    """

    def __init__(self, projectid):
        self.projectid = projectid
        super().__init__(f"Estimation already converted (project={projectid})")


class EstimationConversionService:
    """Convert an accepted EstimationProject into a ConstructionProject.

    Replaces the legacy ``EstimationProjectService.convert_to_project`` with a
    flow that pulls everything from the estimation (no inputs from the caller).

    Spec: docs/superpowers/specs/2026-05-17-conversion-estudio-proyecto-design.md
    """

    @staticmethod
    @transaction.atomic
    def convert(estimation_id: UUID, *, user):
        """Convert the given estimation to a construction project.

        Validates every pre-condition listed in spec §6.1 then creates the
        ConstructionProject with header fields pulled from the estimation
        (OfferAlternative.chosen for amounts, EstimationFinancialSettings for
        anticipo).
        """
        from apps.budgets.models import CostCategory, CostTypeCode
        from apps.budgets.services import DEFAULT_INDIRECT_CATEGORIES
        from apps.projects.models import ConstructionProject, ProjectStateCode, ProjectZone
        from apps.projects.services import ProjectService

        try:
            estimation = EstimationProject.objects.select_for_update().get(
                estimationprojectid=estimation_id,
            )
        except EstimationProject.DoesNotExist:
            raise NotFound(f"Estimation project with ID {estimation_id} not found")

        # Idempotency: if already converted, signal the existing projectid.
        if estimation.statecode == EstimationStateCode.CONVERTED:
            raise AlreadyConvertedError(estimation.generatedprojectid_id)

        EstimationConversionService._validate_preconditions(estimation)

        chosen = OfferAlternative.objects.get(projectid=estimation, ischosen=True)
        settings = getattr(estimation, 'financial_settings', None)
        advance_notax = settings.advanceamountnotax if settings else Decimal('0')
        advance_withtax = advance_notax * Decimal('1.16')

        project = ConstructionProject.objects.create(
            projectnumber=ProjectService.generate_project_number(),
            name=estimation.name,
            description=estimation.description,
            statecode=ProjectStateCode.DRAFT,
            accountid=estimation.accountid,
            opportunityid=estimation.opportunityid,
            presentationdate=estimation.presentationdate,
            awarddate=timezone.now().date(),
            startdate=estimation.estimatedstartdate,
            contractenddate=estimation.estimatedenddate,
            durationmonths=estimation.durationmonths,
            projecttype=estimation.projecttype,
            biddingtype=estimation.biddingtype,
            periodtype=estimation.periodtype,
            contractamount_notax=chosen.salepricenet,
            contractamount_withtax=chosen.salepricetotal,
            advancepayment_notax=advance_notax,
            advancepayment_withtax=advance_withtax,
            exchangerate_mxn_usd=estimation.exchangerate_mxn_usd,
            ownerid=user,
            createdby=user,
            modifiedby=user,
        )

        default_zone = EstimationConversionService._seed_default_zone(project, user)
        family_to_category = EstimationConversionService._seed_direct_categories(estimation, project, user)
        indirect_categories = EstimationConversionService._seed_indirect_categories(project, user)
        period_by_sortorder = EstimationConversionService._initialize_periods(project, user)
        EstimationConversionService._seed_direct_imputation_codes_and_budgets(
            estimation, project, family_to_category, default_zone, period_by_sortorder, user,
        )
        EstimationConversionService._seed_indirect_imputation_codes_and_budgets(
            estimation, project, indirect_categories, period_by_sortorder, user,
        )

        # Lock the estimation: state → CONVERTED and link to the new project.
        estimation.statecode = EstimationStateCode.CONVERTED
        estimation.generatedprojectid = project
        estimation.modifiedby = user
        estimation.save(update_fields=['statecode', 'generatedprojectid', 'modifiedby', 'modifiedon'])

        return project

    @staticmethod
    def _initialize_periods(project, user):
        from apps.budgets.services import PeriodService
        periods = PeriodService.initialize_periods(project.projectid, user)
        return {p.sortorder: p for p in periods}

    @staticmethod
    def _seed_default_zone(project, user):
        from apps.projects.models import ProjectZone
        return ProjectZone.objects.create(
            projectid=project,
            prefix='GEN',
            name='General',
            sortorder=0,
            createdby=user,
            modifiedby=user,
        )

    @staticmethod
    def _seed_direct_categories(estimation, project, user):
        """Create 1 CostCategory direct per ConceptFamily, mapped P1..P10 (Opción A).

        Returns ``dict[familyid: UUID → CostCategory]`` so downstream code can
        assign each ImputationCode to the correct category.
        """
        from apps.budgets.models import CostCategory, CostTypeCode

        family_to_category = {}
        families = list(
            ConceptFamily.objects.filter(projectid=estimation).order_by('sortorder')
        )
        if not families:
            return family_to_category

        max_individual = 9
        head = families[:max_individual]
        tail = families[max_individual:]

        for idx, family in enumerate(head, start=1):
            cat = CostCategory.objects.create(
                projectid=project,
                costtype=CostTypeCode.DIRECT,
                code=f'P{idx}',
                name=family.name,
                sortorder=idx,
                createdby=user,
                modifiedby=user,
            )
            family_to_category[family.familyid] = cat

        if tail:
            cat = CostCategory.objects.create(
                projectid=project,
                costtype=CostTypeCode.DIRECT,
                code='P10',
                name='Otros',
                sortorder=10,
                createdby=user,
                modifiedby=user,
            )
            for family in tail:
                family_to_category[family.familyid] = cat

        return family_to_category

    @staticmethod
    def _seed_direct_imputation_codes_and_budgets(
        estimation, project, family_to_category, default_zone, period_by_sortorder, user,
    ):
        """For each BudgetConcept: create one direct ImputationCode and its per-period budget rows.

        plannedamount(period) = SUM(breakdown.amount × concept.quantity × fraction(breakdown, period))
        plannedvolume(period) = WorkPlanEntry.distributedquantity (entrytype=PLANNED)
        """
        from apps.budgets.models import ImputationCode, ImputationCodeBudget, CostTypeCode

        concepts = list(
            BudgetConcept.objects.filter(projectid=estimation)
            .select_related('subfamilyid', 'subfamilyid__familyid')
            .order_by('subfamilyid__familyid__sortorder', 'subfamilyid__sortorder', 'sequencenumber')
        )
        if not concepts:
            return

        # Pre-fetch CostDistribution for direct breakdowns (per concept).
        breakdown_dist = {
            (row['breakdownid_id'], row['periodnumber']): row['fraction']
            for row in CostDistribution.objects.filter(
                projectid=estimation, linetype=CostLineType.BREAKDOWN
            ).values('breakdownid_id', 'periodnumber', 'fraction')
        }
        # Pre-fetch workplan PLANNED entries.
        workplan = {
            (row['conceptid_id'], row['periodnumber']): row['distributedquantity']
            for row in WorkPlanEntry.objects.filter(
                projectid=estimation, entrytype=WorkPlanEntryType.PLANNED
            ).values('conceptid_id', 'periodnumber', 'distributedquantity')
        }

        # Track sequence counters per (category, zone) for code generation.
        seq_counters: dict = {}

        for concept in concepts:
            family_id = concept.subfamilyid.familyid_id
            category = family_to_category.get(family_id)
            if category is None:
                # No category mapped for this concept's family — skip defensively.
                continue

            key = (category.categoryid, default_zone.zoneid)
            seq_counters[key] = seq_counters.get(key, 0) + 1
            seq = seq_counters[key]
            code_str = f"{default_zone.prefix}-{category.code}-{seq}"

            imputation_code = ImputationCode.objects.create(
                projectid=project,
                categoryid=category,
                zoneid=default_zone,
                costtype=CostTypeCode.DIRECT,
                code=code_str,
                sequencenumber=seq,
                name=concept.description,
                unit=concept.unit,
                contractcode=concept.code,
                contractunitprice=concept.clientunitprice or concept.unitprice or None,
                sourceconceptid=concept,
                unitcost=concept.directunitcost + concept.indirectunitcost,
                quantity=concept.quantity,
                totalbudget=concept.totalamount,
                remainingbudget=concept.totalamount,
                createdby=user,
                modifiedby=user,
            )

            # Build per-period budget rows.
            breakdown_ids_for_concept = list(
                UnitCostBreakdown.objects.filter(conceptid=concept).values_list('breakdownid', 'amount')
            )
            concept_qty = concept.quantity or Decimal('0')

            budgets_to_create = []
            for sortorder, period in period_by_sortorder.items():
                planned_amount = Decimal('0')
                for bd_id, bd_amount in breakdown_ids_for_concept:
                    frac = breakdown_dist.get((bd_id, sortorder), Decimal('0'))
                    planned_amount += (bd_amount or Decimal('0')) * concept_qty * frac

                planned_volume = workplan.get((concept.conceptid, sortorder), Decimal('0'))

                budgets_to_create.append(ImputationCodeBudget(
                    imputationcodeid=imputation_code,
                    periodid=period,
                    periodlabel=period.label,
                    plannedamount=planned_amount,
                    plannedvolume=planned_volume,
                ))

            ImputationCodeBudget.objects.bulk_create(budgets_to_create)

    @staticmethod
    def _seed_indirect_categories(project, user):
        """Create the 8 standard indirect categories (C1-C8). Returns ``dict[code: str → CostCategory]``."""
        from apps.budgets.models import CostCategory
        from apps.budgets.services import DEFAULT_INDIRECT_CATEGORIES

        CostCategory.objects.bulk_create([
            CostCategory(
                projectid=project,
                costtype=costtype,
                code=code,
                name=name,
                sortorder=sortorder,
                createdby=user,
                modifiedby=user,
            )
            for code, name, costtype, sortorder in DEFAULT_INDIRECT_CATEGORIES
        ])
        return {
            c.code: c
            for c in CostCategory.objects.filter(projectid=project, costtype=1)
        }

    @staticmethod
    def _seed_indirect_imputation_codes_and_budgets(
        estimation, project, indirect_categories, period_by_sortorder, user,
    ):
        """For each IndirectCostDetail: create one indirect ImputationCode (zoneid=None) and its per-period budgets.

        plannedamount(period) = IndirectCostDetail.amount × CostDistribution.fraction(detail, period)
        """
        from apps.budgets.models import ImputationCode, ImputationCodeBudget, CostTypeCode

        details = list(
            IndirectCostDetail.objects.filter(projectid=estimation)
            .order_by('categorycode', 'linenumber')
        )
        if not details:
            return

        indirect_dist = {
            (row['indirectcostid_id'], row['periodnumber']): row['fraction']
            for row in CostDistribution.objects.filter(
                projectid=estimation, linetype=CostLineType.INDIRECT
            ).values('indirectcostid_id', 'periodnumber', 'fraction')
        }

        seq_counters: dict = {}

        for detail in details:
            category = indirect_categories.get(detail.categorycode)
            if category is None:
                # Skip silently if the detail's categorycode doesn't match a seeded C-category.
                continue

            seq_counters[category.code] = seq_counters.get(category.code, 0) + 1
            seq = seq_counters[category.code]
            code_str = f"{category.code}-{seq}"

            is_c1 = detail.categorycode == 'C1'
            imputation_code = ImputationCode.objects.create(
                projectid=project,
                categoryid=category,
                zoneid=None,
                costtype=CostTypeCode.INDIRECT,
                code=code_str,
                sequencenumber=seq,
                name=detail.description,
                personnelrole=detail.area if is_c1 else None,
                monthlycost=detail.monthlycost,
                units=detail.units,
                executionmonths=int(detail.months) if detail.months else None,
                totalbudget=detail.amount,
                remainingbudget=detail.amount,
                createdby=user,
                modifiedby=user,
            )

            budgets_to_create = []
            for sortorder, period in period_by_sortorder.items():
                frac = indirect_dist.get((detail.indirectcostid, sortorder), Decimal('0'))
                planned_amount = (detail.amount or Decimal('0')) * frac
                budgets_to_create.append(ImputationCodeBudget(
                    imputationcodeid=imputation_code,
                    periodid=period,
                    periodlabel=period.label,
                    plannedamount=planned_amount,
                ))
            ImputationCodeBudget.objects.bulk_create(budgets_to_create)

    @staticmethod
    def _validate_preconditions(estimation):
        if estimation.statecode != EstimationStateCode.ACCEPTED:
            raise ValidationError(
                f"Estimation must be in ACCEPTED state to convert (current: "
                f"{EstimationStateCode(estimation.statecode).label})"
            )
        if estimation.accountid_id is None:
            raise ValidationError(
                "Estimation has no account linked; cannot convert to a project"
            )
        if estimation.estimatedstartdate is None:
            raise ValidationError("Estimation is missing the estimated start date")
        if estimation.estimatedenddate is None:
            raise ValidationError("Estimation is missing the estimated end date")
        if not estimation.durationmonths or estimation.durationmonths <= 0:
            raise ValidationError("Estimation duration (months) must be greater than zero")

        chosen = OfferAlternative.objects.filter(
            projectid=estimation, ischosen=True
        ).first()
        if chosen is None:
            raise ValidationError(
                "Estimation has no chosen offer alternative — set one as ischosen=True before converting"
            )


# Default external cost checklist items
DEFAULT_EXTERNAL_COSTS = [
    'Fianza de anticipo',
    'Fianza de cumplimiento',
    'Fianza de vicios ocultos',
    'Seguro de responsabilidad civil',
    'Seguro de obra',
    'Seguro de equipo',
    'Gastos notariales',
    'Permisos de construccion',
    'Licencias ambientales',
    'Permisos municipales',
    'Derechos de via',
    'Estudios topograficos',
    'Estudios de mecanica de suelos',
    'Estudios de impacto ambiental',
    'Dictamen estructural',
    'Gastos de licitacion',
    'Supervision externa',
    'Laboratorio de control de calidad',
    'Gastos financieros',
    'Impuestos (ISR provisional)',
]


class ConceptCatalogService:
    """Service class for ConceptFamily, ConceptSubfamily, and BudgetConcept."""

    # -------------------------------------------------------------------------
    # Families
    # -------------------------------------------------------------------------

    @staticmethod
    def list_families(project_id: UUID, user) -> QuerySet[ConceptFamily]:
        """List all concept families for a project."""
        return ConceptFamily.objects.filter(
            projectid=project_id,
            statecode=0,
        ).select_related('createdby', 'modifiedby')

    @staticmethod
    def create_family(dto: CreateConceptFamilyDto, user) -> ConceptFamily:
        """Create a new concept family."""
        family = ConceptFamily(
            projectid_id=dto.projectid,
            name=dto.name,
            code=dto.code,
            sortorder=dto.sortorder or 0,
            createdby=user,
            modifiedby=user,
        )
        family.save()
        return family

    @staticmethod
    def update_family(family_id: UUID, dto: UpdateConceptFamilyDto, user) -> ConceptFamily:
        """Update an existing concept family."""
        try:
            family = ConceptFamily.objects.get(familyid=family_id)
        except ConceptFamily.DoesNotExist:
            raise NotFound(f"ConceptFamily with ID {family_id} not found")

        update_fields = ['name', 'code', 'sortorder', 'statecode']
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(family, field, value)

        family.modifiedby = user
        family.save()
        return family

    @staticmethod
    def delete_family(family_id: UUID, user) -> ConceptFamily:
        """Soft delete a concept family (statecode=1)."""
        try:
            family = ConceptFamily.objects.get(familyid=family_id)
        except ConceptFamily.DoesNotExist:
            raise NotFound(f"ConceptFamily with ID {family_id} not found")

        family.statecode = 1
        family.modifiedby = user
        family.save()
        ConceptSubfamily.objects.filter(familyid=family).update(statecode=1, modifiedby=user)
        return family

    # -------------------------------------------------------------------------
    # Subfamilies
    # -------------------------------------------------------------------------

    @staticmethod
    def list_subfamilies(family_id: UUID, user) -> QuerySet[ConceptSubfamily]:
        """List all subfamilies for a family."""
        return ConceptSubfamily.objects.filter(
            familyid=family_id,
            statecode=0,
        ).select_related('familyid', 'createdby', 'modifiedby')

    @staticmethod
    def create_subfamily(dto: CreateConceptSubfamilyDto, user) -> ConceptSubfamily:
        """Create a new concept subfamily."""
        # Resolve projectid from the parent family if not provided
        if not dto.projectid:
            try:
                family = ConceptFamily.objects.get(familyid=dto.familyid)
            except ConceptFamily.DoesNotExist:
                raise NotFound(f"Concept family with ID {dto.familyid} not found")
            dto.projectid = family.projectid_id

        subfamily = ConceptSubfamily(
            familyid_id=dto.familyid,
            projectid_id=dto.projectid,
            name=dto.name,
            code=dto.code,
            sortorder=dto.sortorder or 0,
            createdby=user,
            modifiedby=user,
        )
        subfamily.save()
        return subfamily

    @staticmethod
    def update_subfamily(subfamily_id: UUID, dto: UpdateConceptSubfamilyDto, user) -> ConceptSubfamily:
        """Update an existing concept subfamily."""
        try:
            subfamily = ConceptSubfamily.objects.get(subfamilyid=subfamily_id)
        except ConceptSubfamily.DoesNotExist:
            raise NotFound(f"ConceptSubfamily with ID {subfamily_id} not found")

        update_fields = ['name', 'code', 'sortorder', 'statecode']
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(subfamily, field, value)

        subfamily.modifiedby = user
        subfamily.save()
        return subfamily

    @staticmethod
    def delete_subfamily(subfamily_id: UUID, user) -> ConceptSubfamily:
        """Soft delete a concept subfamily (statecode=1)."""
        try:
            subfamily = ConceptSubfamily.objects.get(subfamilyid=subfamily_id)
        except ConceptSubfamily.DoesNotExist:
            raise NotFound(f"ConceptSubfamily with ID {subfamily_id} not found")

        subfamily.statecode = 1
        subfamily.modifiedby = user
        subfamily.save()
        return subfamily

    # -------------------------------------------------------------------------
    # Concepts
    # -------------------------------------------------------------------------

    @staticmethod
    def list_concepts(
        project_id: UUID,
        user,
        subfamilyid: Optional[UUID] = None,
    ) -> QuerySet[BudgetConcept]:
        """List all budget concepts for a project, optionally filtered by subfamily."""
        queryset = BudgetConcept.objects.filter(projectid=project_id, statecode=0)

        if subfamilyid is not None:
            queryset = queryset.filter(subfamilyid=subfamilyid)

        return queryset.select_related(
            'subfamilyid', 'subfamilyid__familyid', 'createdby', 'modifiedby'
        )

    @staticmethod
    def get_concept(concept_id: UUID, user) -> BudgetConcept:
        """Get a single budget concept by ID."""
        try:
            return BudgetConcept.objects.select_related(
                'subfamilyid', 'subfamilyid__familyid', 'createdby', 'modifiedby'
            ).get(conceptid=concept_id)
        except BudgetConcept.DoesNotExist:
            raise NotFound(f"BudgetConcept with ID {concept_id} not found")

    @staticmethod
    def create_concept(dto: CreateBudgetConceptDto, user) -> BudgetConcept:
        """Create a new budget concept with auto-generated code and sequence number."""
        # Fetch the subfamily and family for code generation
        try:
            subfamily = ConceptSubfamily.objects.select_related('familyid').get(
                subfamilyid=dto.subfamilyid
            )
        except ConceptSubfamily.DoesNotExist:
            raise ValidationError(f"ConceptSubfamily with ID {dto.subfamilyid} not found")

        family = subfamily.familyid

        # Calculate next sequence number within this subfamily
        max_seq = BudgetConcept.objects.filter(
            subfamilyid=dto.subfamilyid,
        ).aggregate(max_seq=Max('sequencenumber'))['max_seq'] or 0
        next_seq = max_seq + 1

        # Auto-generate code: F{family_code}-S{subfamily_code}-{seq}
        code = f"F{family.code}-S{subfamily.code}-{next_seq}"

        concept = BudgetConcept(
            projectid_id=dto.projectid,
            subfamilyid_id=dto.subfamilyid,
            code=code,
            sequencenumber=next_seq,
            description=dto.description,
            unit=dto.unit,
            quantity=dto.quantity,
            directunitcost=Decimal('0'),
            indirectunitcost=Decimal('0'),
            utilityunitcost=Decimal('0'),
            unitprice=Decimal('0'),
            totalamount=Decimal('0'),
            clientunitprice=dto.clientunitprice,
            breakdownmethod=dto.breakdownmethod or 0,
            isprintable=dto.isprintable if dto.isprintable is not None else True,
            createdby=user,
            modifiedby=user,
        )
        concept.save()
        return concept

    @staticmethod
    def update_concept(concept_id: UUID, dto: UpdateBudgetConceptDto, user) -> BudgetConcept:
        """Update an existing budget concept."""
        try:
            concept = BudgetConcept.objects.get(conceptid=concept_id)
        except BudgetConcept.DoesNotExist:
            raise NotFound(f"BudgetConcept with ID {concept_id} not found")

        update_fields = [
            'description', 'unit', 'quantity', 'directunitcost', 'indirectunitcost',
            'utilityunitcost', 'unitprice', 'totalamount', 'clientunitprice',
            'breakdownmethod', 'isprintable', 'statecode',
        ]
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(concept, field, value)

        concept.modifiedby = user
        concept.save()
        return concept

    @staticmethod
    def delete_concept(concept_id: UUID, user) -> BudgetConcept:
        """Soft delete a budget concept (statecode=1)."""
        try:
            concept = BudgetConcept.objects.get(conceptid=concept_id)
        except BudgetConcept.DoesNotExist:
            raise NotFound(f"BudgetConcept with ID {concept_id} not found")

        concept.statecode = 1
        concept.modifiedby = user
        concept.save()
        return concept

    @staticmethod
    def recalculate_concept(concept_id: UUID, user) -> BudgetConcept:
        """Recalculate concept costs from its unit cost breakdowns.

        Sums breakdown amounts by category to get directunitcost, then computes:
        - utilityunitcost = (directunitcost + indirectunitcost) * project.profitpercent / 100
        - unitprice = directunitcost + indirectunitcost + utilityunitcost
        - totalamount = unitprice * quantity

        Mirrors the Excel "E7 Fase Estudio" formulas K = (I+J)*K2, L = I+J+K, M = H*L.
        """
        try:
            concept = BudgetConcept.objects.select_related('projectid').get(conceptid=concept_id)
        except BudgetConcept.DoesNotExist:
            raise NotFound(f"BudgetConcept with ID {concept_id} not found")

        # Sum all active breakdown amounts grouped by category
        breakdowns = UnitCostBreakdown.objects.filter(
            conceptid=concept_id,
            statecode=0,
        )

        total_direct = breakdowns.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')

        concept.directunitcost = total_direct
        profitpercent = concept.projectid.profitpercent or Decimal('0')
        concept.utilityunitcost = (
            (concept.directunitcost + concept.indirectunitcost) * profitpercent / Decimal('100')
        )
        concept.unitprice = concept.directunitcost + concept.indirectunitcost + concept.utilityunitcost
        concept.totalamount = concept.unitprice * concept.quantity

        concept.modifiedby = user
        concept.save()
        return concept


class UnitCostBreakdownService:
    """Service class for UnitCostBreakdown business logic."""

    @staticmethod
    def _recalc_concept(concept_id: UUID, user) -> None:
        """Recompute concept totals after a breakdown mutation.

        Mirrors the Excel "E7 Fase Estudio" cascade:
            I = directunitcost (Σ breakdown amounts)
            K = (I + J) * project.profitpercent / 100
            L = I + J + K
            M = H * L
        """
        try:
            concept = BudgetConcept.objects.select_related('projectid').get(conceptid=concept_id)
        except BudgetConcept.DoesNotExist:
            return

        total_direct = UnitCostBreakdown.objects.filter(
            conceptid=concept_id,
            statecode=0,
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        concept.directunitcost = total_direct
        profitpercent = concept.projectid.profitpercent or Decimal('0')
        concept.utilityunitcost = (
            (concept.directunitcost + concept.indirectunitcost) * profitpercent / Decimal('100')
        )
        concept.unitprice = (
            concept.directunitcost + concept.indirectunitcost + concept.utilityunitcost
        )
        concept.totalamount = concept.unitprice * concept.quantity
        concept.modifiedby = user
        concept.save(update_fields=[
            'directunitcost', 'utilityunitcost', 'unitprice', 'totalamount',
            'modifiedby', 'modifiedon',
        ])

    @staticmethod
    def list_breakdowns(concept_id: UUID, user) -> QuerySet[UnitCostBreakdown]:
        """List all active breakdown lines for a concept.

        Excludes soft-deleted lines (``statecode=1``); otherwise a deleted line
        keeps reappearing in the UI after refetch.
        """
        return UnitCostBreakdown.objects.filter(
            conceptid=concept_id,
            statecode=0,
        ).select_related('supplyid')

    @staticmethod
    def create_breakdown(dto: CreateUnitCostBreakdownDto, user) -> UnitCostBreakdown:
        """Create a new breakdown line with computed amount and auto linenumber."""
        # Validate categorycode
        if dto.categorycode not in [c.value for c in BreakdownCategoryCode]:
            raise ValidationError(f"Invalid category code: {dto.categorycode}")

        # Calculate next linenumber within this concept and category
        max_line = UnitCostBreakdown.objects.filter(
            conceptid=dto.conceptid,
            categorycode=dto.categorycode,
        ).aggregate(max_line=Max('linenumber'))['max_line'] or 0
        next_line = max_line + 1

        # Compute amount = quantity * unitprice * yieldvalue
        quantity = dto.quantity or Decimal('0')
        unitprice = dto.unitprice or Decimal('0')
        yieldvalue = dto.yieldvalue or Decimal('1')
        amount = quantity * unitprice * yieldvalue

        breakdown = UnitCostBreakdown(
            conceptid_id=dto.conceptid,
            categorycode=dto.categorycode,
            linenumber=next_line,
            description=dto.description,
            unit=dto.unit,
            quantity=quantity,
            unitprice=unitprice,
            yieldvalue=yieldvalue,
            amount=amount,
            supplyid_id=dto.supplyid,
        )
        breakdown.save()
        UnitCostBreakdownService._recalc_concept(dto.conceptid, user)
        return breakdown

    @staticmethod
    def bulk_create_breakdowns(
        concept_id: UUID, dtos: list[CreateUnitCostBreakdownDto], user
    ) -> list[UnitCostBreakdown]:
        """Create multiple breakdown lines for a concept in a single transaction."""
        from django.db import transaction

        if not dtos:
            return []

        valid_codes = {c.value for c in BreakdownCategoryCode}
        for dto in dtos:
            if dto.categorycode not in valid_codes:
                raise ValidationError(f"Invalid category code: {dto.categorycode}")

        # Pre-compute next line numbers per category to avoid repeated queries
        max_lines = {}
        existing = (
            UnitCostBreakdown.objects.filter(conceptid=concept_id)
            .values('categorycode')
            .annotate(max_line=Max('linenumber'))
        )
        for row in existing:
            max_lines[row['categorycode']] = row['max_line'] or 0

        created: list[UnitCostBreakdown] = []
        with transaction.atomic():
            for dto in dtos:
                next_line = max_lines.get(dto.categorycode, 0) + 1
                max_lines[dto.categorycode] = next_line

                quantity = dto.quantity or Decimal('0')
                unitprice = dto.unitprice or Decimal('0')
                yieldvalue = dto.yieldvalue or Decimal('1')
                amount = quantity * unitprice * yieldvalue

                breakdown = UnitCostBreakdown(
                    conceptid_id=concept_id,
                    categorycode=dto.categorycode,
                    linenumber=next_line,
                    description=dto.description,
                    unit=dto.unit,
                    quantity=quantity,
                    unitprice=unitprice,
                    yieldvalue=yieldvalue,
                    amount=amount,
                    supplyid_id=dto.supplyid,
                )
                breakdown.save()
                created.append(breakdown)

        UnitCostBreakdownService._recalc_concept(concept_id, user)
        return created

    @staticmethod
    def update_breakdown(breakdown_id: UUID, dto: UpdateUnitCostBreakdownDto, user) -> UnitCostBreakdown:
        """Update a breakdown line and recompute amount."""
        try:
            breakdown = UnitCostBreakdown.objects.get(breakdownid=breakdown_id)
        except UnitCostBreakdown.DoesNotExist:
            raise NotFound(f"UnitCostBreakdown with ID {breakdown_id} not found")

        update_fields = [
            'categorycode', 'description', 'unit', 'quantity',
            'unitprice', 'yieldvalue', 'statecode',
        ]
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(breakdown, field, value)

        # Handle supplyid FK separately (can be set to None)
        if dto.supplyid is not None:
            breakdown.supplyid_id = dto.supplyid

        # Recompute amount
        breakdown.amount = breakdown.quantity * breakdown.unitprice * breakdown.yieldvalue

        breakdown.save()
        UnitCostBreakdownService._recalc_concept(breakdown.conceptid_id, user)
        return breakdown

    @staticmethod
    def delete_breakdown(breakdown_id: UUID, user) -> UnitCostBreakdown:
        """Soft delete a breakdown line (statecode=1)."""
        try:
            breakdown = UnitCostBreakdown.objects.get(breakdownid=breakdown_id)
        except UnitCostBreakdown.DoesNotExist:
            raise NotFound(f"UnitCostBreakdown with ID {breakdown_id} not found")

        breakdown.statecode = 1
        breakdown.save()
        UnitCostBreakdownService._recalc_concept(breakdown.conceptid_id, user)
        return breakdown

    @staticmethod
    @transaction.atomic
    def auto_generate_hm_epp(concept_id: UUID, user) -> list[UnitCostBreakdown]:
        """Auto-create HerramientaMenor + EPP lines at 3% of labor cost.

        If category 4 (Labor) breakdowns exist, add:
        - Category 6 (Minor Tools) at 3% of total labor cost
        - Category 7 (PPE) at 3% of total labor cost
        """
        # Check if labor breakdowns exist (category 4)
        labor_total = UnitCostBreakdown.objects.filter(
            conceptid=concept_id,
            categorycode=BreakdownCategoryCode.LABOR,
            statecode=0,
        ).aggregate(total=Sum('amount'))['total']

        if not labor_total or labor_total <= 0:
            return []

        hm_epp_amount = labor_total * Decimal('0.03')
        created = []

        for cat_code, cat_desc in [
            (BreakdownCategoryCode.MINOR_TOOLS, 'Herramienta Menor (3% M.O.)'),
            (BreakdownCategoryCode.PPE, 'EPP (3% M.O.)'),
        ]:
            # Calculate next linenumber
            max_line = UnitCostBreakdown.objects.filter(
                conceptid=concept_id,
                categorycode=cat_code,
            ).aggregate(max_line=Max('linenumber'))['max_line'] or 0
            next_line = max_line + 1

            breakdown = UnitCostBreakdown(
                conceptid_id=concept_id,
                categorycode=cat_code,
                linenumber=next_line,
                description=cat_desc,
                unit='%',
                quantity=Decimal('1'),
                unitprice=hm_epp_amount,
                yieldvalue=Decimal('1'),
                amount=hm_epp_amount,
            )
            breakdown.save()
            created.append(breakdown)

        UnitCostBreakdownService._recalc_concept(concept_id, user)
        return created

    @staticmethod
    def regenerate_hm_epp(concept_id, user):
        """Regenerate Herramienta Menor (3% labor) and EPP (3% labor) lines.

        Deletes any existing HM/EPP lines for the concept and creates fresh ones
        based on the current labor total. If labor amount is 0, no HM/EPP is
        created (and any existing HM/EPP are still deleted).

        Returns: tuple (hm_created: bool, epp_created: bool)
        """
        from decimal import Decimal as D, ROUND_HALF_UP

        UnitCostBreakdown.objects.filter(
            conceptid_id=concept_id,
            categorycode__in=[
                BreakdownCategoryCode.MINOR_TOOLS,
                BreakdownCategoryCode.PPE,
            ],
        ).delete()

        labor_amount = UnitCostBreakdown.objects.filter(
            conceptid_id=concept_id,
            categorycode=BreakdownCategoryCode.LABOR,
        ).aggregate(total=Sum('amount'))['total'] or D('0')

        if labor_amount <= 0:
            return (False, False)

        quantity = D('0.03')
        unitprice = labor_amount.quantize(D('0.01'), ROUND_HALF_UP)
        yieldvalue = D('1')
        amount = (quantity * unitprice * yieldvalue).quantize(D('0.01'), ROUND_HALF_UP)

        for category, desc in [
            (BreakdownCategoryCode.MINOR_TOOLS, 'HERRAMIENTA MENOR'),
            (BreakdownCategoryCode.PPE, 'EPP'),
        ]:
            UnitCostBreakdown.objects.create(
                conceptid_id=concept_id,
                categorycode=category,
                linenumber=1,
                description=desc,
                unit='%',
                quantity=quantity,
                unitprice=unitprice,
                yieldvalue=yieldvalue,
                amount=amount,
            )

        return (True, True)

    # -------------------------------------------------------------------------
    # Skeleton generation rules (hardcoded, mirrors frontend config)
    # -------------------------------------------------------------------------
    SKELETON_RULES = [
        {
            'ruleid': 'rule-excavacion',
            'name': 'Excavacion / Terraceria',
            'subfamily_patterns': ['excavac', 'terraceria', 'desmonte', 'despalme', 'deshierbe'],
            'unit_patterns': ['m3', 'm2', 'ha'],
            'include_hm_epp': True,
            'lines': [
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Diesel', 'unit': 'Lt', 'qty': 1, 'price': '23.09', 'yield': '0.862'},
                {'cat': BreakdownCategoryCode.MACHINERY, 'desc': 'Excavadora 336', 'unit': 'Hr', 'qty': 1, 'price': '900', 'yield': '0.034'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Operador 336', 'unit': 'jor', 'qty': 1, 'price': '1071.43', 'yield': '0.00426'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Viatico Operador 336', 'unit': 'semana', 'qty': 1, 'price': '2000', 'yield': '0.00071'},
            ],
        },
        {
            'ruleid': 'rule-concreto',
            'name': 'Concreto',
            'subfamily_patterns': ['concreto', 'ciment', 'colado', 'zapata', 'losa'],
            'unit_patterns': ['m3'],
            'include_hm_epp': True,
            'lines': [
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Cemento Portland CPC 40', 'unit': 'ton', 'qty': 1, 'price': '3200', 'yield': '0.35'},
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Arena', 'unit': 'm3', 'qty': 1, 'price': '280', 'yield': '0.56'},
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Grava 3/4"', 'unit': 'm3', 'qty': 1, 'price': '320', 'yield': '0.756'},
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Agua', 'unit': 'Lt', 'qty': 1, 'price': '0.15', 'yield': '180'},
                {'cat': BreakdownCategoryCode.MACHINERY, 'desc': 'Revolvedora 1 saco', 'unit': 'Hr', 'qty': 1, 'price': '150', 'yield': '0.5'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Cuadrilla (1 oficial + 2 ayudantes)', 'unit': 'jor', 'qty': 1, 'price': '3500', 'yield': '0.125'},
            ],
        },
        {
            'ruleid': 'rule-acero',
            'name': 'Acero Estructural',
            'subfamily_patterns': ['acero', 'armado', 'refuerzo', 'habilitado'],
            'unit_patterns': ['kg', 'ton'],
            'include_hm_epp': True,
            'lines': [
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Acero de refuerzo', 'unit': 'kg', 'qty': 1, 'price': '19', 'yield': '1.05'},
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Alambre recocido', 'unit': 'kg', 'qty': 1, 'price': '28', 'yield': '0.04'},
                {'cat': BreakdownCategoryCode.MACHINERY, 'desc': 'Equipo de corte y doblado', 'unit': 'Hr', 'qty': 1, 'price': '80', 'yield': '0.01'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Armador fierrero', 'unit': 'jor', 'qty': 1, 'price': '850', 'yield': '0.01'},
            ],
        },
        {
            'ruleid': 'rule-malla',
            'name': 'Malla / Gavion',
            'subfamily_patterns': ['malla', 'gavion', 'triple torsion', 'torsion'],
            'unit_patterns': ['m2'],
            'include_hm_epp': True,
            'lines': [
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Malla triple torsion 8x10 calibre 12', 'unit': 'm2', 'qty': 1, 'price': '96', 'yield': '1.2'},
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Cable de acero Galv 1/2"', 'unit': 'ml', 'qty': 1, 'price': '42', 'yield': '0.17'},
                {'cat': BreakdownCategoryCode.MATERIALS, 'desc': 'Varilla corrugada 1/2"', 'unit': 'kg', 'qty': 1, 'price': '19', 'yield': '0.75'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Cuadrilla 1 cabo + 4 instaladores', 'unit': 'jor', 'qty': 1, 'price': '5500', 'yield': '0.01818'},
            ],
        },
        {
            'ruleid': 'rule-acarreo',
            'name': 'Acarreo',
            'subfamily_patterns': ['acarreo', 'carga y acarreo', 'transporte'],
            'unit_patterns': ['m3'],
            'include_hm_epp': True,
            'lines': [
                {'cat': BreakdownCategoryCode.MACHINERY, 'desc': 'Cargador Frontal', 'unit': 'Hr', 'qty': 1, 'price': '900', 'yield': '0.048'},
                {'cat': BreakdownCategoryCode.MACHINERY, 'desc': 'Camion de volteo 14m3', 'unit': 'jor', 'qty': 1, 'price': '6000', 'yield': '0.006'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Operador Cargador Frontal', 'unit': 'jor', 'qty': 1, 'price': '928.57', 'yield': '0.0048'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Viatico Operador Cargador', 'unit': 'semana', 'qty': 1, 'price': '2000', 'yield': '0.0008'},
            ],
        },
        {
            'ruleid': 'rule-topografia',
            'name': 'Topografia',
            'subfamily_patterns': ['topograf', 'trazo', 'nivelacion'],
            'unit_patterns': ['m2', 'ml'],
            'include_hm_epp': True,
            'lines': [
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Topografo + Ayudantes', 'unit': 'jor', 'qty': 1, 'price': '4500', 'yield': '0.003'},
                {'cat': BreakdownCategoryCode.LABOR, 'desc': 'Trabajos de gabinete', 'unit': 'jor', 'qty': 1, 'price': '2500', 'yield': '0.001'},
            ],
        },
    ]

    @classmethod
    def _normalize_text(cls, text: str) -> str:
        """Normalize text: lowercase, remove accents."""
        import unicodedata
        normalized = unicodedata.normalize('NFD', text.lower())
        return ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')

    @classmethod
    def _match_rule_against_text(cls, text: str, unit: str):
        """Try to match a rule against a given text and unit."""
        normalized = cls._normalize_text(text)
        norm_unit = unit.lower().strip()

        for rule in cls.SKELETON_RULES:
            subfamily_match = any(p in normalized for p in rule['subfamily_patterns'])
            if not subfamily_match:
                continue
            if rule['unit_patterns']:
                if norm_unit in rule['unit_patterns']:
                    return rule
            else:
                return rule
        return None

    @classmethod
    def _find_matching_rule(cls, subfamily_name: str, unit: str, description: str = ''):
        """Find a skeleton rule: first by subfamily name, then fallback to concept description."""
        rule = cls._match_rule_against_text(subfamily_name, unit)
        if rule:
            return rule
        if description:
            return cls._match_rule_against_text(description, unit)
        return None

    @classmethod
    @transaction.atomic
    def auto_generate_skeleton(cls, concept_id: UUID, subfamily_name: str, unit: str, user, description: str = '') -> list:
        """Auto-generate skeleton breakdown lines based on subfamily and unit matching rules.
        Falls back to concept description if subfamily doesn't match."""
        rule = cls._find_matching_rule(subfamily_name, unit, description)
        if not rule:
            return []

        created = []
        for line_def in rule['lines']:
            cat_code = line_def['cat']
            max_line = UnitCostBreakdown.objects.filter(
                conceptid=concept_id,
                categorycode=cat_code,
            ).aggregate(max_line=Max('linenumber'))['max_line'] or 0

            qty = Decimal(str(line_def['qty']))
            price = Decimal(line_def['price'])
            yld = Decimal(line_def['yield'])
            amount = qty * price * yld

            breakdown = UnitCostBreakdown(
                conceptid_id=concept_id,
                categorycode=cat_code,
                linenumber=max_line + 1,
                description=line_def['desc'],
                unit=line_def['unit'],
                quantity=qty,
                unitprice=price,
                yieldvalue=yld,
                amount=amount,
            )
            breakdown.save()
            created.append(breakdown)

        # Optionally add HM + EPP lines
        if rule['include_hm_epp']:
            hm_epp = cls.auto_generate_hm_epp(concept_id, user)
            created.extend(hm_epp)

        cls._recalc_concept(concept_id, user)
        return created

    @staticmethod
    def duplicate_line(breakdown_id: UUID, user) -> UnitCostBreakdown:
        """Duplicate an existing breakdown line within the same concept and category."""
        original = UnitCostBreakdown.objects.select_related('supplyid').get(
            breakdownid=breakdown_id, statecode=0
        )
        max_line = UnitCostBreakdown.objects.filter(
            conceptid=original.conceptid,
            categorycode=original.categorycode,
            statecode=0,
        ).aggregate(Max('linenumber'))['linenumber__max'] or 0

        new_line = UnitCostBreakdown(
            conceptid=original.conceptid,
            categorycode=original.categorycode,
            linenumber=max_line + 1,
            description=original.description,
            unit=original.unit,
            quantity=original.quantity,
            unitprice=original.unitprice,
            yieldvalue=original.yieldvalue,
            amount=original.amount,
            supplyid=original.supplyid,
        )
        new_line.save()
        UnitCostBreakdownService._recalc_concept(original.conceptid_id, user)
        return new_line

    @staticmethod
    @transaction.atomic
    def copy_from_concept(target_concept_id: UUID, source_concept_id: UUID, user) -> list[UnitCostBreakdown]:
        """Copy all active breakdown lines from source concept to target concept."""
        source_lines = UnitCostBreakdown.objects.filter(
            conceptid_id=source_concept_id, statecode=0
        ).select_related('supplyid').order_by('categorycode', 'linenumber')

        if not source_lines.exists():
            raise ValidationError("El concepto origen no tiene lineas de desglose.")

        existing = UnitCostBreakdown.objects.filter(
            conceptid_id=target_concept_id, statecode=0
        ).values('categorycode').annotate(max_line=Max('linenumber'))
        max_lines = {row['categorycode']: row['max_line'] for row in existing}

        new_lines = []
        for line in source_lines:
            current_max = max_lines.get(line.categorycode, 0)
            current_max += 1
            max_lines[line.categorycode] = current_max

            new_lines.append(UnitCostBreakdown(
                conceptid_id=target_concept_id,
                categorycode=line.categorycode,
                linenumber=current_max,
                description=line.description,
                unit=line.unit,
                quantity=line.quantity,
                unitprice=line.unitprice,
                yieldvalue=line.yieldvalue,
                amount=line.amount,
                supplyid=line.supplyid,
            ))

        created = UnitCostBreakdown.objects.bulk_create(new_lines)
        UnitCostBreakdownService._recalc_concept(target_concept_id, user)
        return created


class IndirectCostDetailService:
    """Service class for IndirectCostDetail business logic."""

    @staticmethod
    def list_details(
        project_id: UUID,
        user,
        categorycode: Optional[str] = None,
    ) -> QuerySet[IndirectCostDetail]:
        """List indirect cost details for a project, optionally filtered by category."""
        queryset = IndirectCostDetail.objects.filter(projectid=project_id)

        if categorycode is not None:
            queryset = queryset.filter(categorycode=categorycode)

        return queryset.select_related('createdby', 'modifiedby')

    @staticmethod
    def create_detail(dto: CreateIndirectCostDetailDto, user) -> IndirectCostDetail:
        """Create a new indirect cost detail line with computed amount and auto linenumber."""
        # Calculate next linenumber within this project and category
        max_line = IndirectCostDetail.objects.filter(
            projectid=dto.projectid,
            categorycode=dto.categorycode,
        ).aggregate(max_line=Max('linenumber'))['max_line'] or 0
        next_line = max_line + 1

        # Compute amount = monthlycost * units * months
        monthlycost = dto.monthlycost or Decimal('0')
        units = dto.units or Decimal('1')
        months = dto.months or Decimal('0')
        amount = monthlycost * units * months

        detail = IndirectCostDetail(
            projectid_id=dto.projectid,
            categorycode=dto.categorycode,
            linenumber=next_line,
            imputationcode=dto.imputationcode or '',
            area=dto.area or '',
            description=dto.description,
            monthlycost=monthlycost,
            units=units,
            months=months,
            amount=amount,
            createdby=user,
            modifiedby=user,
        )
        detail.save()
        return detail

    @staticmethod
    def update_detail(detail_id: UUID, dto: UpdateIndirectCostDetailDto, user) -> IndirectCostDetail:
        """Update an indirect cost detail line and recompute amount."""
        try:
            detail = IndirectCostDetail.objects.get(indirectcostid=detail_id)
        except IndirectCostDetail.DoesNotExist:
            raise NotFound(f"IndirectCostDetail with ID {detail_id} not found")

        update_fields = [
            'categorycode', 'imputationcode', 'area', 'description',
            'monthlycost', 'units', 'months', 'statecode',
        ]
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(detail, field, value)

        # Recompute amount
        detail.amount = detail.monthlycost * detail.units * detail.months

        detail.modifiedby = user
        detail.save()
        return detail

    @staticmethod
    def delete_detail(detail_id: UUID, user) -> IndirectCostDetail:
        """Soft delete an indirect cost detail (statecode=1)."""
        try:
            detail = IndirectCostDetail.objects.get(indirectcostid=detail_id)
        except IndirectCostDetail.DoesNotExist:
            raise NotFound(f"IndirectCostDetail with ID {detail_id} not found")

        detail.statecode = 1
        detail.modifiedby = user
        detail.save()
        return detail

    @staticmethod
    @transaction.atomic
    def apply_template(project_id: UUID, projectsize: int, user) -> list[IndirectCostDetail]:
        """Replace project's indirect cost details with lines from the matching template.

        Existing details (and their cascading CostDistribution cells) are deleted before
        seeding to honor the UI contract that applying a template replaces current data.
        """
        if projectsize not in [c.value for c in ProjectSizeCode]:
            raise ValidationError(f"Invalid project size: {projectsize}")

        templates = IndirectCostTemplate.objects.filter(
            projectsize=projectsize,
            statecode=0,
        ).order_by('categorycode', 'sortorder')

        if not templates.exists():
            raise ValidationError(f"No templates found for project size {projectsize}")

        IndirectCostDetail.objects.filter(projectid=project_id).delete()

        created = []
        line_counters = defaultdict(int)

        for template in templates:
            line_counters[template.categorycode] += 1
            amount = template.monthlycost * template.units * template.months

            detail = IndirectCostDetail(
                projectid_id=project_id,
                categorycode=template.categorycode,
                linenumber=line_counters[template.categorycode],
                description=template.description,
                monthlycost=template.monthlycost,
                units=template.units,
                months=template.months,
                amount=amount,
                createdby=user,
                modifiedby=user,
            )
            created.append(detail)

        IndirectCostDetail.objects.bulk_create(created)
        return created

    @staticmethod
    def get_total(project_id: UUID, user) -> Decimal:
        """Get total indirect cost amount for a project."""
        total = IndirectCostDetail.objects.filter(
            projectid=project_id,
            statecode=0,
        ).aggregate(total=Sum('amount'))['total']
        return total or Decimal('0')

    @staticmethod
    @transaction.atomic
    def prorate_to_concepts(project_id: UUID, user) -> list[BudgetConcept]:
        """Distribute total indirect cost across concepts proportionally by direct cost.

        Each concept gets:
            indirectunitcost = (directunitcost * quantity / Σ direct) * total_indirect / quantity
            utilityunitcost  = (directunitcost + indirectunitcost) * project.profitpercent / 100
            unitprice        = directunitcost + indirectunitcost + utilityunitcost
            totalamount      = unitprice * quantity
        """
        total_indirect = IndirectCostDetailService.get_total(project_id, user)

        try:
            project = EstimationProject.objects.get(estimationprojectid=project_id)
        except EstimationProject.DoesNotExist:
            raise NotFound(f"EstimationProject {project_id} not found")
        profitpercent = project.profitpercent or Decimal('0')

        concepts = BudgetConcept.objects.filter(
            projectid=project_id,
            statecode=0,
        )

        # Calculate total direct cost across all concepts (DB aggregation,
        # avoids loading every concept into Python just to sum).
        total_direct = concepts.aggregate(
            total=Sum(F('directunitcost') * F('quantity'), output_field=models.DecimalField())
        )['total'] or Decimal('0')

        if total_direct <= 0:
            return list(concepts)

        updated = []
        for concept in concepts:
            concept_direct_total = concept.directunitcost * concept.quantity
            proportion = concept_direct_total / total_direct
            concept_indirect_share = total_indirect * proportion

            # Distribute as unit cost
            if concept.quantity > 0:
                concept.indirectunitcost = concept_indirect_share / concept.quantity
            else:
                concept.indirectunitcost = Decimal('0')

            concept.utilityunitcost = (
                (concept.directunitcost + concept.indirectunitcost) * profitpercent / Decimal('100')
            )
            concept.unitprice = concept.directunitcost + concept.indirectunitcost + concept.utilityunitcost
            concept.totalamount = concept.unitprice * concept.quantity
            concept.modifiedby = user
            updated.append(concept)

        BudgetConcept.objects.bulk_update(
            updated,
            ['indirectunitcost', 'utilityunitcost', 'unitprice', 'totalamount',
             'modifiedby', 'modifiedon'],
        )
        return updated


class OfferAlternativeService:
    """Service class for OfferAlternative business logic."""

    @staticmethod
    def list_alternatives(project_id: UUID, user) -> QuerySet[OfferAlternative]:
        """List active offer alternatives for a project (excludes soft-deleted)."""
        return OfferAlternative.objects.filter(
            projectid=project_id,
            statecode=0,  # Active only — soft-deleted have statecode=1
        ).select_related('createdby', 'modifiedby')

    @staticmethod
    def create_alternative(dto: CreateOfferAlternativeDto, user) -> OfferAlternative:
        """Create a new offer alternative. Max 4 per project."""
        existing_count = OfferAlternative.objects.filter(
            projectid=dto.projectid,
            statecode=0,
        ).count()

        if existing_count >= 4:
            raise ValidationError("Maximum 4 active alternatives per project")

        # Auto-assign alternative number
        max_num = OfferAlternative.objects.filter(
            projectid=dto.projectid,
        ).aggregate(max_num=Max('alternativenumber'))['max_num'] or 0
        next_num = max_num + 1

        # Compute totals from project concepts + indirects
        concepts = BudgetConcept.objects.filter(
            projectid=dto.projectid,
            statecode=0,
        )
        direct_total = concepts.aggregate(
            total=Sum(F('directunitcost') * F('quantity'), output_field=models.DecimalField())
        )['total'] or Decimal('0')

        indirect_total = IndirectCostDetail.objects.filter(
            projectid=dto.projectid,
            statecode=0,
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        construction_cost = direct_total + indirect_total

        transversal = dto.transversalpercent or Decimal('0')
        profit = dto.profitpercent or Decimal('0')

        # coefficient = 1 + transversal/100 + profit/100
        coefficient = Decimal('1') + transversal / Decimal('100') + profit / Decimal('100')
        sale_price_net = construction_cost * coefficient
        tax_amount = sale_price_net * Decimal('0.16')
        sale_price_total = sale_price_net + tax_amount

        alternative = OfferAlternative(
            projectid_id=dto.projectid,
            alternativenumber=next_num,
            name=dto.name,
            description=dto.description or '',
            transversalpercent=transversal,
            profitpercent=profit,
            coefficient=coefficient,
            directcosttotal=direct_total,
            indirectcosttotal=indirect_total,
            constructioncost=construction_cost,
            salepricenet=sale_price_net,
            taxamount=tax_amount,
            salepricetotal=sale_price_total,
            authorizationname=dto.authorizationname or '',
            authorizationposition=dto.authorizationposition or '',
            createdby=user,
            modifiedby=user,
        )
        alternative.save()
        return alternative

    @staticmethod
    def update_alternative(alternative_id: UUID, dto: UpdateOfferAlternativeDto, user) -> OfferAlternative:
        """Update an offer alternative and recompute derived fields."""
        try:
            alternative = OfferAlternative.objects.get(alternativeid=alternative_id)
        except OfferAlternative.DoesNotExist:
            raise NotFound(f"OfferAlternative with ID {alternative_id} not found")

        update_fields = [
            'name', 'description', 'transversalpercent', 'profitpercent',
            'authorizationname', 'authorizationposition', 'statecode',
        ]
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(alternative, field, value)

        # Recompute coefficient and derived totals
        alternative.coefficient = (
            Decimal('1')
            + alternative.transversalpercent / Decimal('100')
            + alternative.profitpercent / Decimal('100')
        )
        alternative.constructioncost = alternative.directcosttotal + alternative.indirectcosttotal
        alternative.salepricenet = alternative.constructioncost * alternative.coefficient
        alternative.taxamount = alternative.salepricenet * Decimal('0.16')
        alternative.salepricetotal = alternative.salepricenet + alternative.taxamount

        alternative.modifiedby = user
        alternative.save()
        return alternative

    @staticmethod
    def delete_alternative(alternative_id: UUID, user) -> OfferAlternative:
        """Soft delete an offer alternative (statecode=1)."""
        try:
            alternative = OfferAlternative.objects.get(alternativeid=alternative_id)
        except OfferAlternative.DoesNotExist:
            raise NotFound(f"OfferAlternative with ID {alternative_id} not found")

        alternative.statecode = 1
        alternative.modifiedby = user
        alternative.save()
        return alternative

    @staticmethod
    @transaction.atomic
    def choose_alternative(alternative_id: UUID, user) -> OfferAlternative:
        """Set one alternative as chosen, unset all others for the same project,
        sync ``project.profitpercent`` to the chosen alternative's profitpercent
        and re-prorate indirect costs so every concept's utilityunitcost picks
        up the new rate (Excel "E7 Fase Estudio" K2 behavior).
        """
        try:
            alternative = OfferAlternative.objects.get(alternativeid=alternative_id)
        except OfferAlternative.DoesNotExist:
            raise NotFound(f"OfferAlternative with ID {alternative_id} not found")

        # Unset all others
        OfferAlternative.objects.filter(
            projectid=alternative.projectid,
        ).update(ischosen=False)

        # Set this one
        alternative.ischosen = True
        alternative.modifiedby = user
        alternative.save()

        # Sync the project-level profit rate and recompute every concept so the
        # CD / CI / Utilidad / P.U. / Importe columns stay consistent with the
        # chosen offer (utility = (CD+CI) * profit% / 100).
        project = alternative.projectid
        project.profitpercent = alternative.profitpercent or Decimal('0')
        project.modifiedby = user
        project.save(update_fields=['profitpercent', 'modifiedby', 'modifiedon'])
        IndirectCostDetailService.prorate_to_concepts(project.estimationprojectid, user)

        return alternative


class ExternalCostService:
    """Service class for ExternalCostItem business logic."""

    @staticmethod
    def list_costs(project_id: UUID, user) -> QuerySet[ExternalCostItem]:
        """List all external cost items for a project."""
        return ExternalCostItem.objects.filter(
            projectid=project_id
        )

    @staticmethod
    @transaction.atomic
    def initialize_checklist(project_id: UUID, user) -> list[ExternalCostItem]:
        """Create default external cost checklist items (~21 items)."""
        existing = ExternalCostItem.objects.filter(projectid=project_id).exists()
        if existing:
            raise ValidationError("External cost checklist already initialized for this project")

        items = []
        for idx, name in enumerate(DEFAULT_EXTERNAL_COSTS, start=1):
            item = ExternalCostItem(
                projectid_id=project_id,
                itemname=name,
                applies=0,  # N/A by default
                percentofsale=None,
                amount=Decimal('0'),
                sortorder=idx,
            )
            items.append(item)

        ExternalCostItem.objects.bulk_create(items)
        return items

    @staticmethod
    def update_cost(cost_id: UUID, dto: UpdateExternalCostItemDto, user) -> ExternalCostItem:
        """Update an external cost item.

        ``amount`` is derived: when the item applies (Yes), it is recomputed as
        ``(percentofsale / 100) * project.estimatedcontractamount``. When applies
        is N/A or No, ``amount`` is forced to 0. Any client-supplied ``amount``
        in the DTO is ignored.
        """
        try:
            cost = ExternalCostItem.objects.select_related('projectid').get(externalcostid=cost_id)
        except ExternalCostItem.DoesNotExist:
            raise NotFound(f"ExternalCostItem with ID {cost_id} not found")

        if dto.applies is not None:
            cost.applies = dto.applies
        if dto.percentofsale is not None:
            cost.percentofsale = dto.percentofsale
        if dto.statecode is not None:
            cost.statecode = dto.statecode

        if cost.applies == ChecklistStatusCode.YES:
            contract_amount = cost.projectid.estimatedcontractamount or Decimal('0')
            percent = cost.percentofsale or Decimal('0')
            cost.amount = (percent / Decimal('100')) * contract_amount
        else:
            cost.amount = Decimal('0')

        cost.save()
        return cost


class SupplyExplosionService:
    """Service class for supply explosion reports (read-only, computed)."""

    @staticmethod
    def generate_auxiliary(project_id: UUID, user) -> list[dict]:
        """Generate the auxiliary supply explosion: iterate all UnitCostBreakdowns for a project.

        Returns a list of dicts matching SupplyExplosionItemSchema.
        """
        breakdowns = UnitCostBreakdown.objects.filter(
            conceptid__projectid=project_id,
            statecode=0,
        ).select_related('conceptid', 'supplyid')

        results = []
        for bd in breakdowns:
            concept = bd.conceptid
            results.append({
                'conceptid': concept.conceptid,
                'conceptcode': concept.code,
                'conceptdescription': concept.description,
                'conceptquantity': concept.quantity,
                'categorycode': bd.categorycode,
                'supplyid': bd.supplyid.supplyid if bd.supplyid else None,
                'supplycode': bd.supplyid.code if bd.supplyid else None,
                'description': bd.description,
                'unit': bd.unit,
                'quantity': bd.quantity,
                'unitprice': bd.unitprice,
                'amount': bd.amount,
            })

        return results

    @staticmethod
    def generate_consolidated(project_id: UUID, user) -> list[dict]:
        """Generate consolidated supply explosion grouped by supply code.

        Aggregates each breakdown line by its supply (catalog item), scaling
        ``quantity`` and ``amount`` by the parent concept's quantity. This
        makes the totals reflect the **whole project** (e.g., total kg of
        cement to buy, total cost of cement) rather than per-unit-of-concept,
        so the consolidated total reconciles with ``Σ directunitcost × quantity``
        used by the cuadre indicator.

        Returns a list of dicts matching SupplyExplosionConsolidatedSchema.
        """
        breakdowns = UnitCostBreakdown.objects.filter(
            conceptid__projectid=project_id,
            statecode=0,
            supplyid__isnull=False,
        ).select_related('supplyid', 'conceptid')

        # Group by supply code
        groups = defaultdict(lambda: {
            'description': '',
            'unit': '',
            'supplytype': 0,
            'totalquantity': Decimal('0'),
            'totalamount': Decimal('0'),
            'concepts': set(),
        })

        for bd in breakdowns:
            supply = bd.supplyid
            concept_qty = bd.conceptid.quantity or Decimal('0')
            key = supply.code
            group = groups[key]
            group['description'] = supply.description
            group['unit'] = supply.unit
            group['supplytype'] = supply.supplytype
            group['totalquantity'] += bd.quantity * concept_qty
            group['totalamount'] += bd.amount * concept_qty
            group['concepts'].add(bd.conceptid_id)

        results = []
        for supplycode, data in sorted(groups.items()):
            total_qty = data['totalquantity']
            total_amt = data['totalamount']
            avg_price = total_amt / total_qty if total_qty > 0 else Decimal('0')

            results.append({
                'supplycode': supplycode,
                'description': data['description'],
                'unit': data['unit'],
                'supplytype': data['supplytype'],
                'totalquantity': total_qty,
                'averageprice': avg_price,
                'totalamount': total_amt,
                'conceptcount': len(data['concepts']),
            })

        return results

    # Labels for human-readable export — single source of truth.
    _CATEGORY_LABELS = {
        1: 'Materiales',
        2: 'Acarreos',
        3: 'Maquinaria',
        4: 'Mano de Obra',
        5: 'Subcontratos',
        6: 'Herramienta Menor',
        7: 'EPP',
    }
    _SUPPLY_TYPE_LABELS = {
        0: 'Material',
        1: 'Mano de Obra',
        2: 'Maquinaria',
        3: 'Subcontrato',
        4: 'Acarreo',
    }

    @classmethod
    def export_excel(cls, project_id: UUID, user) -> bytes:
        """Export the supply explosion to an .xlsx with two sheets.

        Sheet 1 'Auxiliar': one row per UnitCostBreakdown line, with full
        concept + supply context (cod, descripción, unidad, qty, PU, importe).
        Sheet 2 'Consolidado': aggregated by supply code (totales del proyecto).

        Includes a header row 1 with the project name and row 2 with the
        project UUID (informativo; no se re-importa, este export es read-only).
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        project = EstimationProject.objects.filter(pk=project_id).first()
        project_name = getattr(project, 'name', '') if project else ''

        aux_data = cls.generate_auxiliary(project_id, user)
        cons_data = cls.generate_consolidated(project_id, user)

        wb = Workbook()

        # ---------- Sheet 1: Auxiliar ----------
        ws_aux = wb.active
        ws_aux.title = 'Auxiliar'

        ws_aux.cell(row=1, column=1, value=f'Proyecto: {project_name}').font = Font(bold=True, size=12)
        ws_aux.cell(row=2, column=1, value=str(project_id)).font = Font(italic=True, color='888888')

        aux_headers = [
            'COD. CONCEPTO', 'DESCRIPCION CONCEPTO', 'CANTIDAD CONCEPTO',
            'CATEGORIA', 'COD. INSUMO', 'DESCRIPCION INSUMO', 'UNIDAD',
            'CANTIDAD', 'P. UNITARIO', 'IMPORTE U.', 'IMPORTE TOTAL',
        ]
        header_fill = PatternFill('solid', fgColor='DDDDDD')
        for i, h in enumerate(aux_headers, start=1):
            cell = ws_aux.cell(row=3, column=i, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        ws_aux.freeze_panes = 'A4'
        aux_widths = [14, 50, 16, 14, 14, 50, 10, 12, 14, 14, 16]
        for i, w in enumerate(aux_widths, start=1):
            ws_aux.column_dimensions[get_column_letter(i)].width = w

        row = 4
        for item in aux_data:
            concept_qty = item['conceptquantity'] or Decimal('0')
            line_amount = item['amount'] or Decimal('0')
            total_amount = line_amount * concept_qty
            ws_aux.cell(row=row, column=1, value=item['conceptcode'])
            ws_aux.cell(row=row, column=2, value=item['conceptdescription'])
            ws_aux.cell(row=row, column=3, value=float(concept_qty))
            ws_aux.cell(
                row=row, column=4,
                value=cls._CATEGORY_LABELS.get(item['categorycode'], f"Cat {item['categorycode']}"),
            )
            ws_aux.cell(row=row, column=5, value=item['supplycode'] or '')
            ws_aux.cell(row=row, column=6, value=item['description'])
            ws_aux.cell(row=row, column=7, value=item['unit'])
            ws_aux.cell(row=row, column=8, value=float(item['quantity']))
            ws_aux.cell(row=row, column=9, value=float(item['unitprice']))
            ws_aux.cell(row=row, column=10, value=float(line_amount))
            ws_aux.cell(row=row, column=11, value=float(total_amount))
            row += 1

        # ---------- Sheet 2: Consolidado ----------
        ws_cons = wb.create_sheet('Consolidado')

        ws_cons.cell(row=1, column=1, value=f'Proyecto: {project_name}').font = Font(bold=True, size=12)
        ws_cons.cell(row=2, column=1, value=str(project_id)).font = Font(italic=True, color='888888')

        cons_headers = [
            'COD. INSUMO', 'DESCRIPCION', 'UNIDAD', 'TIPO',
            'CANTIDAD TOTAL', 'PRECIO PROMEDIO', 'IMPORTE TOTAL',
            '% INCIDENCIA', '# CONCEPTOS',
        ]
        for i, h in enumerate(cons_headers, start=1):
            cell = ws_cons.cell(row=3, column=i, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        ws_cons.freeze_panes = 'A4'
        cons_widths = [14, 50, 10, 16, 16, 16, 16, 14, 12]
        for i, w in enumerate(cons_widths, start=1):
            ws_cons.column_dimensions[get_column_letter(i)].width = w

        project_total = sum(
            (Decimal(str(r['totalamount'])) for r in cons_data),
            Decimal('0'),
        )

        row = 4
        for item in cons_data:
            total_amt = item['totalamount'] or Decimal('0')
            pct = (total_amt / project_total * Decimal('100')) if project_total > 0 else Decimal('0')
            ws_cons.cell(row=row, column=1, value=item['supplycode'])
            ws_cons.cell(row=row, column=2, value=item['description'])
            ws_cons.cell(row=row, column=3, value=item['unit'])
            ws_cons.cell(
                row=row, column=4,
                value=cls._SUPPLY_TYPE_LABELS.get(item['supplytype'], f"Tipo {item['supplytype']}"),
            )
            ws_cons.cell(row=row, column=5, value=float(item['totalquantity']))
            ws_cons.cell(row=row, column=6, value=float(item['averageprice']))
            ws_cons.cell(row=row, column=7, value=float(total_amt))
            ws_cons.cell(row=row, column=8, value=float(pct))
            ws_cons.cell(row=row, column=9, value=item['conceptcount'])
            row += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


class WorkPlanService:
    """Service class for WorkPlanEntry business logic."""

    @staticmethod
    def list_entries(
        project_id: UUID,
        user,
        conceptid: Optional[UUID] = None,
        entrytype: Optional[int] = None,
    ) -> QuerySet[WorkPlanEntry]:
        """List work plan entries for a project, optionally filtered by concept/entrytype."""
        queryset = WorkPlanEntry.objects.filter(projectid=project_id)

        if conceptid is not None:
            queryset = queryset.filter(conceptid=conceptid)
        if entrytype is not None:
            queryset = queryset.filter(entrytype=entrytype)

        return queryset.select_related('conceptid', 'createdby', 'modifiedby')

    @staticmethod
    def create_entry(dto: CreateWorkPlanEntryDto, user) -> WorkPlanEntry:
        """Create a work plan entry with computed distributedamount."""
        # Fetch concept to get unitprice
        try:
            concept = BudgetConcept.objects.get(conceptid=dto.conceptid)
        except BudgetConcept.DoesNotExist:
            raise NotFound(f"BudgetConcept with ID {dto.conceptid} not found")

        distributed_amount = dto.distributedquantity * concept.unitprice

        entry = WorkPlanEntry(
            conceptid_id=dto.conceptid,
            projectid_id=dto.projectid,
            periodnumber=dto.periodnumber,
            periodlabel=dto.periodlabel,
            entrytype=getattr(dto, 'entrytype', WorkPlanEntryType.PLANNED),
            distributedquantity=dto.distributedquantity,
            distributedamount=distributed_amount,
            createdby=user,
            modifiedby=user,
        )
        entry.save()
        return entry

    @staticmethod
    def update_entry(entry_id: UUID, dto: UpdateWorkPlanEntryDto, user) -> WorkPlanEntry:
        """Update a work plan entry and recompute amount."""
        try:
            entry = WorkPlanEntry.objects.select_related('conceptid').get(
                workplanentryid=entry_id
            )
        except WorkPlanEntry.DoesNotExist:
            raise NotFound(f"WorkPlanEntry with ID {entry_id} not found")

        if dto.distributedquantity is not None:
            entry.distributedquantity = dto.distributedquantity
            entry.distributedamount = dto.distributedquantity * entry.conceptid.unitprice

        entry.modifiedby = user
        entry.save()
        return entry

    @staticmethod
    def delete_entry(entry_id: UUID, user) -> None:
        """Hard delete a work plan entry."""
        try:
            entry = WorkPlanEntry.objects.get(workplanentryid=entry_id)
        except WorkPlanEntry.DoesNotExist:
            raise NotFound(f"WorkPlanEntry with ID {entry_id} not found")

        entry.delete()

    @staticmethod
    @transaction.atomic
    def bulk_distribute(project_id: UUID, entries_data: list, user) -> list[WorkPlanEntry]:
        """Bulk create/update work plan entries.

        entries_data: list of dicts with {conceptid, periodnumber, periodlabel, distributedquantity}

        Note: Previously this validated that sum per concept <= concept.quantity, but the
        Excel source of truth does not enforce that cap on actual production, so we keep
        the check out of the path. The UI still highlights rows that exceed capacity.
        """
        # Validate that every referenced concept exists (fail fast with a clean 404).
        # Single query + dict reuse instead of one .exists() per concept and one
        # .get() per entry inside the loop (avoids N+1).
        referenced_concepts = {entry_data['conceptid'] for entry_data in entries_data}
        # Key by str so lookups work whether the payload sends UUID objects or
        # string UUIDs (c.conceptid is a UUID, payload values are strings).
        concepts_by_id = {
            str(c.conceptid): c
            for c in BudgetConcept.objects.filter(conceptid__in=referenced_concepts)
        }
        missing = {str(cid) for cid in referenced_concepts if str(cid) not in concepts_by_id}
        if missing:
            raise NotFound(f"BudgetConcept with ID {next(iter(missing))} not found")

        created_or_updated = []
        for entry_data in entries_data:
            concept_id_val = entry_data['conceptid']
            period_num = entry_data['periodnumber']
            period_label = entry_data['periodlabel']
            etype = int(entry_data.get('entrytype', WorkPlanEntryType.PLANNED))
            dist_qty = Decimal(str(entry_data['distributedquantity']))

            concept = concepts_by_id[str(concept_id_val)]
            dist_amount = dist_qty * concept.unitprice

            entry, _created = WorkPlanEntry.objects.update_or_create(
                conceptid_id=concept_id_val,
                periodnumber=period_num,
                entrytype=etype,
                defaults={
                    'projectid_id': project_id,
                    'periodlabel': period_label,
                    'distributedquantity': dist_qty,
                    'distributedamount': dist_amount,
                    'createdby': user,
                    'modifiedby': user,
                },
            )
            created_or_updated.append(entry)

        return created_or_updated

    # ------------------------------------------------------------------
    # Matrix / Summary aggregations (Excel "Plan de obra" replica)
    # ------------------------------------------------------------------

    @staticmethod
    def _build_value_dict(entries, concept_by_id):
        """Build {'total_qty','total_amount','by_period'} for a list of entries of one entrytype."""
        total_qty = Decimal('0')
        total_amount = Decimal('0')
        by_period = {}
        for e in entries:
            total_qty += e.distributedquantity
            total_amount += e.distributedamount
            by_period[str(e.periodnumber)] = by_period.get(str(e.periodnumber), Decimal('0')) + e.distributedquantity
        return {
            'total_qty': total_qty,
            'total_amount': total_amount,
            'by_period': by_period,
        }

    @staticmethod
    def get_matrix(project_id: UUID, user) -> dict:
        """Return the full (family → subfamily → concept × period) matrix with totals.

        Mirrors the Excel "Plan de obra" layout with both planned and actual blocks.
        """
        concepts = (
            BudgetConcept.objects
            .filter(projectid=project_id)
            .select_related('subfamilyid', 'subfamilyid__familyid')
            .order_by('subfamilyid__familyid__sortorder', 'subfamilyid__sortorder', 'sequencenumber')
        )

        entries = (
            WorkPlanEntry.objects
            .filter(projectid=project_id)
            .only('conceptid_id', 'periodnumber', 'periodlabel', 'entrytype',
                  'distributedquantity', 'distributedamount')
        )

        # Group entries by (conceptid, entrytype)
        grouped = defaultdict(list)
        periods_map = {}
        for e in entries:
            grouped[(str(e.conceptid_id), e.entrytype)].append(e)
            periods_map[e.periodnumber] = e.periodlabel

        periods = [
            {'number': n, 'label': periods_map[n]}
            for n in sorted(periods_map)
        ]

        # Build family/subfamily tree
        family_tree = {}  # familyid -> {family meta, subfamilies: {subfamilyid: {...}}}
        for c in concepts:
            sf = c.subfamilyid
            fm = sf.familyid
            fnode = family_tree.setdefault(str(fm.familyid), {
                'familyid': fm.familyid,
                'code': fm.code,
                'name': fm.name,
                'sortorder': getattr(fm, 'sortorder', 0),
                'contract_amount': Decimal('0'),
                'subfamilies': {},
                'planned_amount': Decimal('0'),
                'actual_amount': Decimal('0'),
                'planned_by_period_amount': defaultdict(lambda: Decimal('0')),
                'actual_by_period_amount': defaultdict(lambda: Decimal('0')),
            })
            sfnode = fnode['subfamilies'].setdefault(str(sf.subfamilyid), {
                'subfamilyid': sf.subfamilyid,
                'code': sf.code,
                'name': sf.name,
                'sortorder': getattr(sf, 'sortorder', 0),
                'concepts': [],
            })

            planned_entries = grouped.get((str(c.conceptid), WorkPlanEntryType.PLANNED), [])
            actual_entries = grouped.get((str(c.conceptid), WorkPlanEntryType.ACTUAL), [])

            planned_val = WorkPlanService._build_value_dict(planned_entries, None)
            actual_val = WorkPlanService._build_value_dict(actual_entries, None)

            sfnode['concepts'].append({
                'conceptid': c.conceptid,
                'code': c.code,
                'description': c.description,
                'unit': c.unit,
                'quantity': c.quantity,
                'unitprice': c.unitprice,
                'totalamount': c.totalamount,
                'planned': planned_val,
                'actual': actual_val,
            })

            fnode['contract_amount'] += c.totalamount or Decimal('0')
            fnode['planned_amount'] += planned_val['total_amount']
            fnode['actual_amount'] += actual_val['total_amount']
            # Per-period amount (SUMPRODUCT equivalent): qty * unitprice
            for pnum, qty in planned_val['by_period'].items():
                fnode['planned_by_period_amount'][pnum] += qty * (c.unitprice or Decimal('0'))
            for pnum, qty in actual_val['by_period'].items():
                fnode['actual_by_period_amount'][pnum] += qty * (c.unitprice or Decimal('0'))

        # Flatten tree to lists with totals
        families = []
        gt_contract = Decimal('0')
        gt_planned = Decimal('0')
        gt_actual = Decimal('0')
        gt_planned_period = defaultdict(lambda: Decimal('0'))
        gt_actual_period = defaultdict(lambda: Decimal('0'))

        for fnode in sorted(family_tree.values(), key=lambda x: (x['sortorder'], x['code'])):
            subfams = sorted(fnode['subfamilies'].values(), key=lambda x: (x['sortorder'], x['code']))
            families.append({
                'familyid': fnode['familyid'],
                'code': fnode['code'],
                'name': fnode['name'],
                'contract_amount': fnode['contract_amount'],
                'subfamilies': [
                    {
                        'subfamilyid': sf['subfamilyid'],
                        'code': sf['code'],
                        'name': sf['name'],
                        'concepts': sf['concepts'],
                    }
                    for sf in subfams
                ],
                'totals': {
                    'planned_amount': fnode['planned_amount'],
                    'actual_amount': fnode['actual_amount'],
                    'planned_by_period_amount': dict(fnode['planned_by_period_amount']),
                    'actual_by_period_amount': dict(fnode['actual_by_period_amount']),
                },
            })
            gt_contract += fnode['contract_amount']
            gt_planned += fnode['planned_amount']
            gt_actual += fnode['actual_amount']
            for k, v in fnode['planned_by_period_amount'].items():
                gt_planned_period[k] += v
            for k, v in fnode['actual_by_period_amount'].items():
                gt_actual_period[k] += v

        return {
            'periods': periods,
            'families': families,
            'grand_totals': {
                'contract_amount': gt_contract,
                'planned_amount': gt_planned,
                'actual_amount': gt_actual,
                'planned_by_period_amount': dict(gt_planned_period),
                'actual_by_period_amount': dict(gt_actual_period),
            },
        }

    @staticmethod
    def get_summary(project_id: UUID, user) -> dict:
        """Compact per-family summary: contract, planned, actual amounts + percents."""
        matrix = WorkPlanService.get_matrix(project_id, user)
        families = []
        for f in matrix['families']:
            contract = f['contract_amount'] or Decimal('0')
            planned = f['totals']['planned_amount']
            actual = f['totals']['actual_amount']
            families.append({
                'familyid': f['familyid'],
                'code': f['code'],
                'name': f['name'],
                'contract_amount': contract,
                'planned_amount': planned,
                'actual_amount': actual,
                'percent_planned': float(planned / contract) if contract else 0.0,
                'percent_actual': float(actual / contract) if contract else 0.0,
            })
        return {
            'families': families,
            'grand_totals': matrix['grand_totals'],
        }


class TemporalDistributionService:
    """Service class for temporal distribution reports (read-only, computed)."""

    @staticmethod
    def calculate(project_id: UUID, user) -> list[dict]:
        """Aggregate workplan entries by period, compute invoiced, costs, results with cumulatives.

        Returns a list of dicts matching TemporalDistributionSchema.
        """
        entries = WorkPlanEntry.objects.filter(
            projectid=project_id,
        ).select_related('conceptid').order_by('periodnumber')

        # Group by period
        periods = defaultdict(lambda: {
            'periodlabel': '',
            'invoicedamount': Decimal('0'),
            'costamount': Decimal('0'),
        })

        for entry in entries:
            period = periods[entry.periodnumber]
            period['periodlabel'] = entry.periodlabel

            concept = entry.conceptid
            # Invoiced amount = distributedquantity * client unit price (or unitprice if no client price)
            client_price = concept.clientunitprice if concept.clientunitprice else concept.unitprice
            period['invoicedamount'] += entry.distributedquantity * client_price

            # Cost amount = distributedquantity * direct unit cost
            period['costamount'] += entry.distributedquantity * concept.directunitcost

        # Build result with cumulatives
        results = []
        cumulative_invoiced = Decimal('0')
        cumulative_cost = Decimal('0')
        cumulative_result = Decimal('0')

        for period_num in sorted(periods.keys()):
            data = periods[period_num]
            invoiced = data['invoicedamount']
            cost = data['costamount']
            result = invoiced - cost

            cumulative_invoiced += invoiced
            cumulative_cost += cost
            cumulative_result += result

            results.append({
                'periodnumber': period_num,
                'periodlabel': data['periodlabel'],
                'invoicedamount': invoiced,
                'costamount': cost,
                'resultamount': result,
                'cumulativeinvoiced': cumulative_invoiced,
                'cumulativecost': cumulative_cost,
                'cumulativeresult': cumulative_result,
            })

        return results


class SupplyCatalogService:
    """Service class for SupplyCatalogItem business logic."""

    @staticmethod
    def list_items(
        user,
        search: Optional[str] = None,
        supplytype: Optional[int] = None,
    ) -> QuerySet[SupplyCatalogItem]:
        """List supply catalog items with optional text search and type filter."""
        queryset = SupplyCatalogItem.objects.filter(statecode=0)

        if supplytype is not None:
            queryset = queryset.filter(supplytype=supplytype)

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) | Q(description__icontains=search)
            )

        return queryset.select_related('createdby', 'modifiedby')

    @staticmethod
    def create_item(dto: CreateSupplyCatalogItemDto, user) -> SupplyCatalogItem:
        """Create a new supply catalog item."""
        item = SupplyCatalogItem(
            code=dto.code,
            description=dto.description,
            unit=dto.unit,
            supplytype=dto.supplytype,
            referenceprice=dto.referenceprice or Decimal('0'),
            referencedate=dto.referencedate,
            geographiczone=dto.geographiczone or '',
            createdby=user,
            modifiedby=user,
        )
        item.save()
        return item

    @staticmethod
    def update_item(item_id: UUID, dto: UpdateSupplyCatalogItemDto, user) -> SupplyCatalogItem:
        """Update a supply catalog item."""
        try:
            item = SupplyCatalogItem.objects.get(supplyid=item_id)
        except SupplyCatalogItem.DoesNotExist:
            raise NotFound(f"SupplyCatalogItem with ID {item_id} not found")

        update_fields = [
            'code', 'description', 'unit', 'supplytype',
            'referenceprice', 'referencedate', 'geographiczone', 'statecode',
        ]
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(item, field, value)

        item.modifiedby = user
        item.save()
        return item

    @staticmethod
    def delete_item(item_id: UUID, user) -> SupplyCatalogItem:
        """Soft delete a supply catalog item (statecode=1)."""
        try:
            item = SupplyCatalogItem.objects.get(supplyid=item_id)
        except SupplyCatalogItem.DoesNotExist:
            raise NotFound(f"SupplyCatalogItem with ID {item_id} not found")

        item.statecode = 1
        item.modifiedby = user
        item.save()
        return item


class EquipmentYieldService:
    """Service class for EquipmentYield business logic."""

    @staticmethod
    def list_yields(
        user,
        category: Optional[str] = None,
    ) -> QuerySet[EquipmentYield]:
        """List equipment yields with optional category filter."""
        queryset = EquipmentYield.objects.filter(statecode=0)

        if category is not None:
            queryset = queryset.filter(category=category)

        return queryset.select_related('createdby', 'modifiedby')

    @staticmethod
    def create_yield(dto: CreateEquipmentYieldDto, user) -> EquipmentYield:
        """Create an equipment yield record with computed fields.

        Calculations:
        - realyield = theoreticalyield * trafficfactor
        - dailyfuelconsumption = fuelconsumption * effectivehours
        - monthlycubicmeters = realyield * effectivehours * effectivedays * numberofequipment
        - monthlydiesel = dailyfuelconsumption * effectivedays * numberofequipment
        - costpercubicmeter = monthlycost / monthlycubicmeters (if > 0)
        """
        theoretical = dto.theoreticalyield or Decimal('0')
        traffic = dto.trafficfactor or Decimal('0.8')
        fuel = dto.fuelconsumption or Decimal('0')
        hours = dto.effectivehours or Decimal('0')
        days = dto.effectivedays or Decimal('0')
        num_equip = dto.numberofequipment or 1
        monthly_cost = dto.monthlycost or Decimal('0')

        real_yield = theoretical * traffic
        daily_fuel = fuel * hours
        monthly_cubic = real_yield * hours * days * num_equip
        monthly_diesel = daily_fuel * days * num_equip
        cost_per_cubic = monthly_cost / monthly_cubic if monthly_cubic > 0 else Decimal('0')

        equipment = EquipmentYield(
            category=dto.category,
            description=dto.description,
            suppliername=dto.suppliername or '',
            monthlycost=monthly_cost,
            numberofequipment=num_equip,
            theoreticalyield=theoretical,
            effectivehours=hours,
            realyield=real_yield,
            fuelconsumption=fuel,
            dailyfuelconsumption=daily_fuel,
            effectivedays=days,
            trafficfactor=traffic,
            monthlycubicmeters=monthly_cubic,
            monthlydiesel=monthly_diesel,
            costpercubicmeter=cost_per_cubic,
            createdby=user,
            modifiedby=user,
        )
        equipment.save()
        return equipment

    @staticmethod
    def update_yield(yield_id: UUID, dto: UpdateEquipmentYieldDto, user) -> EquipmentYield:
        """Update an equipment yield record and recompute calculated fields."""
        try:
            equipment = EquipmentYield.objects.get(equipmentyieldid=yield_id)
        except EquipmentYield.DoesNotExist:
            raise NotFound(f"EquipmentYield with ID {yield_id} not found")

        update_fields = [
            'category', 'description', 'suppliername', 'monthlycost',
            'numberofequipment', 'theoreticalyield', 'effectivehours',
            'fuelconsumption', 'effectivedays', 'trafficfactor', 'statecode',
        ]
        for field in update_fields:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(equipment, field, value)

        # Recompute calculated fields
        equipment.realyield = equipment.theoreticalyield * equipment.trafficfactor
        equipment.dailyfuelconsumption = equipment.fuelconsumption * equipment.effectivehours
        equipment.monthlycubicmeters = (
            equipment.realyield
            * equipment.effectivehours
            * equipment.effectivedays
            * equipment.numberofequipment
        )
        equipment.monthlydiesel = (
            equipment.dailyfuelconsumption
            * equipment.effectivedays
            * equipment.numberofequipment
        )
        equipment.costpercubicmeter = (
            equipment.monthlycost / equipment.monthlycubicmeters
            if equipment.monthlycubicmeters > 0
            else Decimal('0')
        )

        equipment.modifiedby = user
        equipment.save()
        return equipment

    @staticmethod
    def delete_yield(yield_id: UUID, user) -> EquipmentYield:
        """Soft delete an equipment yield record (statecode=1)."""
        try:
            equipment = EquipmentYield.objects.get(equipmentyieldid=yield_id)
        except EquipmentYield.DoesNotExist:
            raise NotFound(f"EquipmentYield with ID {yield_id} not found")

        equipment.statecode = 1
        equipment.modifiedby = user
        equipment.save()
        return equipment


class ConceptPriceCatalogService:
    """Service class for ConceptPriceCatalogItem and ConceptPriceReference."""

    @staticmethod
    def _next_code(source: int) -> str:
        """Generate the next auto-incremented code for a source."""
        prefix = {0: 'SICT', 1: 'HIST', 2: 'MAN'}.get(source, 'HIST')
        last = (
            ConceptPriceCatalogItem.objects
            .filter(code__startswith=f'{prefix}-')
            .order_by('-code')
            .values_list('code', flat=True)
            .first()
        )
        if last:
            try:
                num = int(last.split('-')[-1]) + 1
            except (ValueError, IndexError):
                num = 1
        else:
            num = 1
        return f"{prefix}-{num:05d}"

    @staticmethod
    def list_items(
        user,
        search: Optional[str] = None,
        source: Optional[int] = None,
        unit: Optional[str] = None,
    ) -> QuerySet[ConceptPriceCatalogItem]:
        """List catalog items with optional filters."""
        queryset = ConceptPriceCatalogItem.objects.filter(statecode=0)

        if source is not None:
            queryset = queryset.filter(source=source)

        if unit:
            queryset = queryset.filter(unit__iexact=unit)

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) | Q(description__icontains=search)
            )

        return queryset.select_related('createdby', 'modifiedby')

    @staticmethod
    def get_item(item_id: UUID) -> ConceptPriceCatalogItem:
        """Get a single catalog item with its references."""
        try:
            return ConceptPriceCatalogItem.objects.prefetch_related(
                'price_references'
            ).get(catalogitemid=item_id)
        except ConceptPriceCatalogItem.DoesNotExist:
            from core.exceptions import NotFound
            raise NotFound(f"ConceptPriceCatalogItem with ID {item_id} not found")

    @staticmethod
    def create_item(dto, user) -> ConceptPriceCatalogItem:
        """Create a new catalog item."""
        code = dto.code or ConceptPriceCatalogService._next_code(dto.source)
        item = ConceptPriceCatalogItem(
            code=code,
            description=dto.description,
            unit=dto.unit,
            source=dto.source,
            category=dto.category or '',
            createdby=user,
            modifiedby=user,
        )
        item.save()
        return item

    @staticmethod
    def update_item(item_id: UUID, dto, user) -> ConceptPriceCatalogItem:
        """Update a catalog item."""
        try:
            item = ConceptPriceCatalogItem.objects.get(catalogitemid=item_id)
        except ConceptPriceCatalogItem.DoesNotExist:
            from core.exceptions import NotFound
            raise NotFound(f"ConceptPriceCatalogItem with ID {item_id} not found")

        for field in ['code', 'description', 'unit', 'source', 'category', 'statecode']:
            value = getattr(dto, field, None)
            if value is not None:
                setattr(item, field, value)

        item.modifiedby = user
        item.save()
        return item

    @staticmethod
    def delete_item(item_id: UUID, user) -> ConceptPriceCatalogItem:
        """Soft delete a catalog item."""
        try:
            item = ConceptPriceCatalogItem.objects.get(catalogitemid=item_id)
        except ConceptPriceCatalogItem.DoesNotExist:
            from core.exceptions import NotFound
            raise NotFound(f"ConceptPriceCatalogItem with ID {item_id} not found")

        item.statecode = 1
        item.modifiedby = user
        item.save()
        return item

    # --- Price References ---

    @staticmethod
    def list_references(
        catalog_item_id: UUID,
    ) -> QuerySet[ConceptPriceReference]:
        """List references for a specific catalog item."""
        return ConceptPriceReference.objects.filter(
            catalogitemid=catalog_item_id,
            statecode=0,
        ).select_related('createdby', 'modifiedby')

    @staticmethod
    def create_reference(dto, user) -> ConceptPriceReference:
        """Create a price reference and update parent stats."""
        ref = ConceptPriceReference(
            catalogitemid_id=dto.catalogitemid,
            projectname=dto.projectname,
            projectlocation=dto.projectlocation or '',
            unitprice=dto.unitprice,
            quantity=dto.quantity,
            totalamount=dto.totalamount,
            referencedate=dto.referencedate,
            notes=dto.notes or '',
            createdby=user,
            modifiedby=user,
        )
        ref.save()

        # Update parent stats
        parent = ref.catalogitemid
        parent.update_price_stats()
        parent.save()

        return ref

    @staticmethod
    def delete_reference(reference_id: UUID, user) -> ConceptPriceReference:
        """Soft delete a reference and update parent stats."""
        try:
            ref = ConceptPriceReference.objects.select_related(
                'catalogitemid'
            ).get(referenceid=reference_id)
        except ConceptPriceReference.DoesNotExist:
            from core.exceptions import NotFound
            raise NotFound(f"ConceptPriceReference with ID {reference_id} not found")

        ref.statecode = 1
        ref.modifiedby = user
        ref.save()

        # Update parent stats
        parent = ref.catalogitemid
        parent.update_price_stats()
        parent.save()

        return ref

    @staticmethod
    @transaction.atomic
    def bulk_import(items: list, user) -> dict:
        """Bulk import concepts with their project references.

        Each item: {description, unit, source, category?, references: [{projectname, unitprice, quantity?, totalamount?}]}
        Returns: {created: int, references_created: int}
        """
        created = 0
        refs_created = 0

        for item_data in items:
            code = ConceptPriceCatalogService._next_code(
                item_data.get('source', CatalogSourceCode.HISTORICO)
            )

            catalog_item = ConceptPriceCatalogItem(
                code=code,
                description=item_data['description'],
                unit=item_data['unit'],
                source=item_data.get('source', CatalogSourceCode.HISTORICO),
                category=item_data.get('category', ''),
                createdby=user,
                modifiedby=user,
            )
            catalog_item.save()
            created += 1

            references = item_data.get('references', [])
            for ref_data in references:
                if not ref_data.get('unitprice') or ref_data['unitprice'] <= 0:
                    continue
                ref = ConceptPriceReference(
                    catalogitemid=catalog_item,
                    projectname=ref_data['projectname'],
                    projectlocation=ref_data.get('projectlocation', ''),
                    unitprice=ref_data['unitprice'],
                    quantity=ref_data.get('quantity'),
                    totalamount=ref_data.get('totalamount'),
                    referencedate=ref_data.get('referencedate'),
                    createdby=user,
                    modifiedby=user,
                )
                ref.save()
                refs_created += 1

            # Update stats
            catalog_item.update_price_stats()
            catalog_item.save()

        return {'created': created, 'references_created': refs_created}


class FamilyTemplateService:
    """Service for managing reusable family/subfamily templates."""

    @staticmethod
    def list_template_sets(user, category=None, search=None):
        from django.db.models import Count
        qs = FamilyTemplateSet.objects.filter(statecode=0).annotate(
            _family_count=Count('items__familycode', filter=Q(items__statecode=0), distinct=True),
            _subfamily_count=Count('items', filter=Q(items__statecode=0)),
        )
        if category:
            qs = qs.filter(category=category)
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(description__icontains=search))
        return qs.prefetch_related('items')

    @staticmethod
    def get_template_set(template_set_id, user):
        try:
            return FamilyTemplateSet.objects.prefetch_related('items').get(
                templatesetid=template_set_id, statecode=0
            )
        except FamilyTemplateSet.DoesNotExist:
            raise NotFound(f"FamilyTemplateSet with ID {template_set_id} not found")

    @staticmethod
    def create_template_set(dto, user):
        ts = FamilyTemplateSet(
            name=dto.name,
            description=dto.description or '',
            category=dto.category,
            issystem=False,
            createdby=user,
            modifiedby=user,
        )
        ts.save()
        return ts

    @staticmethod
    def delete_template_set(template_set_id, user):
        try:
            ts = FamilyTemplateSet.objects.get(templatesetid=template_set_id)
        except FamilyTemplateSet.DoesNotExist:
            raise NotFound(f"FamilyTemplateSet with ID {template_set_id} not found")
        if ts.issystem:
            raise ValidationError("No se pueden eliminar plantillas del sistema.")
        ts.statecode = 1  # Inactive
        ts.modifiedby = user
        ts.save()
        return ts

    @staticmethod
    @transaction.atomic
    def save_project_as_template(dto: SaveProjectAsTemplateDto, user):
        """Extract families+subfamilies from a project and save as a new template."""
        families = ConceptFamily.objects.filter(
            projectid=dto.projectid, statecode=0
        ).order_by('sortorder')

        if not families.exists():
            raise ValidationError("El proyecto no tiene familias para guardar como plantilla.")

        ts = FamilyTemplateSet(
            name=dto.name,
            description=dto.description or '',
            category=dto.category,
            issystem=False,
            createdby=user,
            modifiedby=user,
        )
        ts.save()

        items = []
        for family in families:
            subfamilies = ConceptSubfamily.objects.filter(
                familyid=family, statecode=0
            ).order_by('sortorder')
            if subfamilies.exists():
                for sf in subfamilies:
                    items.append(FamilyTemplateItem(
                        templatesetid=ts,
                        familycode=family.code,
                        familyname=family.name,
                        subfamilycode=sf.code,
                        subfamilyname=sf.name,
                        familysortorder=family.sortorder,
                        subfamilysortorder=sf.sortorder,
                    ))
            else:
                items.append(FamilyTemplateItem(
                    templatesetid=ts,
                    familycode=family.code,
                    familyname=family.name,
                    subfamilycode='',
                    subfamilyname='',
                    familysortorder=family.sortorder,
                    subfamilysortorder=0,
                ))
        FamilyTemplateItem.objects.bulk_create(items)
        return ts

    @staticmethod
    @transaction.atomic
    def apply_template_to_project(dto: ApplyFamilyTemplateDto, user):
        """Create families+subfamilies in a project from a template. Skips existing codes."""
        try:
            ts = FamilyTemplateSet.objects.prefetch_related('items').get(
                templatesetid=dto.templatesetid, statecode=0
            )
        except FamilyTemplateSet.DoesNotExist:
            raise NotFound("Plantilla no encontrada")

        items = ts.items.filter(statecode=0)
        if dto.familycodes:
            items = items.filter(familycode__in=dto.familycodes)

        existing_codes = set(
            ConceptFamily.objects.filter(
                projectid=dto.projectid, statecode=0
            ).values_list('code', flat=True)
        )

        # Group items by family
        families_map = defaultdict(list)
        for item in items.order_by('familysortorder', 'subfamilysortorder'):
            if item.familycode not in existing_codes:
                families_map[
                    (item.familycode, item.familyname, item.familysortorder)
                ].append(item)

        created = []
        for (fcode, fname, fsort), sub_items in families_map.items():
            family = ConceptFamily(
                projectid_id=dto.projectid,
                name=fname,
                code=fcode,
                sortorder=fsort,
                createdby=user,
                modifiedby=user,
            )
            family.save()
            created.append(family)

            for item in sub_items:
                if item.subfamilycode:
                    ConceptSubfamily(
                        familyid=family,
                        projectid_id=dto.projectid,
                        name=item.subfamilyname,
                        code=item.subfamilycode,
                        sortorder=item.subfamilysortorder,
                        createdby=user,
                        modifiedby=user,
                    ).save()

        return created


# =============================================================================
# ExcelImportService — Analyze & import concepts from Excel
# =============================================================================

class ExcelImportService:
    """Handles Excel upload analysis and concept import into a project."""

    @staticmethod
    def analyze(project_id, file, user):
        """
        Parse an Excel file and match each row against the concept catalog.
        Returns analysis results WITHOUT writing to the database.
        """
        import openpyxl
        from apps.proyeccion.matching import match_concepts

        project = EstimationProject.objects.get(estimationprojectid=project_id)

        # Parse Excel
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # Find header row (look in first 10 rows for "PARTIDA")
        header_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            cell_val = str(ws.cell(row=row_idx, column=1).value or '').strip().upper()
            if 'PARTIDA' in cell_val:
                header_row = row_idx
                break

        if header_row is None:
            from core.exceptions import ValidationError
            raise ValidationError(
                'No se encontro la fila de encabezados. '
                'La primera columna debe contener "PARTIDA".'
            )

        # Parse data rows
        rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            partida = str(ws.cell(row=row_idx, column=1).value or '').strip()
            code = str(ws.cell(row=row_idx, column=2).value or '').strip()
            description = str(ws.cell(row=row_idx, column=3).value or '').strip()
            unit = str(ws.cell(row=row_idx, column=4).value or '').strip()
            quantity_raw = ws.cell(row=row_idx, column=5).value

            if not description:
                continue

            try:
                quantity = float(quantity_raw) if quantity_raw else 0.0
            except (ValueError, TypeError):
                quantity = 0.0

            # Inherit partida from previous row if empty
            if not partida and rows:
                partida = rows[-1]['partida']

            rows.append({
                'row': row_idx,
                'partida': partida,
                'code': code,
                'description': description,
                'unit': unit,
                'quantity': quantity,
            })

        wb.close()

        if not rows:
            from core.exceptions import ValidationError
            raise ValidationError('El archivo no contiene filas de datos.')

        # Match against catalog
        matched_rows = match_concepts(rows)

        # Build partidas list (check against existing subfamilies)
        existing_subfamilies = {
            sf.name.strip().upper(): sf.subfamilyid
            for sf in ConceptSubfamily.objects.filter(
                projectid=project, statecode=0
            )
        }

        partida_names = []
        seen = set()
        for r in matched_rows:
            p = r['partida']
            if p and p not in seen:
                seen.add(p)
                sf_id = existing_subfamilies.get(p.upper())
                partida_names.append({
                    'name': p,
                    'subfamilyid': sf_id,
                    'is_new': sf_id is None,
                })

        # Build response
        concepts = []
        for r in matched_rows:
            candidate = r.get('match_candidate')
            candidate_data = None
            if candidate:
                candidate_data = {
                    'catalogitemid': candidate.catalogitemid,
                    'code': candidate.code,
                    'description': candidate.description,
                    'unit': candidate.unit,
                    'averageprice': float(candidate.averageprice or 0),
                    'classificationl2': candidate.classificationl2 or '',
                    'classificationl3': candidate.classificationl3 or '',
                }

            concepts.append({
                'row': r['row'],
                'partida': r['partida'],
                'code': r['code'],
                'description': r['description'],
                'unit': r['unit'],
                'quantity': r['quantity'],
                'match_status': r['match_status'],
                'match_score': r['match_score'],
                'match_candidate': candidate_data,
            })

        summary = {
            'total': len(concepts),
            'exact': sum(1 for c in concepts if c['match_status'] == 'exact'),
            'partial': sum(1 for c in concepts if c['match_status'] == 'partial'),
            'none': sum(1 for c in concepts if c['match_status'] == 'none'),
        }

        return {
            'partidas': partida_names,
            'concepts': concepts,
            'summary': summary,
        }

    @staticmethod
    @transaction.atomic
    def do_import(project_id, payload, user):
        """
        Import analyzed concepts into the project as BudgetConcepts.
        Creates missing subfamilies if requested.
        """
        project = EstimationProject.objects.get(estimationprojectid=project_id)

        # Get or create the default family for imports
        family = ConceptFamily.objects.filter(
            projectid=project, statecode=0
        ).order_by('sortorder').first()

        if not family:
            family = ConceptFamily.objects.create(
                projectid=project,
                name='GENERAL',
                code='IMP',
                sortorder=0,
                createdby=user,
                modifiedby=user,
            )

        # Resolve partidas -> subfamilies
        existing_sfs = {
            sf.name.strip().upper(): sf
            for sf in ConceptSubfamily.objects.filter(
                projectid=project, statecode=0
            )
        }

        partida_map = {}  # partida_name_upper -> ConceptSubfamily
        subfamilies_created = 0
        next_sort = (
            ConceptSubfamily.objects.filter(projectid=project).count() + 1
        )

        for item in payload.items:
            p_name = item.partida.strip()
            p_upper = p_name.upper()

            if p_upper in partida_map:
                continue

            if p_upper in existing_sfs:
                partida_map[p_upper] = existing_sfs[p_upper]
            elif payload.create_missing_subfamilies:
                # Extract code from partida (e.g., "01. GABINETE" -> code="01", name="GABINETE")
                parts = p_name.split('.', 1)
                sf_code = parts[0].strip() if len(parts) > 1 else str(next_sort).zfill(2)
                sf_name = parts[1].strip() if len(parts) > 1 else p_name

                sf = ConceptSubfamily.objects.create(
                    familyid=family,
                    projectid=project,
                    name=sf_name or p_name,
                    code=sf_code,
                    sortorder=next_sort,
                    createdby=user,
                    modifiedby=user,
                )
                partida_map[p_upper] = sf
                existing_sfs[p_upper] = sf
                subfamilies_created += 1
                next_sort += 1

        # Create BudgetConcepts
        created_count = 0
        matched_count = 0

        for item in payload.items:
            p_upper = item.partida.strip().upper()
            subfamily = partida_map.get(p_upper)
            if not subfamily:
                continue

            # Resolve catalog price if matched
            client_price = None
            if item.accepted_catalog_id and item.use_catalog_price:
                try:
                    cat_item = ConceptPriceCatalogItem.objects.get(
                        catalogitemid=item.accepted_catalog_id
                    )
                    client_price = cat_item.averageprice if cat_item.averageprice > 0 else None
                except ConceptPriceCatalogItem.DoesNotExist:
                    pass

            from apps.proyeccion.schemas import CreateBudgetConceptDto
            dto = CreateBudgetConceptDto(
                projectid=project_id,
                subfamilyid=subfamily.subfamilyid,
                description=item.description,
                unit=item.unit,
                quantity=Decimal(str(item.quantity)) if item.quantity else Decimal('0'),
                breakdownmethod=0,
                clientunitprice=client_price,
                isprintable=True,
            )

            ConceptCatalogService.create_concept(dto, user)
            created_count += 1

            if item.accepted_catalog_id:
                matched_count += 1

        return {
            'created': created_count,
            'subfamilies_created': subfamilies_created,
            'matched': matched_count,
        }


# =============================================================================
# Temporal Distribution — Period Service
# =============================================================================

from datetime import date, timedelta

from apps.proyeccion.models import (
    ProjectionPeriod,
    CostDistribution,
)

SPANISH_MONTHS = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN',
                  'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC']


class PeriodService:
    """Generate and regenerate ProjectionPeriod rows for an EstimationProject."""

    @staticmethod
    def _iter_periods(startdate, enddate, periodtype):
        """Yield (startdate, enddate) tuples covering [startdate, enddate]."""
        delta_days = 7 if periodtype == 0 else 15  # weekly or fortnightly
        cursor = startdate
        n = 0
        while cursor <= enddate:
            period_end = cursor + timedelta(days=delta_days - 1)
            if period_end > enddate:
                period_end = enddate
            yield (cursor, period_end)
            cursor = period_end + timedelta(days=1)
            n += 1
            if n > 500:
                raise RuntimeError("Period generation exceeded 500 periods — aborting")

    @staticmethod
    def _label(periodnumber, start, periodtype):
        prefix = 'S' if periodtype == 0 else 'Q'
        month_abbr = SPANISH_MONTHS[start.month - 1]
        yr = start.year % 100
        return f"{prefix}{periodnumber:02d} {month_abbr}-{yr:02d}"

    @staticmethod
    @transaction.atomic
    def regenerate_projection_periods(project: EstimationProject, *, confirm=False) -> dict:
        """Regenerate ProjectionPeriod[] from project.estimatedstartdate + estimatedenddate + periodtype.

        Returns: {
          'created': N,
          'deleted': M,
          'kept': K,
          'lost_manual_edits': X  # count of CostDistribution isderived=false that were deleted
        }
        Raises: ValueError if dates missing. Also raises ValueError if lost_manual_edits > 0
                and confirm=False (caller must re-invoke with confirm=True).
        """
        if not project.estimatedstartdate or not project.estimatedenddate:
            raise ValueError("Project needs fechas estimadas (start/end) to regenerate periods")

        # Compute new periods
        new_periods = []
        for i, (s, e) in enumerate(PeriodService._iter_periods(
            project.estimatedstartdate, project.estimatedenddate, project.periodtype
        )):
            new_periods.append({
                'periodnumber': i + 1,
                'periodlabel': PeriodService._label(i + 1, s, project.periodtype),
                'startdate': s,
                'enddate': e,
                'periodtype': project.periodtype,
            })
        new_count = len(new_periods)

        # Check what we'd lose: manual edits in period slots that won't exist anymore
        existing_count = ProjectionPeriod.objects.filter(projectid=project).count()
        lost_manual = 0
        if existing_count > new_count:
            lost_manual = CostDistribution.objects.filter(
                projectid=project, isderived=False,
                periodnumber__gt=new_count,
            ).count()
            if lost_manual > 0 and not confirm:
                raise ValueError(
                    f"regenerate would discard {lost_manual} manual distribution edits "
                    f"in periods > {new_count}. Re-invoke with confirm=True to proceed."
                )

        # Delete existing periods (periodnumber > new_count) and their distributions
        ProjectionPeriod.objects.filter(
            projectid=project, periodnumber__gt=new_count
        ).delete()
        CostDistribution.objects.filter(
            projectid=project, periodnumber__gt=new_count
        ).delete()

        # Upsert new periods
        for p in new_periods:
            ProjectionPeriod.objects.update_or_create(
                projectid=project, periodnumber=p['periodnumber'],
                defaults={
                    'periodlabel': p['periodlabel'],
                    'startdate': p['startdate'],
                    'enddate': p['enddate'],
                    'periodtype': p['periodtype'],
                },
            )

        # Update cached count
        project.periodcount = new_count
        project.save(update_fields=['periodcount'])

        return {
            'created': new_count,
            'deleted': max(0, existing_count - new_count),
            'kept': min(existing_count, new_count),
            'lost_manual_edits': lost_manual,
        }


# =============================================================================
# Cost Distribution Service
# =============================================================================

from django.db.models import DecimalField
from django.db.models.functions import Coalesce
from apps.proyeccion.models import CostLineType


class CostDistributionService:
    """Business logic for CostDistribution — rollups, autofill, bulk edits."""

    @staticmethod
    def compute_rollups(project) -> dict:
        """Compute SUMPRODUCT(amount, fraction) across all lines for each period.

        Returns dict with per-period arrays (length = periodcount) and totals.
        All numeric values are Decimal.
        """
        N = project.periodcount or 0
        zeros = [Decimal("0")] * N

        # Direct: per-line vectors keyed by breakdown UUID + aggregate by period.
        direct_by_period = list(zeros)
        direct_by_period_by_line: dict[str, list[Decimal]] = {}
        lag_by_line: dict[str, int | None] = {}

        # ``UnitCostBreakdown.amount`` is per-unit-of-concept (Σ
        # quantity*unitprice*yieldvalue of an APU ingredient line). The
        # project-level cost a breakdown contributes is
        # ``amount × concept.quantity``, matching the convention used by
        # ``OfferAlternativeService.regenerate_alternatives`` and
        # ``BudgetConceptService.recalculate_concept``. Multiply by
        # ``conceptid.quantity`` here so the PNT rollups are at the same
        # scale as the alternative totals.
        breakdowns = UnitCostBreakdown.objects.filter(
            conceptid__projectid=project,
        ).select_related('conceptid')
        bd_dist_qs = CostDistribution.objects.filter(
            projectid=project, linetype=CostLineType.BREAKDOWN,
        ).values('breakdownid_id', 'periodnumber', 'fraction')
        bd_dist_by_id: dict[str, dict[int, Decimal]] = {}
        for row in bd_dist_qs:
            bd_dist_by_id.setdefault(str(row['breakdownid_id']), {})[row['periodnumber']] = Decimal(row['fraction'])

        for bd in breakdowns:
            bd_id = str(bd.breakdownid)
            line_vec = list(zeros)
            cells = bd_dist_by_id.get(bd_id, {})
            concept_qty = bd.conceptid.quantity or Decimal('0')
            line_total = (bd.amount or Decimal('0')) * concept_qty
            for i in range(N):
                frac = cells.get(i + 1, Decimal('0'))
                value = line_total * frac
                line_vec[i] = value
                direct_by_period[i] += value
            direct_by_period_by_line[bd_id] = line_vec
            lag_by_line[bd_id] = bd.paymentlagperiods

        # Indirect: per-line vectors keyed by indirectcost UUID + aggregate by period.
        indirect_by_period = list(zeros)
        indirect_by_period_by_line: dict[str, list[Decimal]] = {}

        indirects = IndirectCostDetail.objects.filter(projectid=project)
        ind_dist_qs = CostDistribution.objects.filter(
            projectid=project, linetype=CostLineType.INDIRECT,
        ).values('indirectcostid_id', 'periodnumber', 'fraction')
        ind_dist_by_id: dict[str, dict[int, Decimal]] = {}
        for row in ind_dist_qs:
            ind_dist_by_id.setdefault(str(row['indirectcostid_id']), {})[row['periodnumber']] = Decimal(row['fraction'])

        for ind in indirects:
            ind_id = str(ind.indirectcostid)
            line_vec = list(zeros)
            cells = ind_dist_by_id.get(ind_id, {})
            for i in range(N):
                frac = cells.get(i + 1, Decimal('0'))
                value = (ind.amount or Decimal('0')) * frac
                line_vec[i] = value
                indirect_by_period[i] += value
            indirect_by_period_by_line[ind_id] = line_vec
            lag_by_line[ind_id] = ind.paymentlagperiods

        # Retiros via chosen alternative.
        # transversalpercent / profitpercent are stored as RAW percentages
        # (e.g. 30 means 30%), matching the form-to-DB convention used by
        # OfferAlternativeService.create_alternative which also divides by 100
        # to compute the coefficient. Without this division retiros and
        # utility were inflated 100x in the PNT.
        chosen = OfferAlternative.objects.filter(projectid=project, ischosen=True).first()
        trans_pct = chosen.transversalpercent if chosen else Decimal("0")
        prof_pct = chosen.profitpercent if chosen else Decimal("0")
        trans_factor = trans_pct / Decimal("100")
        prof_factor = prof_pct / Decimal("100")

        base_cost = [d + i for d, i in zip(direct_by_period, indirect_by_period)]
        retiro_by_period = [_round2(c * trans_factor) for c in base_cost]
        utility_by_period = [_round2(c * prof_factor) for c in base_cost]
        total_cost_by_period = [b + r + u for b, r, u in zip(base_cost, retiro_by_period, utility_by_period)]

        # Sale from WorkPlanEntry (PLANNED)
        sale_by_period = list(zeros)
        sale_qs = WorkPlanEntry.objects.filter(
            projectid=project, entrytype=WorkPlanEntryType.PLANNED,
        ).values('periodnumber').annotate(total=Coalesce(Sum('distributedamount'), Decimal("0")))
        for row in sale_qs:
            p = row['periodnumber']
            if 1 <= p <= N:
                sale_by_period[p - 1] = Decimal(row['total'] or 0)

        return {
            'sale_by_period': sale_by_period,
            'direct_by_period': direct_by_period,
            'indirect_by_period': indirect_by_period,
            'retiro_by_period': retiro_by_period,
            'utility_by_period': utility_by_period,
            'total_cost_by_period': total_cost_by_period,
            'direct_by_period_by_line': direct_by_period_by_line,
            'indirect_by_period_by_line': indirect_by_period_by_line,
            'lag_by_line': lag_by_line,
            'direct_total': sum(direct_by_period, Decimal("0")),
            'indirect_total': sum(indirect_by_period, Decimal("0")),
            'retiro_total': sum(retiro_by_period, Decimal("0")),
            'utility_total': sum(utility_by_period, Decimal("0")),
            'cost_total': sum(total_cost_by_period, Decimal("0")),
            'sale_total': sum(sale_by_period, Decimal("0")),
            'chosen_alternative_id': chosen.alternativeid if chosen else None,
            'transversalpercent': trans_pct,
            'profitpercent': prof_pct,
        }

    @staticmethod
    @transaction.atomic
    def autofill(project, *, strategy: str, only_empty: bool, scope: str) -> dict:
        """Populate CostDistribution according to strategy/scope.

        strategy: 'proportional_workplan' | 'uniform'
        only_empty: if True, don't overwrite existing rows (even derived)
        scope: 'all' | 'direct_only' | 'indirect_only' | 'family:<CODE>'
        """
        if strategy not in ('proportional_workplan', 'uniform'):
            raise ValueError(f"Unknown strategy: {strategy}")

        N = project.periodcount
        if N == 0:
            raise ValueError("Project has no periods — generate them first.")

        lines_affected = 0
        warnings = []

        # Gather lines to process
        # NOTE: `family:<CODE>` scope refers to a BreakdownCategoryCode pseudo-family
        # as exposed in `build_payload._build_families_hierarchy` (MATERIALS, HAULING,
        # MACHINERY, LABOR, SUBCONTRACTS, MINOR_TOOLS, PPE), NOT a ConceptFamily.code.
        # The frontend dropdown lists these categories, so the filter must match.
        breakdowns = []
        indirects = []
        if scope in ('all', 'direct_only') or scope.startswith('family:'):
            bq = UnitCostBreakdown.objects.filter(conceptid__projectid=project).select_related('conceptid')
            if scope.startswith('family:'):
                fam_code = scope.split(':', 1)[1]
                category_value = next(
                    (value for value, code, _name in CostDistributionService._DIRECT_CATEGORIES
                     if code == fam_code),
                    None,
                )
                if category_value is None:
                    raise ValueError(
                        f"family scope {fam_code!r} no es una categoría direct reconocida "
                        f"(esperado: MATERIALS, HAULING, MACHINERY, LABOR, SUBCONTRACTS, MINOR_TOOLS, PPE)"
                    )
                bq = bq.filter(categorycode=category_value)
            breakdowns = list(bq)
        if scope in ('all', 'indirect_only'):
            indirects = list(IndirectCostDetail.objects.filter(projectid=project))

        # Process breakdowns
        for bd in breakdowns:
            fractions = CostDistributionService._compute_line_fractions_breakdown(
                bd, project, strategy, N, warnings,
            )
            lines_affected += CostDistributionService._upsert_line_distribution(
                project=project, linetype=CostLineType.BREAKDOWN,
                breakdownid=bd.breakdownid, indirectcostid=None,
                fractions=fractions, only_empty=only_empty,
            )

        # Process indirects (always uniform within startmonth-endmonth range)
        for ind in indirects:
            fractions = CostDistributionService._compute_line_fractions_indirect(ind, N)
            lines_affected += CostDistributionService._upsert_line_distribution(
                project=project, linetype=CostLineType.INDIRECT,
                breakdownid=None, indirectcostid=ind.indirectcostid,
                fractions=fractions, only_empty=only_empty,
            )

        return {
            'lines_affected': lines_affected,
            'warnings': warnings,
        }

    @staticmethod
    def _compute_line_fractions_breakdown(bd, project, strategy, N, warnings):
        """Return list of N Decimals summing to 1.0."""
        if strategy == 'uniform':
            return [Decimal(1) / Decimal(N)] * N
        # proportional_workplan
        wp = WorkPlanEntry.objects.filter(
            conceptid=bd.conceptid, projectid=project, entrytype=WorkPlanEntryType.PLANNED,
        ).values('periodnumber', 'distributedamount')
        total = sum((Decimal(e['distributedamount']) for e in wp), Decimal("0"))
        if total <= 0:
            warnings.append(f"breakdown {bd.breakdownid}: no workplan — fallback uniform")
            return [Decimal(1) / Decimal(N)] * N
        buckets = [Decimal("0")] * N
        for e in wp:
            p = e['periodnumber']
            if 1 <= p <= N:
                buckets[p - 1] = Decimal(e['distributedamount']) / total
        return buckets

    @staticmethod
    def _compute_line_fractions_indirect(ind, N):
        start = ind.startmonth or 1
        end = ind.endmonth or N
        start = max(1, min(start, N))
        end = max(start, min(end, N))
        span = end - start + 1
        each = Decimal(1) / Decimal(span)
        buckets = [Decimal("0")] * N
        for p in range(start, end + 1):
            buckets[p - 1] = each
        return buckets

    @staticmethod
    def _upsert_line_distribution(*, project, linetype, breakdownid, indirectcostid, fractions, only_empty) -> int:
        """Return number of rows written/modified.

        Performance: fetches existing distributions for this line in 1 query,
        then issues at most 2 statements (1 bulk_update + 1 bulk_create).
        Previously did 1 SELECT + 1 INSERT/UPDATE per period — for a 26-period
        autofill across 551 lines that was ~28k queries.
        """
        base_lookup = {'projectid': project, 'linetype': linetype}
        if breakdownid:
            base_lookup['breakdownid_id'] = breakdownid
        else:
            base_lookup['indirectcostid_id'] = indirectcostid

        existing_by_period = {
            d.periodnumber: d
            for d in CostDistribution.objects.filter(**base_lookup)
        }

        to_update = []
        to_create = []
        affected = 0

        for idx, frac in enumerate(fractions):
            period = idx + 1
            existing = existing_by_period.get(period)
            if only_empty and existing is not None:
                continue  # preserve
            frac_q = frac.quantize(Decimal("0.00000001"))
            if existing is not None:
                existing.fraction = frac_q
                existing.isderived = True
                existing.version = F('version') + 1
                to_update.append(existing)
            else:
                to_create.append(CostDistribution(
                    **base_lookup,
                    periodnumber=period,
                    fraction=frac_q,
                    isderived=True,
                ))
            affected += 1

        if to_update:
            CostDistribution.objects.bulk_update(
                to_update, ['fraction', 'isderived', 'version'],
            )
        if to_create:
            CostDistribution.objects.bulk_create(to_create)

        return affected

    @staticmethod
    @transaction.atomic
    def apply_bulk_edits(project, *, user, edits: list = None, lag_edits: list = None) -> dict:
        """Apply multiple cell edits and/or per-line lag edits atomically.

        Args:
            edits: list of {lineid, linetype, periodnumber, fraction, expected_version}
            lag_edits: list of {lineid, linetype, paymentlagperiods, expected_lineversion}

        Validations:
            - lag_edits[].paymentlagperiods ∈ [0, 120] ∪ {None}
            - lag_edits[].linetype ∈ {'BREAKDOWN', 'INDIRECT'}
            - cell version mismatch raises VersionConflict
            - line version mismatch raises VersionConflict (with lineid + current_lineversion + kind='lag')

        On any conflict (cell or lag), aborts the entire transaction and raises
        VersionConflict(conflicts=[...]). Conflict items per type:
            cell:  {lineid, periodnumber, your_version, server_version, your_value,
                    server_value, server_modifiedby, server_modifiedon}
            lag:   {lineid, your_lineversion, server_lineversion, your_value, server_value,
                    kind: 'lag'}

        Returns: {
            'updated': int,             # cells changed
            'new_versions': dict,       # cell key → new version
            'lag_updated': int,         # lines whose lag changed
            'new_lineversions': dict,   # lineid → new lineversion
        }

        Note: SQLite (used in development) does not support row-level locking, so
        select_for_update() is a no-op at the SQL level in dev. The all-or-nothing
        atomicity via transaction.atomic() still holds. PostgreSQL in production
        enforces proper row-level locking.
        """
        edits = edits or []
        lag_edits = lag_edits or []

        # -------- Validate lag edits up front --------
        for le in lag_edits:
            plp = le.get('paymentlagperiods')
            if plp is not None:
                if not isinstance(plp, int) or plp < 0 or plp > 120:
                    raise ValueError(
                        f"paymentlagperiods out of range [0,120] or null: {plp!r}"
                    )
            if le['linetype'] not in ('BREAKDOWN', 'INDIRECT'):
                raise ValueError(
                    f"linetype must be BREAKDOWN or INDIRECT, got {le['linetype']!r}"
                )

        # -------- Detect conflicts (cells AND lags) before any mutation --------
        conflicts = []

        # SELECT FOR UPDATE on the affected rows to serialize with other writers
        for edit in edits:
            lt = CostLineType.BREAKDOWN if edit['linetype'] == 'BREAKDOWN' else CostLineType.INDIRECT
            lookup = {'projectid': project, 'linetype': lt, 'periodnumber': edit['periodnumber']}
            if lt == CostLineType.BREAKDOWN:
                lookup['breakdownid_id'] = edit['lineid']
            else:
                lookup['indirectcostid_id'] = edit['lineid']
            existing = CostDistribution.objects.select_for_update().filter(**lookup).first()
            if existing is None:
                # Create with version=0; expected must also be 0
                if edit.get('expected_version', 0) != 0:
                    conflicts.append({
                        'lineid': edit['lineid'], 'periodnumber': edit['periodnumber'],
                        'your_version': edit.get('expected_version'),
                        'server_version': 0,
                        'your_value': float(edit['fraction']),
                        'server_value': None,
                        'server_modifiedby': None,
                        'server_modifiedon': None,
                    })
                continue
            if existing.version != edit['expected_version']:
                conflicts.append({
                    'lineid': edit['lineid'], 'periodnumber': edit['periodnumber'],
                    'your_version': edit['expected_version'],
                    'server_version': existing.version,
                    'your_value': float(edit['fraction']),
                    'server_value': float(existing.fraction),
                    'server_modifiedby': str(existing.modifiedby) if existing.modifiedby else None,
                    'server_modifiedon': existing.modifiedon.isoformat(),
                })

        # ===== LAG CONFLICT DETECTION =====
        for le in lag_edits:
            if le['linetype'] == 'BREAKDOWN':
                line = UnitCostBreakdown.objects.select_for_update().filter(pk=le['lineid']).first()
            else:
                line = IndirectCostDetail.objects.select_for_update().filter(pk=le['lineid']).first()
            if line is None:
                raise ValueError(f"Line not found: {le['lineid']}")
            if line.lineversion != le['expected_lineversion']:
                conflicts.append({
                    'lineid': le['lineid'],
                    'kind': 'lag',
                    'your_lineversion': le['expected_lineversion'],
                    'server_lineversion': line.lineversion,
                    'your_value': le.get('paymentlagperiods'),
                    'server_value': line.paymentlagperiods,
                })

        if conflicts:
            raise VersionConflict(conflicts)

        # -------- No conflicts — apply all --------

        # Apply cell edits
        new_versions = {}
        for edit in edits:
            lt = CostLineType.BREAKDOWN if edit['linetype'] == 'BREAKDOWN' else CostLineType.INDIRECT
            lookup = {'projectid': project, 'linetype': lt, 'periodnumber': edit['periodnumber']}
            if lt == CostLineType.BREAKDOWN:
                lookup['breakdownid_id'] = edit['lineid']
            else:
                lookup['indirectcostid_id'] = edit['lineid']
            existing = CostDistribution.objects.filter(**lookup).first()
            if existing is None:
                new = CostDistribution.objects.create(
                    **lookup,
                    fraction=Decimal(str(edit['fraction'])).quantize(Decimal("0.00000001")),
                    isderived=False, version=1, modifiedby=user,
                )
                new_versions[f"{edit['lineid']}:{edit['periodnumber']}"] = 1
            else:
                existing.fraction = Decimal(str(edit['fraction'])).quantize(Decimal("0.00000001"))
                existing.isderived = False
                existing.version += 1
                existing.modifiedby = user
                existing.save(update_fields=['fraction', 'isderived', 'version', 'modifiedby', 'modifiedon'])
                new_versions[f"{edit['lineid']}:{edit['periodnumber']}"] = existing.version

        # Apply lag edits
        new_lineversions = {}
        for le in lag_edits:
            if le['linetype'] == 'BREAKDOWN':
                line = UnitCostBreakdown.objects.get(pk=le['lineid'])
            else:
                line = IndirectCostDetail.objects.get(pk=le['lineid'])
            line.paymentlagperiods = le.get('paymentlagperiods')
            line.lineversion += 1
            # UnitCostBreakdown has no modifiedby; IndirectCostDetail extends AuditMixin
            if le['linetype'] == 'BREAKDOWN':
                line.save(update_fields=['paymentlagperiods', 'lineversion', 'modifiedon'])
            else:  # INDIRECT
                line.modifiedby = user
                line.save(update_fields=['paymentlagperiods', 'lineversion', 'modifiedby', 'modifiedon'])
            new_lineversions[str(le['lineid'])] = line.lineversion

        return {
            'updated': len(edits),
            'new_versions': new_versions,
            'lag_updated': len(lag_edits),
            'new_lineversions': new_lineversions,
        }

    @staticmethod
    @transaction.atomic
    def reset_line(project, *, lineid: str, linetype: str) -> dict:
        lt = CostLineType.BREAKDOWN if linetype == 'BREAKDOWN' else CostLineType.INDIRECT
        filt = {'projectid': project, 'linetype': lt}
        if lt == CostLineType.BREAKDOWN:
            filt['breakdownid_id'] = lineid
        else:
            filt['indirectcostid_id'] = lineid
        CostDistribution.objects.filter(**filt).delete()

        # Regenerate via autofill (default: proportional_workplan falls back to uniform)
        N = project.periodcount
        warnings = []
        if lt == CostLineType.BREAKDOWN:
            bd = UnitCostBreakdown.objects.get(breakdownid=lineid)
            fractions = CostDistributionService._compute_line_fractions_breakdown(
                bd, project, 'proportional_workplan', N, warnings,
            )
            CostDistributionService._upsert_line_distribution(
                project=project, linetype=lt,
                breakdownid=lineid, indirectcostid=None,
                fractions=fractions, only_empty=False,
            )
        else:
            ind = IndirectCostDetail.objects.get(indirectcostid=lineid)
            fractions = CostDistributionService._compute_line_fractions_indirect(ind, N)
            CostDistributionService._upsert_line_distribution(
                project=project, linetype=lt,
                breakdownid=None, indirectcostid=lineid,
                fractions=fractions, only_empty=False,
            )
        return {'reset': True, 'warnings': warnings}

    @staticmethod
    def build_payload(project) -> dict:
        """Assemble the full GET /cost-distribution response."""
        rollups = CostDistributionService.compute_rollups(project)
        periods = list(ProjectionPeriod.objects.filter(projectid=project).order_by('periodnumber'))

        families = CostDistributionService._build_families_hierarchy(project, rollups)

        # Frontend-friendly flattening
        rollups_payload = {
            'direct_by_period': [float(x) for x in rollups['direct_by_period']],
            'indirect_by_period': [float(x) for x in rollups['indirect_by_period']],
            'retiro_by_period': [float(x) for x in rollups['retiro_by_period']],
            'utility_by_period': [float(x) for x in rollups['utility_by_period']],
            'total_cost_by_period': [float(x) for x in rollups['total_cost_by_period']],
            'sale_by_period': [float(x) for x in rollups['sale_by_period']],
            'margin_by_period': [
                float(s - c) for s, c in zip(rollups['sale_by_period'], rollups['total_cost_by_period'])
            ],
        }
        sale_total = rollups['sale_total']
        cost_total = rollups['cost_total']
        totals = {
            'direct_total': float(rollups['direct_total']),
            'indirect_total': float(rollups['indirect_total']),
            'retiro_total': float(rollups['retiro_total']),
            'utility_total': float(rollups['utility_total']),
            'cost_total': float(cost_total),
            'sale_total': float(sale_total),
            'margin_total': float(sale_total - cost_total),
            'margin_pct': float((sale_total - cost_total) / sale_total) if sale_total else 0.0,
        }
        if rollups['chosen_alternative_id']:
            alt = OfferAlternative.objects.get(alternativeid=rollups['chosen_alternative_id'])
            chosen = {
                'alternativeid': str(alt.alternativeid),
                'name': alt.name,
                'transversalpercent': float(alt.transversalpercent),
                'profitpercent': float(alt.profitpercent),
            }
        else:
            chosen = {
                'alternativeid': None, 'name': None,
                'transversalpercent': 0.0, 'profitpercent': 0.0,
            }

        return {
            'periods': [_period_dto(p) for p in periods],
            'families': families,
            'rollups': rollups_payload,
            'totals': totals,
            'chosen_alternative': chosen,
        }

    # Canonical direct-cost categories, in the order the Excel reference uses them.
    # Maps BreakdownCategoryCode values → display name.
    _DIRECT_CATEGORIES = [
        (1, 'MATERIALS',    'MATERIALES'),
        (2, 'HAULING',      'ACARREOS'),
        (3, 'MACHINERY',    'MAQUINARIA'),
        (4, 'LABOR',        'MANO DE OBRA'),
        (5, 'SUBCONTRACTS', 'SUBCONTRATOS'),
        (6, 'MINOR_TOOLS',  'HERRAMIENTA MENOR'),
        (7, 'PPE',          'EPP'),
    ]

    @staticmethod
    def _build_families_hierarchy(project, rollups) -> list:
        """Build families array: directs grouped by breakdown category (MATERIALES,
        MAQUINARIA, ...), indirects grouped by IndirectCostDetail.categorycode using
        the `area` field for the display name.

        Performance: prefetches *all* CostDistribution rows for the project in one
        query and groups them in Python. Previously each line + each family rollup
        triggered its own filter, producing N+1 (~551 queries for a 56-concept
        project). Now it's exactly 3 queries: breakdowns, indirects, distributions.
        """
        N = project.periodcount
        families = []

        # Single fetch of CostDistribution rows for the whole project.
        dists_by_breakdown: dict = defaultdict(list)
        dists_by_indirect: dict = defaultdict(list)
        for d in CostDistribution.objects.filter(projectid=project):
            if d.breakdownid_id is not None:
                dists_by_breakdown[d.breakdownid_id].append(d)
            elif d.indirectcostid_id is not None:
                dists_by_indirect[d.indirectcostid_id].append(d)

        # DIRECT: group UnitCostBreakdown by categorycode (not by concept family)
        bds_by_cat = defaultdict(list)
        bds = UnitCostBreakdown.objects.filter(
            conceptid__projectid=project,
        ).select_related('conceptid')
        for bd in bds:
            bds_by_cat[bd.categorycode].append(bd)

        for cat_value, cat_code, cat_name in CostDistributionService._DIRECT_CATEGORIES:
            bd_list = bds_by_cat.get(cat_value, [])
            if not bd_list:
                continue  # omit empty categories so the UI stays tight
            # bd.amount is per-unit-of-concept; project-level = bd.amount × concept.quantity.
            # Same convention used by compute_rollups so the family/line totals reconcile
            # with the project-wide direct_total.
            project_amounts_by_id = {
                bd.breakdownid: (bd.amount or Decimal('0')) * (bd.conceptid.quantity or Decimal('0'))
                for bd in bd_list
            }
            rollups_by_period = CostDistributionService._family_rollup_from_dict(
                bd_list, N,
                amounts_by_id=project_amounts_by_id,
                dists_by_id=dists_by_breakdown,
                id_attr='breakdownid',
            )
            total_amount = sum(project_amounts_by_id.values(), Decimal("0"))
            families.append({
                'code': cat_code,
                'name': cat_name,
                'categorytype': 'DIRECT',
                'totalamount': float(total_amount),
                'rollups_by_period': [float(x) for x in rollups_by_period],
                'lines': [
                    CostDistributionService._line_payload_breakdown(
                        bd, N, dists_by_breakdown.get(bd.breakdownid, []),
                    )
                    for bd in bd_list
                ],
            })

        # INDIRECT: group by categorycode; use `area` field for the family name
        ind_by_cat = defaultdict(list)
        for ind in IndirectCostDetail.objects.filter(projectid=project):
            ind_by_cat[ind.categorycode or 'OTHER'].append(ind)
        for cat in sorted(ind_by_cat.keys()):
            inds = ind_by_cat[cat]
            rollups_by_period = CostDistributionService._family_rollup_from_dict(
                inds, N,
                amounts_by_id={ind.indirectcostid: ind.amount for ind in inds},
                dists_by_id=dists_by_indirect,
                id_attr='indirectcostid',
            )
            total_amount = sum((ind.amount for ind in inds), Decimal("0"))
            name = (inds[0].area or '').strip() or f'Indirecto {cat}'
            families.append({
                'code': cat,
                'name': name,
                'categorytype': 'INDIRECT',
                'totalamount': float(total_amount),
                'rollups_by_period': [float(x) for x in rollups_by_period],
                'lines': [
                    CostDistributionService._line_payload_indirect(
                        ind, N, dists_by_indirect.get(ind.indirectcostid, []),
                    )
                    for ind in inds
                ],
            })
        return families

    @staticmethod
    def _family_rollup_from_dict(lines, N, *, amounts_by_id, dists_by_id, id_attr):
        """SUMPRODUCT of given lines x fractions per period using a pre-fetched dict.

        Replaces the legacy ``_family_rollup`` which queried CostDistribution per
        family. ``id_attr`` is ``'breakdownid'`` or ``'indirectcostid'``.
        """
        buckets = [Decimal("0")] * N
        if not lines:
            return buckets
        for line in lines:
            line_id = getattr(line, id_attr)
            amt = amounts_by_id.get(line_id, Decimal("0"))
            for d in dists_by_id.get(line_id, []):
                if 1 <= d.periodnumber <= N:
                    buckets[d.periodnumber - 1] += amt * d.fraction
        return buckets

    @staticmethod
    def _line_payload_breakdown(bd, N, dists):
        """``dists`` is the pre-fetched list of CostDistribution rows for this line.

        ``totalamount`` is project-level: ``bd.amount × concept.quantity``. Matches
        the convention used by compute_rollups so the cell $ values rendered by the
        frontend (lineTotal × fraction) reconcile with the project's direct_total.
        """
        dists_sorted = sorted(dists, key=lambda d: d.periodnumber)
        cells = [
            {'periodnumber': d.periodnumber, 'fraction': float(d.fraction),
             'isderived': d.isderived, 'version': d.version}
            for d in dists_sorted
        ]
        checksum = sum((d.fraction for d in dists_sorted), Decimal("0"))
        project_amount = (bd.amount or Decimal('0')) * (bd.conceptid.quantity or Decimal('0'))
        return {
            'lineid': str(bd.breakdownid),
            'linetype': 'BREAKDOWN',
            'description': bd.description,
            'unit': bd.unit,
            'totalamount': float(project_amount),
            'paymentlagperiods': bd.paymentlagperiods,
            'lineversion': bd.lineversion,
            'distribution': cells,
            'checksum': float(checksum),
        }

    @staticmethod
    def _line_payload_indirect(ind, N, dists):
        """``dists`` is the pre-fetched list of CostDistribution rows for this line."""
        dists_sorted = sorted(dists, key=lambda d: d.periodnumber)
        cells = [
            {'periodnumber': d.periodnumber, 'fraction': float(d.fraction),
             'isderived': d.isderived, 'version': d.version}
            for d in dists_sorted
        ]
        checksum = sum((d.fraction for d in dists_sorted), Decimal("0"))
        return {
            'lineid': str(ind.indirectcostid),
            'linetype': 'INDIRECT',
            'description': ind.description,
            'unit': '',
            'totalamount': float(ind.amount),
            'paymentlagperiods': ind.paymentlagperiods,
            'lineversion': ind.lineversion,
            'distribution': cells,
            'checksum': float(checksum),
        }


class VersionConflict(Exception):
    def __init__(self, conflicts):
        self.conflicts = conflicts
        super().__init__(f"{len(conflicts)} version conflicts")


def _round2(x: Decimal) -> Decimal:
    return x.quantize(Decimal("0.01"))


def _period_dto(p):
    return {
        'periodid': str(p.periodid),
        'periodnumber': p.periodnumber,
        'periodlabel': p.periodlabel,
        'startdate': p.startdate.isoformat(),
        'enddate': p.enddate.isoformat(),
        'periodtype': p.periodtype,
    }


# =============================================================================
# Presence Service
# =============================================================================

from apps.proyeccion.models import DistributionPresence


class PresenceService:
    STALE_THRESHOLD_MINUTES = 2

    @staticmethod
    def heartbeat(project, user, *, mode: str):
        obj, _ = DistributionPresence.objects.update_or_create(
            projectid=project, userid=user,
            defaults={'mode': mode},
        )
        return obj

    @staticmethod
    def list_active(project):
        from django.utils import timezone
        from datetime import timedelta
        cutoff = timezone.now() - timedelta(minutes=PresenceService.STALE_THRESHOLD_MINUTES)
        return list(
            DistributionPresence.objects
            .filter(projectid=project, last_seen__gte=cutoff)
            .select_related('userid')
            .order_by('-last_seen')
        )


# =============================================================================
# Estimation Financial Settings Service
# =============================================================================


class EstimationFinancialSettingsService:
    """Manage 1:1 financial settings for an EstimationProject."""

    _WHITELIST = frozenset({
        'advanceamountnotax', 'advanceentryperiod', 'advanceamortizationrate',
        'imssretentionrate', 'otherretentionrate', 'retentionreturnperiod',
        'directpaymentlag', 'indirectpaymentlag',
        'financecostrate',
    })

    @staticmethod
    def get_or_create(project_id):
        """Idempotent. Materializes defaults on first call."""
        project = EstimationProject.objects.get(pk=project_id)
        settings, _created = EstimationFinancialSettings.objects.get_or_create(projectid=project)
        return settings

    @classmethod
    def update(cls, project_id, dto, user=None):
        """Apply only whitelisted fields. Ignore unknown keys silently."""
        settings = cls.get_or_create(project_id)
        for key, value in (dto or {}).items():
            if key in cls._WHITELIST:
                setattr(settings, key, value)
        if user is not None:
            settings.modifiedby = user
        settings.save()
        return settings


# =============================================================================
# Estimation Billing Rule Service
# =============================================================================


class EstimationBillingRuleService:
    """Manage N billing tranches per estimation project."""

    _SUM_TOLERANCE = Decimal('0.0001')
    _MAX_RULES = 10

    @staticmethod
    def list(project_id):
        return list(
            EstimationBillingRule.objects.filter(projectid=project_id).order_by('sequence')
        )

    @classmethod
    @transaction.atomic
    def replace(cls, project_id, rules, user=None):
        """All-or-nothing replacement. Validates count, sum, sequences."""
        if not rules:
            raise ValueError('Debe proporcionar al menos 1 regla de facturación.')
        if len(rules) > cls._MAX_RULES:
            raise ValueError(f'Máximo {cls._MAX_RULES} reglas permitidas.')

        sequences = [r['sequence'] for r in rules]
        if len(set(sequences)) != len(sequences):
            raise ValueError('Las secuencias deben ser únicas.')

        total = sum((Decimal(str(r['percent'])) for r in rules), Decimal('0'))
        if abs(total - Decimal('1')) > cls._SUM_TOLERANCE:
            raise ValueError(
                f'La suma de porcentajes debe ser 100% (±0.01%). Suma actual: {total * 100:.4f}%.'
            )

        project = EstimationProject.objects.get(pk=project_id)
        EstimationBillingRule.objects.filter(projectid=project).delete()
        created = []
        for r in rules:
            instance = EstimationBillingRule.objects.create(
                projectid=project,
                sequence=r['sequence'],
                percent=Decimal(str(r['percent'])),
                lagperiods=int(r['lagperiods']),
            )
            if user is not None:
                instance.createdby = user
                instance.modifiedby = user
                instance.save(update_fields=['createdby', 'modifiedby'])
            created.append(instance)
        return sorted(created, key=lambda x: x.sequence)


# =============================================================================
# Estimation PNT (Proyección de Necesidad de Tesorería) Calculator
# =============================================================================

from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional
from django.utils import timezone

ZERO = Decimal('0')


@dataclass
class _PNTRow:
    code: str
    label: str
    section: str  # 'RESULTADO' | 'COBROS' | 'PAGOS' | 'CAJA'
    values: list
    emphasis: bool = False
    # Flow that was pushed past the project horizon by lag rules. Currently
    # populated for COBRO_FACTURACION, ANTICIPO_CONCEDIDO, DEVOLUCION and the
    # PAGOS_* family, plus aggregations (COBRO_TOTAL, PAGOS_TOTALES, CAJA_MES).
    # Stays ZERO for devengado rows (RESULTADO section) and cumulative rows
    # (CAJA_ACUMULADA, SALDO_ANTICIPO).
    out_of_horizon: Decimal = ZERO
    # sum(values) + out_of_horizon for flow rows; ZERO for cumulative rows
    # (CAJA_ACUMULADA, SALDO_ANTICIPO) where summing a stock is meaningless.
    total: Decimal = ZERO


@dataclass
class _PNTReport:
    projectid: object
    granularity: str
    periods: list
    rows: list
    stats: dict
    generated_at: datetime


@dataclass
class _InlineBillingRule:
    sequence: int
    percent: Decimal
    lagperiods: int


class EstimationPNTCalculator:
    """Derive PNT from CostDistribution rollups + financial settings + billing rules."""

    _OVERRIDE_ALLOWLIST = frozenset({
        'imssretentionrate', 'otherretentionrate', 'retentionreturnperiod',
        'advanceamountnotax', 'advanceentryperiod', 'advanceamortizationrate',
        'directpaymentlag', 'indirectpaymentlag', 'financecostrate',
    })

    _FLOW_CODES = frozenset({
        'RESULTADO', 'PRODUCCION', 'COSTO_DIRECTO', 'COSTO_INDIRECTO',
        'RETIRO_TRANSV_RES', 'RETIRO_UTIL_RES',
        'COBRO_TOTAL', 'COBRO_FACTURACION', 'ANTICIPO_CONCEDIDO', 'ANTICIPO_AMORT',
        'RET_IMSS', 'OTRAS_RET', 'DEVOLUCION',
        'PAGOS_DIRECTO', 'PAGOS_INDIRECTO', 'RETIRO_TRANSV', 'RETIRO_UTILIDADES',
        'PAGOS_TOTALES', 'CAJA_MES', 'COSTO_FINANCIERO',
    })
    _CUMULATIVE_CODES = frozenset({'CAJA_ACUMULADA', 'SALDO_ANTICIPO'})

    def __init__(self, project_id):
        self.project = EstimationProject.objects.get(pk=project_id)
        self.periods = list(
            ProjectionPeriod.objects.filter(projectid=project_id).order_by('periodnumber')
        )
        self.N = len(self.periods)
        if self.N == 0:
            raise ValueError(
                'No hay periodos. Inicializa el Plan de Obra (Paso 9) antes de consultar el PNT.'
            )
        self.settings = EstimationFinancialSettingsService.get_or_create(project_id)
        self.billing_rules = self._load_billing_rules()
        self.rollups = CostDistributionService.compute_rollups(self.project)

    def _load_billing_rules(self):
        persisted = EstimationBillingRuleService.list(self.project.estimationprojectid)
        if persisted:
            return [
                _InlineBillingRule(sequence=r.sequence, percent=r.percent, lagperiods=r.lagperiods)
                for r in persisted
            ]
        # Sin reglas configuradas → cobro inmediato (cobro_fact[i] = produccion[i]).
        # _compute_cobro_facturacion detecta lista vacía y usa este fallback.
        return []

    def _apply_overrides(self, overrides):
        """Mutate self.settings / self.billing_rules in-memory only."""
        for key, value in (overrides or {}).items():
            if key in self._OVERRIDE_ALLOWLIST:
                if isinstance(value, (int, float, str)):
                    value = Decimal(str(value)) if key not in {
                        'retentionreturnperiod', 'advanceentryperiod',
                        'directpaymentlag', 'indirectpaymentlag',
                    } else int(value)
                setattr(self.settings, key, value)
        if 'billing_rules' in (overrides or {}):
            self.billing_rules = [
                _InlineBillingRule(
                    sequence=int(r['sequence']),
                    percent=Decimal(str(r['percent'])),
                    lagperiods=int(r['lagperiods']),
                )
                for r in overrides['billing_rules']
            ]

    def compute(self, overrides=None, granularity='period'):
        if granularity not in ('period', 'month'):
            raise ValueError(f"granularity must be 'period' or 'month', got {granularity!r}")
        if overrides:
            self._apply_overrides(overrides)

        # Base vectors from rollups
        produccion = list(self.rollups['sale_by_period'])
        costo_directo_neg = [-x for x in self.rollups['direct_by_period']]
        costo_indirecto_neg = [-x for x in self.rollups['indirect_by_period']]
        retiro_transv_neg = [-x for x in self.rollups['retiro_by_period']]
        retiro_util_neg = [-x for x in self.rollups['utility_by_period']]

        # ---- COBROS ----
        cobro_facturacion, cobros_fuera = self._compute_cobro_facturacion(produccion)
        anticipo_concedido, anticipo_concedido_fuera = self._single_period_vector(
            self.settings.advanceamountnotax, self.settings.advanceentryperiod,
        )
        anticipo_amortizado, advance_fully_amortized_at = self._compute_anticipo_amortizado(
            cobro_facturacion,
        )
        retencion_imss = [-self.settings.imssretentionrate * cf for cf in cobro_facturacion]
        otras_retencion = [-self.settings.otherretentionrate * cf for cf in cobro_facturacion]
        devolucion, devolucion_fuera = self._compute_devolucion(retencion_imss, otras_retencion)
        saldo_anticipo = self._cumsum(anticipo_amortizado)
        cobro_total = [
            anticipo_concedido[i] + cobro_facturacion[i] + anticipo_amortizado[i]
            + retencion_imss[i] + otras_retencion[i] + devolucion[i]
            for i in range(self.N)
        ]
        cobro_total_fuera = cobros_fuera + anticipo_concedido_fuera + devolucion_fuera
        # NOTE: SALDO_ANTICIPO is a cumulative running balance derived from
        # ANTICIPO_AMORT and is reported for visibility only. Including it in
        # COBRO_TOTAL would double-count amortization (this was the bug in the
        # discarded operations-side implementation). Locked in by
        # test_cobro_total_excludes_saldo_anticipo.

        # ---- PAGOS ----
        pagos_directo, pagos_dir_fuera = self._apply_pagos_lag(
            self.rollups['direct_by_period_by_line'],
            self.rollups['lag_by_line'],
            self.settings.directpaymentlag,
        )
        pagos_indirecto, pagos_ind_fuera = self._apply_pagos_lag(
            self.rollups['indirect_by_period_by_line'],
            self.rollups['lag_by_line'],
            self.settings.indirectpaymentlag,
        )
        pagos_totales = [
            pagos_directo[i] + pagos_indirecto[i] + retiro_transv_neg[i] + retiro_util_neg[i]
            for i in range(self.N)
        ]
        # Retiros are devengado (no lag) so their contribution to fuera is ZERO.
        pagos_totales_fuera = pagos_dir_fuera + pagos_ind_fuera

        # ---- CAJA ----
        caja_mes = [cobro_total[i] + pagos_totales[i] for i in range(self.N)]
        caja_mes_fuera = cobro_total_fuera + pagos_totales_fuera
        caja_acumulada = self._cumsum(caja_mes)
        costo_financiero = [
            ca * self.settings.financecostrate if ca < 0 else ZERO
            for ca in caja_acumulada
        ]

        # ---- RESULTADO ----
        resultado = [
            produccion[i] + costo_directo_neg[i] + costo_indirecto_neg[i]
            + retiro_transv_neg[i] + retiro_util_neg[i]
            for i in range(self.N)
        ]

        rows = [
            # RESULTADO section
            _PNTRow('RESULTADO', 'Resultado', 'RESULTADO', resultado, emphasis=True),
            _PNTRow('PRODUCCION', 'Producción', 'RESULTADO', produccion),
            _PNTRow('COSTO_DIRECTO', 'Costo Directo', 'RESULTADO', costo_directo_neg),
            _PNTRow('COSTO_INDIRECTO', 'Costo Indirecto', 'RESULTADO', costo_indirecto_neg),
            _PNTRow('RETIRO_TRANSV_RES', 'Transversales', 'RESULTADO', retiro_transv_neg),
            _PNTRow('RETIRO_UTIL_RES', 'Utilidades', 'RESULTADO', retiro_util_neg),
            # COBROS section
            _PNTRow('COBRO_TOTAL', 'Cobro Total sin IVA', 'COBROS', cobro_total, emphasis=True),
            _PNTRow('COBRO_FACTURACION', 'Cobro Facturación', 'COBROS', cobro_facturacion),
            _PNTRow('ANTICIPO_CONCEDIDO', 'Anticipo Concedido', 'COBROS', anticipo_concedido),
            _PNTRow('ANTICIPO_AMORT', 'Anticipo Amortizado', 'COBROS', anticipo_amortizado),
            _PNTRow('RET_IMSS', 'Retenciones IMSS', 'COBROS', retencion_imss),
            _PNTRow('OTRAS_RET', 'Otras Retenciones', 'COBROS', otras_retencion),
            _PNTRow('DEVOLUCION', 'Devolución Retenciones', 'COBROS', devolucion),
            _PNTRow('SALDO_ANTICIPO', 'Saldo Anticipo', 'COBROS', saldo_anticipo),
            # PAGOS section
            _PNTRow('PAGOS_DIRECTO', 'Pagos Costo Directo', 'PAGOS', pagos_directo),
            _PNTRow('PAGOS_INDIRECTO', 'Pagos Costos Indirectos', 'PAGOS', pagos_indirecto),
            _PNTRow('RETIRO_TRANSV', 'Retiro Transversales', 'PAGOS', retiro_transv_neg),
            _PNTRow('RETIRO_UTILIDADES', 'Retiro Utilidades', 'PAGOS', retiro_util_neg),
            _PNTRow('PAGOS_TOTALES', 'Pagos Totales', 'PAGOS', pagos_totales, emphasis=True),
            # CAJA section
            _PNTRow('CAJA_MES', 'Caja Mes', 'CAJA', caja_mes, emphasis=True),
            _PNTRow('CAJA_ACUMULADA', 'Caja Acumulada (PNT)', 'CAJA', caja_acumulada, emphasis=True),
            _PNTRow('COSTO_FINANCIERO', 'Costo Financiero', 'CAJA', costo_financiero),
        ]

        # Per-row out-of-horizon mapping. Devengado rows (RESULTADO section), no-lag
        # retiro rows in PAGOS, and rows derived purely from in-horizon flows have
        # ZERO. Cumulative rows (handled below) are explicitly excluded from totals.
        out_of_horizon_by_code = {
            'COBRO_FACTURACION': cobros_fuera,
            'ANTICIPO_CONCEDIDO': anticipo_concedido_fuera,
            'DEVOLUCION': devolucion_fuera,
            'COBRO_TOTAL': cobro_total_fuera,
            'PAGOS_DIRECTO': pagos_dir_fuera,
            'PAGOS_INDIRECTO': pagos_ind_fuera,
            'PAGOS_TOTALES': pagos_totales_fuera,
            'CAJA_MES': caja_mes_fuera,
        }
        for r in rows:
            r.out_of_horizon = out_of_horizon_by_code.get(r.code, ZERO)
            if r.code in self._CUMULATIVE_CODES:
                # Stocks: total of a running balance is meaningless; leave ZERO.
                r.total = ZERO
            else:
                r.total = sum(r.values, ZERO) + r.out_of_horizon

        chosen = self.rollups.get('chosen_alternative_id')
        advance_fully_amortized_period = (
            self.periods[advance_fully_amortized_at].periodlabel
            if advance_fully_amortized_at is not None
            else None
        )
        stats = {
            'pnt_min': min(caja_acumulada) if caja_acumulada else ZERO,
            'pnt_max': max(caja_acumulada) if caja_acumulada else ZERO,
            'pnt_avg': (sum(caja_acumulada, ZERO) / Decimal(self.N)) if self.N else ZERO,
            'total_costo_financiero': sum(costo_financiero, ZERO),
            'cobros_fuera_horizonte': cobros_fuera,
            'pagos_fuera_horizonte': pagos_dir_fuera + pagos_ind_fuera,
            'chosen_alternative_id': chosen,
            'transversalpercent_aplicado': self.rollups.get('transversalpercent', ZERO),
            'profitpercent_aplicado': self.rollups.get('profitpercent', ZERO),
            'advance_fully_amortized_period': advance_fully_amortized_period,
        }

        periods_out = [
            {'label': p.periodlabel, 'startdate': p.startdate, 'enddate': p.enddate}
            for p in self.periods
        ]

        if granularity == 'month':
            periods_out, rows = self._aggregate_monthly(rows)

        return _PNTReport(
            projectid=self.project.estimationprojectid,
            granularity=granularity,
            periods=periods_out,
            rows=rows,
            stats=stats,
            generated_at=timezone.now(),
        )

    def _apply_lag(self, vec, lag):
        out = [ZERO] * self.N
        fuera = ZERO
        for i in range(self.N):
            target = i + (lag or 0)
            if 0 <= target < self.N:
                out[target] += vec[i]
            else:
                fuera += vec[i]
        return out, fuera

    def _apply_pagos_lag(self, by_line, lag_by_line, default_lag):
        """Apply per-line lag to costs. Falls back to default_lag if lag_by_line[id] is None.

        Args:
            by_line: dict {lineid → vector positivo de N decimales}
            lag_by_line: dict {lineid → int | None}
            default_lag: int — fallback when lag_by_line[lineid] is None

        Returns:
            (out_vec, fuera_total) — output vector with shifted negative amounts;
            amounts pushed beyond N go to fuera.
        """
        out = [ZERO] * self.N
        fuera = ZERO
        for lineid, line_vec in by_line.items():
            line_lag = lag_by_line.get(lineid)
            lag = int(line_lag) if line_lag is not None else (default_lag or 0)
            for i in range(self.N):
                amount = -line_vec[i]
                target = i + lag
                if 0 <= target < self.N:
                    out[target] += amount
                else:
                    fuera += amount
        return out, fuera

    def _aggregate_monthly(self, rows):
        # Defensive: catches typos in row codes when taxonomy is extended
        known = self._FLOW_CODES | self._CUMULATIVE_CODES
        unknown = {r.code for r in rows} - known
        if unknown:
            raise AssertionError(f'Unknown row codes in _aggregate_monthly: {sorted(unknown)}')

        # Group periods by (year, month)
        groups = []
        current_key = None
        for i, p in enumerate(self.periods):
            key = (p.startdate.year, p.startdate.month)
            if key != current_key:
                groups.append([i])
                current_key = key
            else:
                groups[-1].append(i)

        new_periods = []
        for g in groups:
            first = self.periods[g[0]]
            last = self.periods[g[-1]]
            new_periods.append({
                'label': f'{first.startdate.year}-{first.startdate.month:02d}',
                'startdate': first.startdate,
                'enddate': last.enddate,
            })

        new_rows = []
        for r in rows:
            if r.code in self._CUMULATIVE_CODES:
                agg = [r.values[g[-1]] for g in groups]
            else:
                agg = [sum((r.values[i] for i in g), ZERO) for g in groups]
            # out_of_horizon and total are granularity-invariant — they describe
            # the row as a whole, not a particular time bucket.
            new_rows.append(_PNTRow(
                r.code, r.label, r.section, agg, emphasis=r.emphasis,
                out_of_horizon=r.out_of_horizon, total=r.total,
            ))
        return new_periods, new_rows

    # --- internals ---

    def _compute_cobro_facturacion(self, produccion):
        """Milestone billing model — mirrors the Excel "PNT" sheet (R17-R19).

        Each ``EstimationBillingRule`` represents a single project-level
        milestone: ``percent × Σ produccion`` is collected at the absolute
        period number indicated by ``lagperiods`` (1-indexed from project
        start). Rules whose period falls outside the project horizon go to
        the "fuera de horizonte" bucket.

        When no rules are configured, falls back to immediate billing
        (cobro_fact[i] = produccion[i]) so the calculator stays usable
        before the user sets up the milestone schedule.

        Excel formula reference (sheet "PNT"):
            E17 = IF($D17 = E$5, $CQ$9 * $B17, 0)
                       │            │           │
                       periodo del  producción  porcentaje
                       milestone    TOTAL       de la regla
        """
        if not self.billing_rules:
            return list(produccion), ZERO

        total_prod = sum(produccion, ZERO)
        out = [ZERO] * self.N
        fuera = ZERO
        for rule in self.billing_rules:
            amount = total_prod * rule.percent
            target = int(rule.lagperiods) - 1  # 1-indexed absolute period -> 0-indexed bucket
            if 0 <= target < self.N:
                out[target] += amount
            else:
                fuera += amount
        return out, fuera

    def _compute_anticipo_amortizado(self, cobro_facturacion):
        """Capped amortization: per-period descuento del cobro_facturacion según rate,
        pero deja de descontar una vez que el saldo acumulado iguala el monto del anticipo.

        Returns:
            (vector negativo por periodo, índice del periodo donde se completó la amortización
             — None si nunca se llegó al cap o si no hay anticipo).
        """
        out = [ZERO] * self.N
        advance_amt = self.settings.advanceamountnotax or ZERO
        rate = self.settings.advanceamortizationrate or ZERO

        # Sin anticipo o sin tasa → no hay amortización
        if advance_amt <= ZERO or rate <= ZERO:
            return out, None

        cap = -advance_amt  # piso del saldo acumulado (negativo)
        cumulative = ZERO
        fully_amortized_at = None

        for i, cf in enumerate(cobro_facturacion):
            if cumulative <= cap:
                # Ya totalmente amortizado: no más descuentos
                continue
            proposed = -rate * cf  # negativo
            new_cumulative = cumulative + proposed
            if new_cumulative <= cap:
                # Cap: solo descontar lo necesario para cerrar el saldo
                actual = cap - cumulative  # ≤ 0
                out[i] = actual
                cumulative = cap
                if fully_amortized_at is None:
                    fully_amortized_at = i
            else:
                out[i] = proposed
                cumulative = new_cumulative

        return out, fully_amortized_at

    def _single_period_vector(self, amount, period_1indexed):
        """Returns (vector, fuera). If period_1indexed > N the amount falls fuera."""
        out = [ZERO] * self.N
        if not amount:
            return out, ZERO
        p = (period_1indexed or 1) - 1
        amt = Decimal(amount)
        if 0 <= p < self.N:
            out[p] = amt
            return out, ZERO
        return out, amt

    def _compute_devolucion(self, imss, otras):
        """Returns (vector, fuera). If retentionreturnperiod > N the return falls fuera."""
        out = [ZERO] * self.N
        if self.settings.retentionreturnperiod is None:
            return out, ZERO
        amount = -sum(imss) - sum(otras)
        p = self.settings.retentionreturnperiod - 1
        if 0 <= p < self.N:
            out[p] = amount
            return out, ZERO
        return out, amount

    @staticmethod
    def _cumsum(values):
        out = []
        acc = ZERO
        for v in values:
            acc += v
            out.append(acc)
        return out


# =============================================================================
# Breakdown Excel Import / Export Service
# =============================================================================


@dataclass(frozen=True)
class _NormalizedCategory:
    supplytype: int
    category_code: int  # BreakdownCategoryCode value


@dataclass
class _ParsedRow:
    row_num: int
    concepto: str
    categoria: str
    insumo_codigo: str
    insumo_descripcion: str
    unidad: str
    rendimiento: Optional[Decimal]
    precio_unitario: Optional[Decimal]


@dataclass
class _ParsedExcel:
    rows: list  # list[_ParsedRow]
    uploaded_uuid: Optional[str]


class BreakdownExcelService:
    """Round-trip Excel import/export for Unit Cost Breakdowns (CDU)."""

    # Excel CATEGORIA → (SupplyType, BreakdownCategoryCode)
    _CATEGORY_MAP = {
        "MATERIALES":   _NormalizedCategory(supplytype=0, category_code=1),  # Materials
        "MANO_OBRA":    _NormalizedCategory(supplytype=1, category_code=4),  # Labor
        "MAQUINARIA":   _NormalizedCategory(supplytype=2, category_code=3),  # Machinery
        "ACARREOS":     _NormalizedCategory(supplytype=4, category_code=2),  # Hauling
        "SUBCONTRATOS": _NormalizedCategory(supplytype=3, category_code=5),  # Subcontracts
    }

    _REJECTED_CATEGORIES = {"HM", "EPP", "HERRAMIENTA_MENOR", "HERRAMIENTAMENOR", "PPE"}

    @classmethod
    def normalize_category(cls, value: str) -> _NormalizedCategory:
        """Normalize Excel CATEGORIA cell to (supplytype, breakdown category_code).

        Handles case-insensitive comparison, trim, and space/underscore variants.
        Raises ValueError for HM/EPP (auto-generated) or unknown values.
        """
        if not value:
            raise ValueError("Categoría vacía")
        normalized = value.strip().upper().replace(" ", "_")
        if normalized in cls._REJECTED_CATEGORIES:
            raise ValueError("HM/EPP se calculan automáticamente, no incluir en Excel")
        if normalized not in cls._CATEGORY_MAP:
            raise ValueError(f"Categoría no reconocida: {value!r}")
        return cls._CATEGORY_MAP[normalized]

    @staticmethod
    def _build_concept_index(project_id):
        """Return dict {code: conceptid_uuid} for all concepts in the project.

        Single query, no N+1.
        """
        from apps.proyeccion.models import BudgetConcept
        return {
            c.code: c.conceptid
            for c in BudgetConcept.objects.filter(projectid_id=project_id).only("conceptid", "code")
        }

    @staticmethod
    def _match_concept(code_value, concept_index):
        """Match an Excel CONCEPTO cell to a BudgetConcept UUID.

        1. Exact match by `code`
        2. (Future) fallback to FIMP-S<NN>-<idx> derivation if needed
        Returns: UUID or None
        """
        if not code_value:
            return None
        code = str(code_value).strip()
        return concept_index.get(code)

    @staticmethod
    def _build_supply_index():
        """Return dict {code: SupplyCatalogItem} for all catalog items.

        SupplyCatalogItem is global (not per project).
        """
        from apps.proyeccion.models import SupplyCatalogItem
        return {s.code: s for s in SupplyCatalogItem.objects.all()}

    @staticmethod
    def _match_supply(code_value, supply_index):
        """Match Excel INSUMO_CODIGO to a SupplyCatalogItem.

        Case-sensitive (codes are canonical).
        """
        if not code_value:
            return None
        return supply_index.get(str(code_value).strip())

    @staticmethod
    def _parse_excel(file_or_buffer) -> _ParsedExcel:
        """Parse the CDU sheet from an .xlsx file.

        Returns _ParsedExcel with rows and the uploaded UUID from row 2 col 1.
        Raises ValueError if sheet 'CDU' missing or no data rows.
        """
        from decimal import Decimal as D
        from openpyxl import load_workbook
        from zipfile import BadZipFile

        try:
            wb = load_workbook(file_or_buffer, data_only=True)
        except BadZipFile:
            raise ValueError("El archivo no es un .xlsx válido (formato ZIP corrupto o incorrecto)")
        if "CDU" not in wb.sheetnames:
            raise ValueError("El archivo no contiene la hoja 'CDU' esperada")
        ws = wb["CDU"]

        uploaded_uuid = ws.cell(row=2, column=1).value
        if uploaded_uuid is not None:
            uploaded_uuid = str(uploaded_uuid).strip() or None

        rows = []
        last_concepto = ""
        for excel_row in range(4, ws.max_row + 1):
            cells = [ws.cell(row=excel_row, column=c).value for c in range(1, 9)]
            # Skip fully empty rows
            if all(v is None or str(v).strip() == "" for v in cells):
                continue

            concepto_raw = cells[0]
            if concepto_raw is None or str(concepto_raw).strip() == "":
                concepto = last_concepto
            else:
                concepto = str(concepto_raw).strip()
                last_concepto = concepto

            categoria = (str(cells[1]).strip() if cells[1] is not None else "")
            insumo_codigo = (str(cells[2]).strip() if cells[2] is not None else "")
            insumo_descripcion = (str(cells[3]).strip() if cells[3] is not None else "")
            unidad = (str(cells[4]).strip() if cells[4] is not None else "")

            # Skip rows that are only a CONCEPTO marker (visual header)
            if not categoria and not insumo_codigo and not unidad:
                continue

            def to_decimal(v):
                if v is None or str(v).strip() == "":
                    return None
                try:
                    return D(str(v))
                except Exception:
                    return None

            rows.append(_ParsedRow(
                row_num=excel_row,
                concepto=concepto,
                categoria=categoria,
                insumo_codigo=insumo_codigo,
                insumo_descripcion=insumo_descripcion,
                unidad=unidad,
                rendimiento=to_decimal(cells[5]),
                precio_unitario=to_decimal(cells[6]),
            ))

        if not rows:
            raise ValueError("Excel vacío: no se encontraron filas de datos")

        return _ParsedExcel(rows=rows, uploaded_uuid=uploaded_uuid)

    @classmethod
    def analyze(cls, project_id, file, user):
        """Analyze an uploaded CDU Excel file without persisting changes.

        Returns AnalyzeBreakdownsResponseSchema-compatible object.
        """
        from decimal import Decimal as D, ROUND_HALF_UP
        from apps.proyeccion.schemas import (
            AnalyzeBreakdownsResponseSchema,
            BreakdownExcelSummarySchema,
            BreakdownExcelConceptSchema,
            BreakdownExcelLineSchema,
            BreakdownExcelNewSupplySchema,
            BreakdownExcelErrorSchema,
        )

        try:
            parsed = cls._parse_excel(file)
        except ValueError as e:
            return AnalyzeBreakdownsResponseSchema(
                summary=BreakdownExcelSummarySchema(
                    concepts_count=0, lines_count=0, new_supplies_count=0, errors_count=1,
                ),
                concepts=[], new_supplies=[],
                errors=[BreakdownExcelErrorSchema(row=0, message=str(e))],
                project_uuid_match=False,
                uploaded_uuid=None,
            )

        concept_index = cls._build_concept_index(project_id)
        supply_index = cls._build_supply_index()

        concepts_map = {}     # concept_code → dict with lines, etc.
        new_supplies_map = {} # supply_code → new supply data
        errors = []

        for prow in parsed.rows:
            try:
                cat = cls.normalize_category(prow.categoria)
            except ValueError as e:
                errors.append(BreakdownExcelErrorSchema(
                    row=prow.row_num, concept_code=prow.concepto,
                    supply_code=prow.insumo_codigo, message=str(e),
                ))
                continue

            concept_uuid = cls._match_concept(prow.concepto, concept_index)
            if concept_uuid is None:
                errors.append(BreakdownExcelErrorSchema(
                    row=prow.row_num, concept_code=prow.concepto,
                    supply_code=prow.insumo_codigo,
                    message=f"Concepto no encontrado: {prow.concepto}",
                ))
                continue

            if prow.rendimiento is None or prow.rendimiento <= 0:
                errors.append(BreakdownExcelErrorSchema(
                    row=prow.row_num, concept_code=prow.concepto,
                    supply_code=prow.insumo_codigo,
                    message="Rendimiento debe ser > 0",
                ))
                continue

            supply = cls._match_supply(prow.insumo_codigo, supply_index)
            is_new = supply is None
            warnings_for_line = []

            if is_new:
                if not prow.insumo_descripcion:
                    errors.append(BreakdownExcelErrorSchema(
                        row=prow.row_num, concept_code=prow.concepto,
                        supply_code=prow.insumo_codigo,
                        message="Insumo nuevo requiere INSUMO_DESCRIPCION",
                    ))
                    continue
                if not prow.unidad:
                    errors.append(BreakdownExcelErrorSchema(
                        row=prow.row_num, concept_code=prow.concepto,
                        supply_code=prow.insumo_codigo,
                        message="Insumo nuevo requiere UNIDAD",
                    ))
                    continue
                if prow.precio_unitario is None or prow.precio_unitario <= 0:
                    errors.append(BreakdownExcelErrorSchema(
                        row=prow.row_num, concept_code=prow.concepto,
                        supply_code=prow.insumo_codigo,
                        message="Insumo nuevo requiere PRECIO_UNITARIO > 0",
                    ))
                    continue
                effective_unit = prow.unidad
                effective_name = prow.insumo_descripcion
                effective_price = prow.precio_unitario
                if prow.insumo_codigo in new_supplies_map:
                    nsm = new_supplies_map[prow.insumo_codigo]
                    if prow.concepto not in nsm["appears_in_concepts"]:
                        nsm["appears_in_concepts"].append(prow.concepto)
                else:
                    new_supplies_map[prow.insumo_codigo] = {
                        "code": prow.insumo_codigo,
                        "name": effective_name,
                        "unit": effective_unit,
                        "supplytype": cat.supplytype,
                        "reference_price": effective_price,
                        "appears_in_concepts": [prow.concepto],
                    }
            else:
                effective_unit = supply.unit
                effective_name = supply.description
                if prow.precio_unitario is not None and prow.precio_unitario > 0:
                    effective_price = prow.precio_unitario
                elif supply.referenceprice and supply.referenceprice > 0:
                    effective_price = D(str(supply.referenceprice))
                else:
                    errors.append(BreakdownExcelErrorSchema(
                        row=prow.row_num, concept_code=prow.concepto,
                        supply_code=prow.insumo_codigo,
                        message="Insumo sin precio de referencia, especifica PRECIO_UNITARIO",
                    ))
                    continue

            amount = (D('1') * effective_price * prow.rendimiento).quantize(D('0.01'), ROUND_HALF_UP)
            line_dict = {
                "row": prow.row_num,
                "category": prow.categoria.strip().upper().replace(" ", "_"),
                "supply_code": prow.insumo_codigo,
                "supply_name": effective_name,
                "unit": effective_unit,
                "yield_value": prow.rendimiento,
                "unit_price": effective_price,
                "amount": amount,
                "is_new_supply": is_new,
                "warnings": warnings_for_line,
                "_categorycode": cat.category_code,  # internal, not in schema
            }

            if prow.concepto not in concepts_map:
                concepts_map[prow.concepto] = {
                    "code": prow.concepto,
                    "name": "",
                    "lines": [],
                    "_concept_uuid": concept_uuid,
                }
            entry = concepts_map[prow.concepto]

            existing_line = next(
                (l for l in entry["lines"] if l["supply_code"] == line_dict["supply_code"]),
                None,
            )
            if existing_line is not None:
                existing_line["yield_value"] = existing_line["yield_value"] + prow.rendimiento
                existing_line["amount"] = (
                    D('1') * existing_line["unit_price"] * existing_line["yield_value"]
                ).quantize(D('0.01'), ROUND_HALF_UP)
                existing_line["warnings"].append(
                    f"Insumo duplicado en concepto, rendimientos sumados (fila {prow.row_num})"
                )
            else:
                entry["lines"].append(line_dict)

        from apps.proyeccion.models import BudgetConcept
        concept_codes = list(concepts_map.keys())
        name_lookup = {
            c.code: c.description
            for c in BudgetConcept.objects.filter(
                projectid_id=project_id, code__in=concept_codes
            ).only("code", "description")
        }

        out_concepts = []
        total_lines = 0
        for code, entry in concepts_map.items():
            entry["name"] = name_lookup.get(code, "")
            labor_total = sum(
                (D(str(l["amount"])) for l in entry["lines"] if l["_categorycode"] == 4),
                D('0'),
            )
            hm_preview = (D('0.03') * labor_total).quantize(D('0.01'), ROUND_HALF_UP) if labor_total > 0 else D('0.00')
            epp_preview = hm_preview
            lines_total = sum((D(str(l["amount"])) for l in entry["lines"]), D('0'))
            total_preview = (lines_total + hm_preview + epp_preview).quantize(D('0.01'), ROUND_HALF_UP)

            schema_lines = [
                BreakdownExcelLineSchema(**{k: v for k, v in l.items() if not k.startswith("_")})
                for l in entry["lines"]
            ]
            out_concepts.append(BreakdownExcelConceptSchema(
                code=entry["code"],
                name=entry["name"],
                lines=schema_lines,
                hm_preview=hm_preview,
                epp_preview=epp_preview,
                total_preview=total_preview,
            ))
            total_lines += len(schema_lines)

        new_supplies_list = [BreakdownExcelNewSupplySchema(**v) for v in new_supplies_map.values()]

        uuid_match = (
            parsed.uploaded_uuid is not None
            and parsed.uploaded_uuid == str(project_id)
        )

        # Compute CostDistribution rows that will be deleted by cascade.
        # Single aggregated query using related_name="period_distributions"
        # on CostDistribution.breakdownid.
        from django.db.models import Count
        from apps.proyeccion.models import UnitCostBreakdown as _UCB

        concept_uuids = [
            entry["_concept_uuid"] for entry in concepts_map.values()
        ]
        if concept_uuids:
            affected_per_concept = list(
                _UCB.objects
                .filter(conceptid_id__in=concept_uuids)
                .values("conceptid__code")
                .annotate(dist_count=Count("period_distributions"))
                .filter(dist_count__gt=0)
            )
            affected_count = sum(row["dist_count"] for row in affected_per_concept)
            affected_concept_codes = sorted(
                row["conceptid__code"] for row in affected_per_concept
            )
        else:
            affected_count = 0
            affected_concept_codes = []

        return AnalyzeBreakdownsResponseSchema(
            summary=BreakdownExcelSummarySchema(
                concepts_count=len(out_concepts),
                lines_count=total_lines,
                new_supplies_count=len(new_supplies_list),
                errors_count=len(errors),
            ),
            concepts=out_concepts,
            new_supplies=new_supplies_list,
            errors=errors,
            project_uuid_match=uuid_match,
            uploaded_uuid=parsed.uploaded_uuid,
            affected_distributions_count=affected_count,
            affected_concepts_with_distributions=affected_concept_codes,
        )

    @classmethod
    def import_(cls, project_id, payload, user):
        """Persist a previously analyzed import payload.

        Atomic: if any concept fails, the whole transaction rolls back.
        """
        from decimal import Decimal as D, ROUND_HALF_UP
        from django.db import transaction
        from apps.proyeccion.models import (
            BudgetConcept, SupplyCatalogItem, UnitCostBreakdown,
        )
        from apps.proyeccion.schemas import ImportBreakdownsResponseSchema

        # UUID gate
        if (
            payload.uploaded_uuid
            and payload.uploaded_uuid != str(project_id)
            and not payload.override_uuid_mismatch
        ):
            raise ValueError("UUID del proyecto no coincide; marcar override_uuid_mismatch para forzar.")

        with transaction.atomic():
            # 1. Auto-create new supplies (idempotent: get_or_create)
            supplies_created = 0
            for ns in payload.new_supplies:
                _, created = SupplyCatalogItem.objects.get_or_create(
                    code=ns.code,
                    defaults={
                        "description": ns.name,
                        "unit": ns.unit,
                        "supplytype": ns.supplytype,
                        "referenceprice": ns.reference_price,
                    },
                )
                if created:
                    supplies_created += 1

            supply_index = cls._build_supply_index()
            concept_index = cls._build_concept_index(project_id)

            concepts_replaced = 0
            lines_created = 0
            hm_epp_count = 0

            for cdto in payload.concepts:
                if not cdto.lines:
                    raise ValueError(
                        f"Concepto {cdto.code} no tiene líneas en el payload — no se puede vaciar el CDU desde Excel"
                    )
                concept_uuid = concept_index.get(cdto.code)
                if concept_uuid is None:
                    raise ValueError(f"Concepto no encontrado durante import: {cdto.code}")

                # Delete existing breakdowns (HM/EPP included; will be regenerated)
                UnitCostBreakdown.objects.filter(conceptid_id=concept_uuid).delete()

                new_lines = []
                for idx, ldto in enumerate(cdto.lines, start=1):
                    cat = cls.normalize_category(ldto.category)
                    supply = supply_index.get(ldto.supply_code)
                    if supply is None:
                        raise ValueError(
                            f"Insumo no encontrado durante import: {ldto.supply_code} "
                            f"(concepto {cdto.code})"
                        )
                    amount = (D('1') * ldto.unit_price * ldto.yield_value).quantize(
                        D('0.01'), ROUND_HALF_UP
                    )
                    new_lines.append(UnitCostBreakdown(
                        conceptid_id=concept_uuid,
                        categorycode=cat.category_code,
                        linenumber=idx,
                        description=ldto.supply_name or supply.description,
                        unit=ldto.unit or supply.unit,
                        quantity=D('1'),
                        unitprice=ldto.unit_price,
                        yieldvalue=ldto.yield_value,
                        amount=amount,
                        supplyid=supply,
                    ))
                UnitCostBreakdown.objects.bulk_create(new_lines)
                lines_created += len(new_lines)

                hm_created, epp_created = UnitCostBreakdownService.regenerate_hm_epp(
                    concept_uuid, user
                )
                if hm_created and epp_created:
                    hm_epp_count += 1

                UnitCostBreakdownService._recalc_concept(concept_uuid, user)
                concepts_replaced += 1

            # 3. Re-prorate indirect costs across all active concepts.
            # prorate_to_concepts() is a no-op when total_direct == 0 (returns early);
            # it never raises for the empty-data case, so no try/except is needed.
            # Any real exception (DB error, programming bug) must propagate to roll back.
            IndirectCostDetailService.prorate_to_concepts(project_id, user)
            prorate_triggered = True

        return ImportBreakdownsResponseSchema(
            concepts_replaced=concepts_replaced,
            lines_created=lines_created,
            supplies_created=supplies_created,
            hm_epp_regenerated=hm_epp_count,
            prorate_triggered=prorate_triggered,
        )

    # BreakdownCategoryCode integer → Excel CATEGORIA label (excludes HM/EPP)
    _CATEGORY_TO_EXCEL = {
        1: "MATERIALES",     # Materials
        2: "ACARREOS",       # Hauling
        3: "MAQUINARIA",     # Machinery
        4: "MANO_OBRA",      # Labor
        5: "SUBCONTRATOS",   # Subcontracts
    }

    @classmethod
    def export(cls, project_id, user) -> bytes:
        """Export the project's CDU to an .xlsx file (bytes).

        Excludes HM/EPP categories (auto-regenerated on import). For empty
        concepts (no breakdowns yet), emits one placeholder row per concept.
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from apps.proyeccion.models import (
            BudgetConcept, UnitCostBreakdown, EstimationProject,
        )

        project = EstimationProject.objects.filter(pk=project_id).first()
        project_name = getattr(project, "name", "") if project else ""

        wb = Workbook()
        ws = wb.active
        ws.title = "CDU"

        # Row 1: title
        ws.cell(row=1, column=1, value=f"Proyecto: {project_name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)

        # Row 2: UUID (for round-trip identification)
        ws.cell(row=2, column=1, value=str(project_id))
        ws.cell(row=2, column=1).font = Font(italic=True, color="888888")

        # Row 3: column headers
        headers = [
            "CONCEPTO", "CATEGORIA", "INSUMO_CODIGO",
            "INSUMO_DESCRIPCION", "UNIDAD", "RENDIMIENTO",
            "PRECIO_UNITARIO", "IMPORTE",
        ]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")

        ws.freeze_panes = "A4"

        widths = [14, 14, 16, 35, 10, 12, 15, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Fetch concepts ordered alphabetically
        concepts = list(
            BudgetConcept.objects.filter(projectid_id=project_id).order_by("code")
        )

        # Fetch all breakdowns for this project, excluding HM (6) and EPP (7)
        breakdowns_by_concept = {}
        for b in (
            UnitCostBreakdown.objects
            .filter(conceptid__projectid_id=project_id)
            .exclude(categorycode__in=(
                BreakdownCategoryCode.MINOR_TOOLS,
                BreakdownCategoryCode.PPE,
            ))
            .select_related("conceptid", "supplyid")
            .order_by("conceptid__code", "categorycode", "linenumber")
        ):
            breakdowns_by_concept.setdefault(b.conceptid.code, []).append(b)

        excel_row = 4
        for concept in concepts:
            lines = breakdowns_by_concept.get(concept.code, [])
            if not lines:
                # Placeholder row: only CONCEPTO filled (concept has no breakdowns yet)
                ws.cell(row=excel_row, column=1, value=concept.code)
                excel_row += 1
                continue
            for idx, line in enumerate(lines):
                ws.cell(row=excel_row, column=1, value=concept.code if idx == 0 else "")
                ws.cell(row=excel_row, column=2, value=cls._CATEGORY_TO_EXCEL.get(line.categorycode, ""))
                ws.cell(row=excel_row, column=3, value=getattr(line.supplyid, "code", "") if line.supplyid else "")
                ws.cell(row=excel_row, column=4, value=line.description or "")
                ws.cell(row=excel_row, column=5, value=line.unit or "")
                ws.cell(row=excel_row, column=6, value=float(line.yieldvalue))
                ws.cell(row=excel_row, column=7, value=float(line.unitprice))
                ws.cell(row=excel_row, column=8, value=float(line.amount))
                excel_row += 1

        # CATEGORIA dropdown validation
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1='"MATERIALES,MANO_OBRA,MAQUINARIA,ACARREOS,SUBCONTRATOS"',
            allow_blank=True,
        )
        dv.add(f"B4:B{max(excel_row - 1, 4)}")
        ws.add_data_validation(dv)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# =============================================================================
# Indirect Cost Excel Import / Export Service
# =============================================================================


@dataclass
class _ParsedIndirectRow:
    row_num: int
    categoria: str
    codigo: str
    area: str
    descripcion: str
    costo_mensual: Optional[Decimal]
    unidades: Optional[Decimal]
    meses: Optional[Decimal]
    mes_inicio: Optional[int]
    mes_fin: Optional[int]
    lag_pago: Optional[int]


@dataclass
class _ParsedIndirectExcel:
    rows: list
    uploaded_uuid: Optional[str]


class IndirectExcelService:
    """Round-trip Excel import/export for Indirect Cost Details (C1-C8).

    Mirror of BreakdownExcelService for the indirect-costs domain. The import
    strategy is REPLACE: all existing IndirectCostDetail records for the project
    are deleted, then bulk-created from the payload, and prorate_to_concepts is
    triggered to cascade indirect costs to BudgetConcept.indirectunitcost.
    """

    _VALID_CATEGORIES = {f'C{i}' for i in range(1, 9)}
    _SHEET_NAME_CANDIDATES = ("Costos Indirectos", "IndirectCosts", "Indirectos")

    @classmethod
    def _parse_excel(cls, file_or_buffer) -> _ParsedIndirectExcel:
        from decimal import Decimal as D
        from openpyxl import load_workbook
        from zipfile import BadZipFile

        try:
            wb = load_workbook(file_or_buffer, data_only=True)
        except BadZipFile:
            raise ValueError("El archivo no es un .xlsx válido (formato ZIP corrupto o incorrecto)")

        sheet_name = next(
            (s for s in cls._SHEET_NAME_CANDIDATES if s in wb.sheetnames),
            None,
        )
        if sheet_name is None:
            raise ValueError(
                "El archivo no contiene la hoja 'Costos Indirectos' esperada"
            )
        ws = wb[sheet_name]

        uploaded_uuid = ws.cell(row=2, column=1).value
        if uploaded_uuid is not None:
            uploaded_uuid = str(uploaded_uuid).strip() or None

        def to_decimal(v):
            if v is None or str(v).strip() == "":
                return None
            try:
                return D(str(v))
            except Exception:
                return None

        def to_int(v):
            if v is None or str(v).strip() == "":
                return None
            try:
                return int(D(str(v)))
            except Exception:
                return None

        rows = []
        for excel_row in range(4, ws.max_row + 1):
            cells = [ws.cell(row=excel_row, column=c).value for c in range(1, 12)]
            if all(v is None or str(v).strip() == "" for v in cells):
                continue

            categoria = (str(cells[0]).strip().upper() if cells[0] is not None else "")
            codigo = (str(cells[1]).strip() if cells[1] is not None else "")
            area = (str(cells[2]).strip() if cells[2] is not None else "")
            descripcion = (str(cells[3]).strip() if cells[3] is not None else "")

            if not descripcion:
                continue

            rows.append(_ParsedIndirectRow(
                row_num=excel_row,
                categoria=categoria,
                codigo=codigo,
                area=area,
                descripcion=descripcion,
                costo_mensual=to_decimal(cells[4]),
                unidades=to_decimal(cells[5]),
                meses=to_decimal(cells[6]),
                mes_inicio=to_int(cells[7]),
                mes_fin=to_int(cells[8]),
                lag_pago=to_int(cells[9]),
            ))

        if not rows:
            raise ValueError("Excel vacío: no se encontraron filas de datos")

        return _ParsedIndirectExcel(rows=rows, uploaded_uuid=uploaded_uuid)

    @classmethod
    def analyze(cls, project_id, file, user):
        from decimal import Decimal as D, ROUND_HALF_UP
        from apps.proyeccion.schemas import (
            AnalyzeIndirectsResponseSchema,
            IndirectExcelSummarySchema,
            IndirectExcelLineSchema,
            IndirectExcelErrorSchema,
        )

        try:
            parsed = cls._parse_excel(file)
        except ValueError as e:
            return AnalyzeIndirectsResponseSchema(
                summary=IndirectExcelSummarySchema(
                    lines_count=0,
                    total_amount=Decimal('0'),
                    errors_count=1,
                ),
                lines=[],
                errors=[IndirectExcelErrorSchema(row=0, message=str(e))],
                project_uuid_match=False,
                uploaded_uuid=None,
            )

        errors = []
        valid_lines = []
        total_amount = D('0')

        for prow in parsed.rows:
            if not prow.categoria:
                errors.append(IndirectExcelErrorSchema(
                    row=prow.row_num,
                    description=prow.descripcion,
                    message="Categoría vacía",
                ))
                continue
            if prow.categoria not in cls._VALID_CATEGORIES:
                errors.append(IndirectExcelErrorSchema(
                    row=prow.row_num,
                    category=prow.categoria,
                    description=prow.descripcion,
                    message=f"Categoría no reconocida: {prow.categoria} (esperado C1-C8)",
                ))
                continue

            costo = prow.costo_mensual if prow.costo_mensual is not None else D('0')
            unidades = prow.unidades if prow.unidades is not None else D('1')
            meses = prow.meses if prow.meses is not None else D('0')

            if costo < 0 or unidades < 0 or meses < 0:
                errors.append(IndirectExcelErrorSchema(
                    row=prow.row_num,
                    category=prow.categoria,
                    description=prow.descripcion,
                    message="Valores numéricos no pueden ser negativos",
                ))
                continue

            amount = (costo * unidades * meses).quantize(D('0.01'), ROUND_HALF_UP)
            total_amount += amount

            valid_lines.append(IndirectExcelLineSchema(
                row=prow.row_num,
                category=prow.categoria,
                code=prow.codigo,
                area=prow.area,
                description=prow.descripcion,
                monthly_cost=costo,
                units=unidades,
                months=meses,
                start_month=prow.mes_inicio,
                end_month=prow.mes_fin,
                payment_lag=prow.lag_pago,
                amount=amount,
            ))

        uuid_match = (
            parsed.uploaded_uuid is not None
            and parsed.uploaded_uuid == str(project_id)
        )

        return AnalyzeIndirectsResponseSchema(
            summary=IndirectExcelSummarySchema(
                lines_count=len(valid_lines),
                total_amount=total_amount,
                errors_count=len(errors),
            ),
            lines=valid_lines,
            errors=errors,
            project_uuid_match=uuid_match,
            uploaded_uuid=parsed.uploaded_uuid,
        )

    @classmethod
    def import_(cls, project_id, payload, user):
        """Replace all IndirectCostDetail records for the project from payload.

        Atomic: any failure rolls back. Triggers prorate_to_concepts after the
        bulk_create so BudgetConcept.indirectunitcost / unitprice / totalamount
        reflect the new indirect totals.
        """
        from django.db import transaction
        from apps.proyeccion.schemas import ImportIndirectsResponseSchema

        if (
            payload.uploaded_uuid
            and payload.uploaded_uuid != str(project_id)
            and not payload.override_uuid_mismatch
        ):
            raise ValueError(
                "UUID del proyecto no coincide; marcar override_uuid_mismatch para forzar."
            )

        if not payload.lines:
            raise ValueError(
                "Payload vacío: no se puede importar 0 líneas (usa la UI para eliminar todas)"
            )

        with transaction.atomic():
            deleted_count = IndirectCostDetail.objects.filter(
                projectid_id=project_id,
            ).count()
            IndirectCostDetail.objects.filter(projectid_id=project_id).delete()

            cat_counters = {f'C{i}': 0 for i in range(1, 9)}
            new_details = []
            for ldto in payload.lines:
                cat_counters[ldto.category] = cat_counters.get(ldto.category, 0) + 1
                line_num = cat_counters[ldto.category]
                new_details.append(IndirectCostDetail(
                    projectid_id=project_id,
                    categorycode=ldto.category,
                    linenumber=line_num,
                    imputationcode=(ldto.code or '')[:10],
                    area=(ldto.area or '')[:100],
                    description=ldto.description[:500],
                    monthlycost=ldto.monthly_cost,
                    units=ldto.units,
                    months=ldto.months,
                    startmonth=ldto.start_month,
                    endmonth=ldto.end_month,
                    paymentlagperiods=ldto.payment_lag,
                    amount=ldto.amount,
                    statecode=0,
                    createdby=user,
                    modifiedby=user,
                ))
            IndirectCostDetail.objects.bulk_create(new_details)

            IndirectCostDetailService.prorate_to_concepts(project_id, user)
            prorate_triggered = True

        return ImportIndirectsResponseSchema(
            details_deleted=deleted_count,
            details_created=len(new_details),
            prorate_triggered=prorate_triggered,
        )

    @classmethod
    def export(cls, project_id, user) -> bytes:
        """Export the project's indirect costs to an .xlsx file (bytes)."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        project = EstimationProject.objects.filter(pk=project_id).first()
        project_name = getattr(project, "name", "") if project else ""

        wb = Workbook()
        ws = wb.active
        ws.title = "Costos Indirectos"

        ws.cell(row=1, column=1, value=f"Proyecto: {project_name}").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value=str(project_id)).font = Font(italic=True, color="888888")

        headers = [
            "CATEGORIA", "CODIGO", "AREA", "DESCRIPCION",
            "COSTO_MENSUAL", "UNIDADES", "MESES",
            "MES_INICIO", "MES_FIN", "LAG_PAGO", "IMPORTE",
        ]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        ws.freeze_panes = "A4"
        widths = [12, 10, 18, 50, 14, 10, 10, 11, 11, 11, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        details = list(
            IndirectCostDetail.objects
            .filter(projectid_id=project_id, statecode=0)
            .order_by('categorycode', 'linenumber')
        )

        excel_row = 4
        for d in details:
            ws.cell(row=excel_row, column=1, value=d.categorycode)
            ws.cell(row=excel_row, column=2, value=d.imputationcode or "")
            ws.cell(row=excel_row, column=3, value=d.area or "")
            ws.cell(row=excel_row, column=4, value=d.description)
            ws.cell(row=excel_row, column=5, value=float(d.monthlycost))
            ws.cell(row=excel_row, column=6, value=float(d.units))
            ws.cell(row=excel_row, column=7, value=float(d.months))
            ws.cell(row=excel_row, column=8, value=d.startmonth)
            ws.cell(row=excel_row, column=9, value=d.endmonth)
            ws.cell(row=excel_row, column=10, value=d.paymentlagperiods)
            ws.cell(row=excel_row, column=11, value=float(d.amount))
            excel_row += 1

        dv = DataValidation(
            type="list",
            formula1='"C1,C2,C3,C4,C5,C6,C7,C8"',
            allow_blank=False,
        )
        dv.add(f"A4:A{max(excel_row - 1, 4)}")
        ws.add_data_validation(dv)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# =============================================================================
# Concept Excel Export / Import Service (8-column round-trip format)
# =============================================================================

class ConceptExcelService:
    """Export all project concepts to Excel and import new ones from Excel."""

    HEADERS = [
        'FAMILIA', 'COD.FAM', 'SUBFAMILIA', 'COD.SUB',
        'CODIGO', 'DESCRIPCION COMPLETA', 'UNIDAD', 'CANTIDAD',
    ]
    EXAMPLE_ROWS = [
        ('GABINETE', 'GAB', 'Proyecto Ejecutivo', 'GAB-01', 'A1',
         'Elaboracion de proyecto ejecutivo acorde a las necesidades del cliente', 'EST', 1),
        ('GABINETE', 'GAB', 'Proyecto Ejecutivo', 'GAB-01', 'A2',
         'Levantamiento topografico. Incluye brigada de topografia. P.U.O.T.', 'M2', 1400),
        ('PRELIMINARES', 'PRE', 'Movimiento de Tierras', 'PRE-01', 'B1',
         'Deshierbe, desmonte y despalme. Incluye mano de obra y herramienta menor.', 'M2', 5000),
    ]

    @classmethod
    def export(cls, project_id, user) -> bytes:
        """Return an .xlsx workbook with all active families/subfamilies and their concepts.
        Subfamilies without concepts appear as a single blank row so the user can fill them in."""
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        project = EstimationProject.objects.get(estimationprojectid=project_id)

        # Build a structured list: one entry per (subfamily, concept_or_None)
        families = (
            ConceptFamily.objects.filter(projectid=project, statecode=0)
            .prefetch_related(
                'subfamilies',
                'subfamilies__concepts',
            )
            .order_by('sortorder', 'code')
        )

        # Collect rows: (fam_name, fam_code, sf_name, sf_code, concept_or_None)
        data_rows = []
        for fam in families:
            for sf in fam.subfamilies.filter(statecode=0).order_by('sortorder', 'code'):
                active_concepts = list(
                    sf.concepts.filter(statecode=0).order_by('sequencenumber', 'code')
                )
                if active_concepts:
                    for concept in active_concepts:
                        data_rows.append((fam, sf, concept))
                else:
                    # Subfamily has no concepts yet — include a blank placeholder row
                    data_rows.append((fam, sf, None))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Conceptos'

        # Row 1: project name
        ws.cell(row=1, column=1, value=project.name)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)

        # Row 2: project UUID (round-trip marker)
        ws.cell(row=2, column=1, value=str(project.estimationprojectid))
        ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='888888')

        # Row 3: headers
        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(bold=True, color='FFFFFF')
        for col, header in enumerate(cls.HEADERS, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.freeze_panes = 'A4'

        # Column widths: A=20 B=12 C=25 D=12 E=12 F=55 G=10 H=12
        for col, width in enumerate([20, 12, 25, 12, 12, 55, 10, 12], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        placeholder_font = Font(size=10, color='AAAAAA', italic=True)

        if not data_rows:
            # No families at all — show example rows so the user has a template
            gray_font = Font(size=10, color='888888', italic=True)
            for row_offset, ex in enumerate(cls.EXAMPLE_ROWS, 4):
                for col, val in enumerate(ex, 1):
                    cell = ws.cell(row=row_offset, column=col, value=val)
                    cell.font = gray_font
        else:
            for row_offset, (fam, sf, concept) in enumerate(data_rows, 4):
                ws.cell(row=row_offset, column=1, value=fam.name)
                ws.cell(row=row_offset, column=2, value=fam.code)
                ws.cell(row=row_offset, column=3, value=sf.name)
                ws.cell(row=row_offset, column=4, value=sf.code)
                if concept is not None:
                    ws.cell(row=row_offset, column=5, value=concept.code)
                    ws.cell(row=row_offset, column=6, value=concept.description)
                    ws.cell(row=row_offset, column=7, value=concept.unit)
                    ws.cell(row=row_offset, column=8, value=float(concept.quantity))
                else:
                    # Blank placeholder so the user knows to fill in this subfamily
                    for col in range(5, 9):
                        ws.cell(row=row_offset, column=col).font = placeholder_font

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @classmethod
    def analyze(cls, project_id, file, user) -> dict:
        """Parse uploaded 8-column Excel, classify each row as new/skip/error."""
        import openpyxl

        project = EstimationProject.objects.get(estimationprojectid=project_id)

        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # Locate header row: column 4 must contain "COD" and "SUB"
        header_row = None
        for row_idx in range(1, min(11, ws.max_row + 1)):
            val = str(ws.cell(row=row_idx, column=4).value or '').strip().upper()
            if 'COD' in val and 'SUB' in val:
                header_row = row_idx
                break

        if header_row is None:
            from core.exceptions import ValidationError
            raise ValidationError(
                'No se encontro la fila de encabezados. '
                'La columna 4 debe contener "COD.SUB".'
            )

        # Build concept lookup: code.upper() → BudgetConcept (all statecodes)
        # Keying by concept code lets us detect changes for the 'update' status.
        existing_concepts = {
            c.code.strip().upper(): c
            for c in BudgetConcept.objects.filter(projectid=project)
        }

        rows = []
        prev = {'familia': '', 'cod_fam': '', 'subfamilia': '', 'cod_sub': ''}

        for row_idx in range(header_row + 1, ws.max_row + 1):
            familia = str(ws.cell(row=row_idx, column=1).value or '').strip()
            cod_fam = str(ws.cell(row=row_idx, column=2).value or '').strip()
            subfamilia = str(ws.cell(row=row_idx, column=3).value or '').strip()
            cod_sub = str(ws.cell(row=row_idx, column=4).value or '').strip()
            codigo = str(ws.cell(row=row_idx, column=5).value or '').strip()
            description = str(ws.cell(row=row_idx, column=6).value or '').strip()
            unit = str(ws.cell(row=row_idx, column=7).value or '').strip()
            qty_raw = ws.cell(row=row_idx, column=8).value

            # Skip completely empty rows
            if not description and not codigo:
                continue

            # Inherit grouping columns from previous row when blank
            familia = familia or prev['familia']
            cod_fam = cod_fam or prev['cod_fam']
            subfamilia = subfamilia or prev['subfamilia']
            cod_sub = cod_sub or prev['cod_sub']
            prev = {
                'familia': familia,
                'cod_fam': cod_fam,
                'subfamilia': subfamilia,
                'cod_sub': cod_sub,
            }

            try:
                quantity = float(qty_raw) if qty_raw is not None else 0.0
            except (ValueError, TypeError):
                quantity = 0.0

            old_description = old_unit = old_quantity = None

            if not cod_sub or not cod_fam:
                status = 'error'
                error_msg = 'Fila sin COD.SUB o COD.FAM — no se puede determinar la subfamilia'
            else:
                existing_c = existing_concepts.get(codigo.upper()) if codigo else None
                if existing_c is None:
                    status = 'new'
                    error_msg = None
                else:
                    old_description = existing_c.description
                    old_unit = existing_c.unit
                    old_quantity = float(existing_c.quantity)
                    needs_update = (
                        existing_c.statecode != 0
                        or existing_c.description.strip() != description
                        or existing_c.unit.strip() != unit
                        or abs(float(existing_c.quantity) - quantity) > 0.0001
                    )
                    status = 'update' if needs_update else 'skip'
                    error_msg = None

            rows.append({
                'row': row_idx,
                'familia': familia,
                'cod_fam': cod_fam,
                'subfamilia': subfamilia,
                'cod_sub': cod_sub,
                'codigo': codigo,
                'description': description,
                'unit': unit,
                'quantity': quantity,
                'status': status,
                'error_msg': error_msg,
                'old_description': old_description,
                'old_unit': old_unit,
                'old_quantity': old_quantity,
            })

        wb.close()

        if not rows:
            from core.exceptions import ValidationError
            raise ValidationError('El archivo no contiene filas de datos.')

        summary = {
            'new': sum(1 for r in rows if r['status'] == 'new'),
            'update': sum(1 for r in rows if r['status'] == 'update'),
            'skip': sum(1 for r in rows if r['status'] == 'skip'),
            'error': sum(1 for r in rows if r['status'] == 'error'),
            'total': len(rows),
        }

        return {'summary': summary, 'rows': rows}

    @classmethod
    @transaction.atomic
    def import_(cls, project_id, payload, user) -> dict:
        """Create or update BudgetConcept rows. Families/subfamilies are auto-created
        (or reactivated) as needed. Never deletes existing concepts."""
        from django.db.models import Max

        project = EstimationProject.objects.get(estimationprojectid=project_id)

        # Subfamily lookup keyed by (cod_fam.upper(), cod_sub.upper()) so that
        # 'sub-1' under family 'f1' and 'sub-1' under family 'f2' are distinct.
        existing_subfamilies: dict[tuple, object] = {}
        for sf in ConceptSubfamily.objects.filter(projectid=project).select_related('familyid'):
            key = (sf.familyid.code.strip().upper(), sf.code.strip().upper())
            existing_subfamilies[key] = sf

        existing_families = {
            fam.code.strip().upper(): fam
            for fam in ConceptFamily.objects.filter(projectid=project)
        }

        created = 0
        updated = 0
        skipped = 0
        auto_code_counter: dict[str, int] = {}  # (cod_fam, cod_sub) key → last suffix
        seq_num_cache: dict[str, int] = {}       # sf PK str → last sequencenumber

        for item in payload.items:
            cod_sub_upper = item.cod_sub.strip().upper()
            cod_fam_upper = item.cod_fam.strip().upper()
            sf_lookup_key = (cod_fam_upper, cod_sub_upper)

            # ── Handle 'update' items ──────────────────────────────────────────
            if item.status == 'update':
                concept = BudgetConcept.objects.filter(
                    projectid=project, code=item.codigo.strip()
                ).first()
                if concept is None:
                    skipped += 1
                    continue
                concept.description = item.description
                concept.unit = item.unit
                concept.quantity = Decimal(str(item.quantity))
                concept.statecode = 0  # reactivate if it was soft-deleted
                concept.modifiedby = user
                concept.save(update_fields=[
                    'description', 'unit', 'quantity', 'statecode', 'modifiedby', 'modifiedon',
                ])
                # Also reactivate subfamily and family chains
                sf_of_concept = concept.subfamilyid
                if sf_of_concept.statecode != 0:
                    sf_of_concept.statecode = 0
                    sf_of_concept.save(update_fields=['statecode'])
                fam_of_concept = sf_of_concept.familyid
                if fam_of_concept.statecode != 0:
                    fam_of_concept.statecode = 0
                    fam_of_concept.save(update_fields=['statecode'])
                updated += 1
                continue

            # ── Handle 'new' items ────────────────────────────────────────────
            sf = existing_subfamilies.get(sf_lookup_key)
            if sf is None:
                # Ensure family exists (create or reactivate)
                fam = existing_families.get(cod_fam_upper)
                if fam is None:
                    fam_sort = (
                        ConceptFamily.objects.filter(projectid=project)
                        .aggregate(m=Max('sortorder'))['m'] or 0
                    ) + 1
                    fam = ConceptFamily.objects.create(
                        projectid=project,
                        name=item.familia.strip(),
                        code=item.cod_fam.strip(),
                        sortorder=fam_sort,
                        statecode=0,
                    )
                    existing_families[cod_fam_upper] = fam
                elif fam.statecode != 0:
                    fam.statecode = 0
                    fam.save(update_fields=['statecode'])

                # Create subfamily (or handle edge-case where it exists but inactive)
                sf_sort = (
                    ConceptSubfamily.objects.filter(projectid=project)
                    .aggregate(m=Max('sortorder'))['m'] or 0
                ) + 1
                sf = ConceptSubfamily.objects.create(
                    projectid=project,
                    familyid=fam,
                    name=item.subfamilia.strip(),
                    code=item.cod_sub.strip(),
                    sortorder=sf_sort,
                    statecode=0,
                )
                existing_subfamilies[sf_lookup_key] = sf
            elif sf.statecode != 0:
                sf.statecode = 0
                sf.save(update_fields=['statecode'])
                fam = sf.familyid
                if fam.statecode != 0:
                    fam.statecode = 0
                    fam.save(update_fields=['statecode'])

            code = item.codigo.strip()
            if not code:
                n = auto_code_counter.get(sf_lookup_key, 0) + 1
                auto_code_counter[sf_lookup_key] = n
                code = f'{item.cod_sub}-{n:02d}'

            # Skip if concept already exists (unique_together covers all statecodes)
            if BudgetConcept.objects.filter(projectid=project, code=code).exists():
                skipped += 1
                continue

            sf_key = str(sf.subfamilyid)
            if sf_key not in seq_num_cache:
                seq_num_cache[sf_key] = (
                    BudgetConcept.objects.filter(subfamilyid=sf)
                    .aggregate(m=Max('sequencenumber'))['m'] or 0
                )
            seq_num_cache[sf_key] += 1

            BudgetConcept.objects.create(
                projectid=project,
                subfamilyid=sf,
                code=code,
                sequencenumber=seq_num_cache[sf_key],
                description=item.description,
                unit=item.unit,
                quantity=Decimal(str(item.quantity)),
                directunitcost=Decimal('0'),
                indirectunitcost=Decimal('0'),
                utilityunitcost=Decimal('0'),
                unitprice=Decimal('0'),
                totalamount=Decimal('0'),
                breakdownmethod=0,
                isprintable=True,
                statecode=0,
                createdby=user,
                modifiedby=user,
            )
            created += 1

        return {'created': created, 'updated': updated, 'skipped': skipped}
