"""Seed indirect cost forecast data from the SYM KABAT Excel into ImputationCodeBudget."""

import re
from decimal import Decimal

from django.core.management.base import BaseCommand

# Monkey-patch openpyxl to handle encoding issues in defined names
import openpyxl.reader.workbook as wb_module

original_assign = wb_module.WorkbookParser.assign_names


def safe_assign(self):
    try:
        original_assign(self)
    except ValueError:
        pass


wb_module.WorkbookParser.assign_names = safe_assign

import openpyxl  # noqa: E402

PROJECT_ID = "7e642126-d0ef-46c8-af1e-cbc62909cc59"
EXCEL_PATH = "C:/TestAI/erp_project/erp_backend/002. Control y Seguimiento (SYM KABAT).xlsx"
SHEET_NAME = "Indirect Costs"
HEADER_ROW = 8

# Previsión (forecast) section
PREV_COL_START = 11   # K
PREV_COL_END = 58     # BF

# Real (actual) section
REAL_COL_START = 63    # BK
REAL_COL_END = 110     # DF

# Master data columns
CODE_COL = 3           # C - Imputation code
MONTHLY_COST_COL = 6   # F - Costo mensual
UNITS_COL = 7          # G - Unidades

CODE_PATTERN = re.compile(r"^C\d+-\d+$")


class Command(BaseCommand):
    help = "Seed indirect cost forecast data from Excel for SYM KABAT project"

    def handle(self, *args, **options):
        from apps.budgets.models import ImputationCode, ImputationCodeBudget, ImputationPeriod

        # 1. Load Excel
        self.stdout.write("Loading Excel workbook...")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb[SHEET_NAME]

        # 2. Read period headers from PREVISIÓN section (row 8, cols K-BF)
        prev_headers = {}  # col -> label
        for col in range(PREV_COL_START, PREV_COL_END + 1):
            val = ws.cell(row=HEADER_ROW, column=col).value
            if val and isinstance(val, str):
                prev_headers[col] = val.strip()
        self.stdout.write(f"Found {len(prev_headers)} previsión period headers")

        # 3. Read period headers from REAL section (row 8, cols BK-DF)
        real_headers = {}  # col -> label
        for col in range(REAL_COL_START, REAL_COL_END + 1):
            val = ws.cell(row=HEADER_ROW, column=col).value
            if val and isinstance(val, str):
                real_headers[col] = val.strip()
        self.stdout.write(f"Found {len(real_headers)} real period headers")

        # 4. Build lookup: period label -> ImputationPeriod
        periods = {
            p.label: p
            for p in ImputationPeriod.objects.filter(projectid=PROJECT_ID)
        }
        self.stdout.write(f"Found {len(periods)} periods in DB")

        # 5. Build lookup: code -> ImputationCode (indirect only)
        codes = {
            ic.code: ic
            for ic in ImputationCode.objects.filter(
                projectid=PROJECT_ID,
                costtype=1,  # CostTypeCode.INDIRECT
            )
        }
        self.stdout.write(f"Found {len(codes)} indirect imputation codes in DB")

        # 6. Iterate rows, create/update budget lines
        created = 0
        updated = 0
        skipped_codes = set()

        for row in range(HEADER_ROW + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=CODE_COL).value
            if not cell_val or not isinstance(cell_val, str):
                continue
            code_str = cell_val.strip()
            if not CODE_PATTERN.match(code_str):
                continue

            imp_code = codes.get(code_str)
            if not imp_code:
                skipped_codes.add(code_str)
                continue

            # Read master data for conversion
            monthly_cost = ws.cell(row=row, column=MONTHLY_COST_COL).value
            units = ws.cell(row=row, column=UNITS_COL).value
            try:
                monthly_cost = float(monthly_cost) if monthly_cost else 0.0
            except (ValueError, TypeError):
                monthly_cost = 0.0
            try:
                units = float(units) if units else 0.0
            except (ValueError, TypeError):
                units = 0.0

            # Collect all period data for this code
            period_data = {}  # label -> {planned, actual}

            # 6a. Read PREVISIÓN coefficients and convert to amounts
            for col, label in prev_headers.items():
                cell = ws.cell(row=row, column=col).value
                if cell is None:
                    continue
                try:
                    coef = float(cell)
                except (ValueError, TypeError):
                    continue
                if coef <= 0:
                    continue
                planned = Decimal(str(round(coef * monthly_cost * units, 2)))
                if label not in period_data:
                    period_data[label] = {"planned": Decimal("0"), "actual": Decimal("0")}
                period_data[label]["planned"] = planned

            # 6b. Read REAL amounts (absolute values)
            for col, label in real_headers.items():
                cell = ws.cell(row=row, column=col).value
                if cell is None:
                    continue
                try:
                    actual = float(cell)
                except (ValueError, TypeError):
                    continue
                if actual <= 0:
                    continue
                if label not in period_data:
                    period_data[label] = {"planned": Decimal("0"), "actual": Decimal("0")}
                period_data[label]["actual"] = Decimal(str(round(actual, 2)))

            # 6c. Create/update ImputationCodeBudget records
            for label, amounts in period_data.items():
                period = periods.get(label)
                obj, was_created = ImputationCodeBudget.objects.update_or_create(
                    imputationcodeid=imp_code,
                    periodlabel=label,
                    defaults={
                        "plannedamount": amounts["planned"],
                        "actualamount": amounts["actual"],
                        "periodid": period,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Indirect forecast seeding done: {created} created, {updated} updated"
            )
        )
        if skipped_codes:
            self.stdout.write(
                self.style.WARNING(
                    f"Skipped {len(skipped_codes)} codes not in DB: {sorted(skipped_codes)}"
                )
            )

        # 7. Summary
        indirect_budget_lines = ImputationCodeBudget.objects.filter(
            imputationcodeid__projectid=PROJECT_ID,
            imputationcodeid__costtype=1,
        )
        total = indirect_budget_lines.count()
        with_planned = indirect_budget_lines.filter(plannedamount__gt=0).count()
        with_actual = indirect_budget_lines.filter(actualamount__gt=0).count()
        self.stdout.write(f"\nIndirect Costs Summary:")
        self.stdout.write(f"  Total budget lines: {total}")
        self.stdout.write(f"  With planned > 0:   {with_planned}")
        self.stdout.write(f"  With actual > 0:    {with_actual}")
