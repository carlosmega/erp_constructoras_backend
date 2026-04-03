"""Seed forecast (planned) data from the SYM KABAT Excel into ImputationCodeBudget."""

import re
from decimal import Decimal

from django.core.management.base import BaseCommand

# Monkey-patch openpyxl to handle encoding issues in defined names
import openpyxl.reader.workbook as wb_module

original_assign = wb_module.WorkbookParser.assign_names


def safe_assign(self):
    try:
        original_assign(self)
    except ValueError:
        pass


wb_module.WorkbookParser.assign_names = safe_assign

import openpyxl  # noqa: E402

PROJECT_ID = "7e642126-d0ef-46c8-af1e-cbc62909cc59"
EXCEL_PATH = "C:/TestAI/erp_project/erp_backend/002. Control y Seguimiento (SYM KABAT).xlsx"
SHEET_NAME = "Costo Directo"
HEADER_ROW = 9
COL_START = 12  # L
COL_END = 59    # BG
CODE_COL = 4    # D
CODE_PATTERN = re.compile(r"^[A-Z]{3}-[A-Z]\d+-\d+$")


class Command(BaseCommand):
    help = "Seed forecast data from Excel and compute actuals for SYM KABAT project"

    def handle(self, *args, **options):
        from apps.budgets.models import ImputationCode, ImputationCodeBudget, ImputationPeriod
        from apps.budgets.services import BudgetLineService

        # 1. Load Excel
        self.stdout.write("Loading Excel workbook...")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb[SHEET_NAME]

        # 2. Read period headers from row 9
        headers = {}  # col -> label
        for col in range(COL_START, COL_END + 1):
            val = ws.cell(row=HEADER_ROW, column=col).value
            if val and isinstance(val, str):
                headers[col] = val.strip()
        self.stdout.write(f"Found {len(headers)} period headers")

        # 3. Build lookup: period label -> ImputationPeriod
        periods = {
            p.label: p
            for p in ImputationPeriod.objects.filter(projectid=PROJECT_ID)
        }
        self.stdout.write(f"Found {len(periods)} periods in DB")

        # 4. Build lookup: code -> ImputationCode
        codes = {
            ic.code: ic
            for ic in ImputationCode.objects.filter(projectid=PROJECT_ID)
        }
        self.stdout.write(f"Found {len(codes)} imputation codes in DB")

        # 5. Iterate rows, create budget lines
        created = 0
        updated = 0
        skipped_codes = set()

        for row in range(HEADER_ROW + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=CODE_COL).value
            if not cell_val or not isinstance(cell_val, str):
                continue
            code_str = cell_val.strip()
            if not CODE_PATTERN.match(code_str):
                continue

            imp_code = codes.get(code_str)
            if not imp_code:
                skipped_codes.add(code_str)
                continue

            total_budget = imp_code.totalbudget or Decimal("0")

            for col, label in headers.items():
                cell = ws.cell(row=row, column=col).value
                if cell is None:
                    continue
                try:
                    pct = float(cell)
                except (ValueError, TypeError):
                    continue
                if pct <= 0:
                    continue

                planned = Decimal(str(round(pct * float(total_budget), 2)))
                period = periods.get(label)

                obj, was_created = ImputationCodeBudget.objects.update_or_create(
                    imputationcodeid=imp_code,
                    periodlabel=label,
                    defaults={
                        "plannedamount": planned,
                        "periodid": period,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Forecast seeding done: {created} created, {updated} updated"
            )
        )
        if skipped_codes:
            self.stdout.write(
                self.style.WARNING(
                    f"Skipped {len(skipped_codes)} codes not in DB: {sorted(skipped_codes)}"
                )
            )

        # 6. Compute actuals from classified expenses
        self.stdout.write("Computing actuals from classified expenses...")
        actuals_count = BudgetLineService.compute_actuals(PROJECT_ID)
        self.stdout.write(
            self.style.SUCCESS(f"Actuals computed: {actuals_count} budget lines updated")
        )

        # 7. Summary
        total = ImputationCodeBudget.objects.count()
        with_planned = ImputationCodeBudget.objects.filter(plannedamount__gt=0).count()
        with_actual = ImputationCodeBudget.objects.filter(actualamount__gt=0).count()
        self.stdout.write(f"\nSummary:")
        self.stdout.write(f"  Total budget lines: {total}")
        self.stdout.write(f"  With planned > 0:   {with_planned}")
        self.stdout.write(f"  With actual > 0:    {with_actual}")
